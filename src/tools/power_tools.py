"""
Power Tools - MCP tools for analyzing power measurement data (PACS & FlexLogger formats).

Session-safe: all mutable state is keyed by ctx.session_id so multiple
concurrent users never collide.

MCP Tools exposed:
  parse_power_config             - Parse config file (XML / CSV / pickle) for rail mappings
  analyze_power_summary          - Analyze summary file — primary tool (95% of queries)
  analyze_power_traces           - Time-series trace analysis for a time window
  load_power_csv                 - Load arbitrary CSV into per-session storage
  load_power_json                - Load arbitrary JSON into per-session storage
  analyze_power_dataframe        - LLM-driven pandas query on a loaded DataFrame
  detect_power_rail_config       - Generate PowerRailConfig.txt from folder (Step 1)
  process_summary_rails_to_json  - Extract target rails from summaries → JSON (Step 2)
  create_power_comparison_matrix - Build cross-run comparison Excel/CSV/Markdown (Step 3)

All other functions are internal helpers (no @mcp.tool decorator).
"""

import os
import sys
import json
import pickle
import shutil
import logging
import tempfile
import traceback
import uuid
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Annotated, Dict, Any, List, Optional

import csv
import re as _re_module
import pandas as pd
from pydantic import Field
from fastmcp import Context
from concurrent.futures import ThreadPoolExecutor

from app import mcp
from utils.decorators import embed_if_large, async_tool

# ---------------------------------------------------------------------------
# Paths  (same layout as socwatch_tools / pnp_tools)
# ---------------------------------------------------------------------------
_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))          # src/tools/
_SRC_DIR = os.path.dirname(_CURRENT_DIR)                           # src/
_PROJECT_ROOT = os.path.dirname(_SRC_DIR)                          # fastmcp-server-template/
_WORKSPACE_ROOT = os.path.dirname(os.path.dirname(_PROJECT_ROOT))  # autobots_sdk_20251009_010150/
_PNP_AGENTS_PATH = os.path.join(_WORKSPACE_ROOT, "new_v2_agents", "PnP_agents")
_SDK_PATH = os.path.join(
    _WORKSPACE_ROOT,
    "applications.services.design-system.autobots.autobots-sdk_new_version",
)

# Ensure SDK is importable (for LLM wrapper used by XML parser)
if _SDK_PATH not in sys.path:
    sys.path.insert(0, _SDK_PATH)
if _PNP_AGENTS_PATH not in sys.path:
    sys.path.insert(0, _PNP_AGENTS_PATH)

# Environment variables expected by the SDK
os.environ.setdefault("AUTOBOTS_SDK_TOOL_PATH", _SDK_PATH)
os.environ.setdefault("AUTOBOTS_CONFIG_PATH", os.path.join(_WORKSPACE_ROOT, "crt"))

logging.basicConfig(level=logging.DEBUG, filename="debug_power.log", filemode="w")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Staging / fallback copy — used when source network files can't be read
# ---------------------------------------------------------------------------
_DEFAULT_STAGING_ROOT = r"\\BAPNPSPEED03.gar.corp.intel.com\share\astra_demo_data"
_STAGING_COPY_EXTENSIONS = {".csv", ".txt", ".xlsx", ".json"}

# ---------------------------------------------------------------------------
# Per-session state store  (replaces module-level globals)
# ---------------------------------------------------------------------------
# Dict[session_id → session_state_dict]
_SESSION_STORE: Dict[str, Dict[str, Any]] = {}


def _get_session(ctx: Context) -> Dict[str, Any]:
    """Return (or create) the isolated state bucket for this session."""
    sid = getattr(ctx, "session_id", None) or "default"
    if sid not in _SESSION_STORE:
        _SESSION_STORE[sid] = {
            "dataframes": {},          # name → pickle_path
            "power_rail_context": {
                "soc_rails": {},
                "platform_rails": {},
                "rail_formulas": {},
                "updated": False,
            },
            "debug_logs": [],
        }
    return _SESSION_STORE[sid]


def _debug(session: Dict[str, Any], msg: str, level: str = "INFO"):
    """Append a timestamped debug line to the session's log buffer."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    entry = f"[{ts}] {level}: {msg}"
    logs = session["debug_logs"]
    logs.append(entry)
    if len(logs) > 200:
        session["debug_logs"] = logs[-200:]
    logger.info(msg) if level == "INFO" else logger.error(msg)


def _debug_summary(session: Dict[str, Any]) -> Dict[str, Any]:
    """Return the tail of the session debug log."""
    logs = session["debug_logs"]
    return {
        "recent_logs": logs[-20:] if logs else [],
        "total_log_entries": len(logs),
    }


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEFAULT_PRIORITY_RAILS = [
    "P_SOC", "P_CPU", "P_VCCCORE", "P_VCCGT", "P_VCCSA",
    "P_VDD2_CPU", "P_VDDQ_CPU",
    "P_VCCPRIM", "P_IPU", "P_ISP",
    "P_MEMORY", "P_DDR_AB", "P_DDR_CD", "P_DISPLAY", "P_WIFI",
    "P_STORAGE", "P_THUNDERBOLT", "P_SENSOR_HUB", "P_EC",
]

_SOC_KEYWORDS = [
    "SOC", "CPU", "VCCCORE", "VCCGT", "VCCSA", "ECORE",
    "VDD2_CPU", "VDDQ_CPU", "VCCPRIM", "VCCRTC", "VNNAON", "LP_ECORE", "VCCST",
]


# ---------------------------------------------------------------------------
# Pure helper functions (no state, no side-effects)
# ---------------------------------------------------------------------------

def _fuzzy_match(requested: str, available: dict, threshold: int = 80):
    """Return (best_match_name, score) or (None, 0).

    Also handles the PACS P_VAL_ naming convention where a requested
    name like 'P_VCCCORE' must match 'P_VAL_VCCCORE_PH1_R5E8_0.002'.
    """
    best, best_score = None, 0
    req = requested.upper()
    # Core keyword: strip P_ or P_VAL_ prefix and trailing resistor designator
    # e.g. 'P_VCCCORE' -> 'VCCCORE', 'P_VAL_VCCCORE_PH1_R5E8_0.002' -> 'VCCCORE'
    import re as _re
    _core = _re.sub(r'^P_VAL_|^P_', '', req)           # strip prefix
    _core = _re.sub(r'_R\d+[A-Z]\d+_.*$', '', _core)   # strip resistor suffix
    for name in available:
        n = name.upper()
        score = int(SequenceMatcher(None, req, n).ratio() * 100)
        if req in n or n in req:
            score = max(score, 85)
        # P_VCCCORE  matches  P_VAL_VCCCORE_PH1_R5E8_0.002
        if _core and len(_core) >= 3 and _core in n:
            score = max(score, 87)
        if score > best_score:
            best, best_score = name, score
    return (best, best_score) if best_score >= threshold else (None, 0)


def _is_soc_rail(name: str) -> bool:
    upper = name.upper()
    return any(kw in upper for kw in _SOC_KEYWORDS)


def _safe_file_path(file_path: str) -> str:
    """Copy UNC-path files to a local temp folder; return local paths as-is."""
    if not (file_path.startswith("\\\\") or file_path.startswith("//")):
        return file_path
    temp_dir = Path(tempfile.gettempdir()) / "power_tools_cache"
    temp_dir.mkdir(parents=True, exist_ok=True)
    original_name = Path(file_path).name
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    uid = uuid.uuid4().hex[:8]
    local = temp_dir / f"{ts}_{uid}_{original_name}"
    shutil.copy2(file_path, local)
    return str(local)


def _safe_float(value) -> Optional[float]:
    if value is None or (isinstance(value, float) and pd.isna(value)) or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Format-specific loaders (internal — no @mcp.tool)
# ---------------------------------------------------------------------------

def _load_flexlogger_summary_csv(csv_path: str) -> Dict[str, Any]:
    """Parse a FlexLogger *_Raw_Summary.csv (horizontal layout)."""
    try:
        safe_path = _safe_file_path(csv_path)
        df = pd.read_csv(safe_path, index_col=0)
        property_rows = {}
        for prop in ("Min_Value", "Max_Value", "Total_Average", "Total_Energy"):
            if prop in df.index:
                property_rows[prop] = df.loc[prop]
        cols = [c for c in df.columns if c != "TimeStamp" and not c.startswith("Unnamed")]
        stats: Dict[str, Any] = {}
        for sig in cols:
            try:
                mn = _safe_float(property_rows["Min_Value"][sig]) if "Min_Value" in property_rows else None
                mx = _safe_float(property_rows["Max_Value"][sig]) if "Max_Value" in property_rows else None
                av = _safe_float(property_rows["Total_Average"][sig]) if "Total_Average" in property_rows else None
                en = _safe_float(property_rows["Total_Energy"][sig]) if "Total_Energy" in property_rows else None
                if mn is not None or mx is not None or av is not None or en is not None:
                    stats[sig] = {"min": mn, "max": mx, "average": av, "energy": en}
            except Exception:
                continue
        return {"success": True, "summary_stats": stats, "total_signals": len(stats)}
    except Exception as exc:
        return {"success": False, "error": f"FlexLogger summary CSV parse error: {exc}"}


def _load_flexlogger_xml_config(xml_path: str, session: Dict[str, Any]) -> Dict[str, Any]:
    """LLM-based XML config parser for FlexLogger (uses AutoBots SDK LLM wrapper)."""
    try:
        from autobots_sdk.base.models.langchain.base_chat_model import (
            AutobotsLangchain_AzureChatOpenAI,
            HumanMessage,
        )

        llm = AutobotsLangchain_AzureChatOpenAI()
        safe_path = _safe_file_path(xml_path)

        with open(safe_path, "r", encoding="utf-8") as f:
            xml_content = f.read()

        chunk_size = 8000
        chunks = [xml_content[i:i + chunk_size] for i in range(0, len(xml_content), chunk_size)]
        _debug(session, f"[XML] {len(xml_content)} chars → {len(chunks)} chunks")

        chunk_results = []
        for idx, chunk in enumerate(chunks):
            prompt = (
                f"You are analyzing chunk {idx+1}/{len(chunks)} of a FlexLogger XML config.\n"
                "Extract ALL relevant info as JSON:\n"
                '{"soc_power_formula": "...", "rail_mappings": [...], '
                '"test_metadata": {...}, "power_rail_names": [...], "voltage_ranges": {...}}\n'
                "Rules: return ONLY valid JSON, no markdown.\n"
                f"XML CHUNK:\n{chunk}"
            )
            try:
                resp = llm.invoke([HumanMessage(content=prompt)])
                text = resp.content.strip()
                if text.startswith("```"):
                    text = text.split("```")[1]
                    if text.startswith("json"):
                        text = text[4:]
                    text = text.strip()
                chunk_results.append(json.loads(text))
            except Exception as e:
                _debug(session, f"[XML] chunk {idx+1} error: {e}", "ERROR")
                chunk_results.append({"rail_mappings": [], "soc_power_formula": None,
                                      "test_metadata": {}, "power_rail_names": [], "voltage_ranges": {}})

        # Merge
        merged: Dict[str, Any] = {
            "soc_power_formula": None, "rail_mappings": [], "test_metadata": {},
            "power_rail_names": [], "voltage_ranges": {}, "errors": [],
        }
        existing_sigs: set = set()
        for cr in chunk_results:
            if not isinstance(cr, dict):
                continue
            if not merged["soc_power_formula"] and cr.get("soc_power_formula"):
                merged["soc_power_formula"] = cr["soc_power_formula"]
            for rm in cr.get("rail_mappings") or []:
                if isinstance(rm, dict) and rm.get("signal_name") and rm["signal_name"] not in existing_sigs:
                    merged["rail_mappings"].append(rm)
                    existing_sigs.add(rm["signal_name"])
            tm = cr.get("test_metadata")
            if isinstance(tm, dict):
                for k, v in tm.items():
                    if v and not merged["test_metadata"].get(k):
                        merged["test_metadata"][k] = v
            for rn in cr.get("power_rail_names") or []:
                if rn and rn not in merged["power_rail_names"]:
                    merged["power_rail_names"].append(rn)
            vr = cr.get("voltage_ranges")
            if isinstance(vr, dict):
                merged["voltage_ranges"].update(vr)

        # Convert to backward-compatible structure
        channels: Dict[str, Any] = {}
        rail_mapping: Dict[str, Any] = {}
        for ri in merged.get("rail_mappings", []):
            sig = ri.get("signal_name", "")
            mt = ri.get("measurement_type", "").upper()
            rn = sig
            for pfx in ("V_VAL_", "I_VAL_", "V_", "I_"):
                if rn.upper().startswith(pfx):
                    rn = sig[len(pfx):]
                    break
            if rn not in channels:
                channels[rn] = {}
            ci = {"physical_channel": ri.get("physical_channel", "Unknown"), "measurement_type": mt}
            if "VOLTAGE" in mt or "VOLT" in mt:
                channels[rn]["voltage_signal"] = sig
                channels[rn]["voltage_info"] = ci
                daq_parts = ci["physical_channel"].split("/")
                rail_mapping[sig] = {"signal_name": sig, "daq_slot": daq_parts[0] if daq_parts else "Unknown",
                                     "channel": daq_parts[1] if len(daq_parts) > 1 else "Unknown", "type": "Voltage"}
            elif "CURRENT" in mt or "CURR" in mt:
                channels[rn]["current_signal"] = sig
                channels[rn]["current_info"] = ci
                daq_parts = ci["physical_channel"].split("/")
                rail_mapping[sig] = {"signal_name": sig, "daq_slot": daq_parts[0] if daq_parts else "Unknown",
                                     "channel": daq_parts[1] if len(daq_parts) > 1 else "Unknown", "type": "Current"}
            elif "VIRTUAL" in mt or ri.get("formula"):
                channels[rn]["power_formula"] = ri.get("formula", "")
                channels[rn]["power_signal"] = sig
                rail_mapping[sig] = {"signal_name": sig, "daq_slot": "Virtual", "channel": "Calculated",
                                     "type": "Power", "formula": ri.get("formula", "")}

        meta = merged.get("test_metadata", {})
        config = {
            "test_name": meta.get("test_name", "Unknown"),
            "sample_rate": meta.get("sample_rate", 100),
            "acquisition_mode": meta.get("acquisition_mode", "Continuous"),
            "daq": "NI FlexLogger",
        }
        return {
            "success": True, "config": config, "channels": channels,
            "rail_mapping": rail_mapping, "rail_mappings": merged.get("rail_mappings", []),
            "soc_power_formula": merged.get("soc_power_formula"),
            "power_rail_names": merged.get("power_rail_names", []),
            "voltage_ranges": merged.get("voltage_ranges", {}),
            "total_rails": len(channels), "total_signals": len(rail_mapping),
            "llm_extracted": True,
        }
    except Exception as exc:
        return {"success": False, "error": f"XML parse error: {exc}",
                "traceback": traceback.format_exc()[:1000]}


# =====================================================================
# MCP TOOL 1 — Parse power config file
# =====================================================================
def parse_power_config(
    config_path: Annotated[str, Field(
        description="Path to the configuration file (.xml, .csv, or .pkl)"
    )],
    ctx: Context,
) -> Dict[str, Any]:
    """Parse power measurement configuration file and store rail context in session."""
    sess = _get_session(ctx)
    _debug(sess, f"parse_power_config called: {config_path}")

    try:
        # --- FlexLogger XML ---
        if config_path.lower().endswith(".xml"):
            xml_result = _load_flexlogger_xml_config(config_path, sess)
            if not xml_result["success"]:
                return xml_result

            soc_formulas: Dict[str, str] = {}
            plat_formulas: Dict[str, str] = {}
            all_formulas: Dict[str, str] = {}
            for ri in xml_result.get("rail_mappings", []):
                sig = ri.get("signal_name", "")
                formula = ri.get("formula")
                if formula and sig.startswith("P_"):
                    all_formulas[sig] = formula
                    if _is_soc_rail(sig):
                        soc_formulas[sig] = formula
                    else:
                        plat_formulas[sig] = formula
            if not all_formulas:
                for rn, ci in xml_result["channels"].items():
                    v = ci.get("voltage_signal", f"V_{rn}")
                    i = ci.get("current_signal", f"I_{rn}")
                    bucket = soc_formulas if _is_soc_rail(rn) else plat_formulas
                    bucket[f"P_{rn}"] = f"{v} * {i}"

            sess["power_rail_context"] = {
                "soc_power_formula": xml_result.get("soc_power_formula"),
                "soc_rail_formulas": soc_formulas,
                "platform_rail_formulas": plat_formulas,
                "power_rail_names": xml_result.get("power_rail_names", []),
                "voltage_ranges": xml_result.get("voltage_ranges", {}),
                "rail_mapping": xml_result["rail_mapping"],
                "config": xml_result["config"],
                "format": "FlexLogger",
                "updated": True,
            }
            return {
                "success": True, "format": "FlexLogger",
                "soc_power_formula": xml_result.get("soc_power_formula"),
                "rail_mapping": xml_result["rail_mapping"],
                "soc_rail_formulas": soc_formulas,
                "platform_rail_formulas": plat_formulas,
                "power_rail_names": xml_result.get("power_rail_names", []),
                "voltage_ranges": xml_result.get("voltage_ranges", {}),
                "total_signals": xml_result.get("total_signals", 0),
                "total_formulas": len(all_formulas),
                "llm_extracted": xml_result.get("llm_extracted", False),
                "message": (
                    f"FlexLogger config: {xml_result.get('total_signals', 0)} signals, "
                    f"{len(soc_formulas)} SOC + {len(plat_formulas)} platform formulas"
                ),
            }

        # --- PACS CSV / pickle ---
        safe_path = _safe_file_path(config_path)
        if safe_path.endswith(".pkl"):
            with open(safe_path, "rb") as f:
                config_df = pickle.load(f)
        else:
            config_df = pd.read_csv(safe_path)

        rail_mapping: Dict[str, Any] = {}
        voltage_sigs: list = []
        current_sigs: list = []
        power_formulas: Dict[str, str] = {}
        soc_formulas = {}
        plat_formulas = {}

        for _, row in config_df.iterrows():
            sig = str(row.get("Signal Name", "")).strip()
            if not sig:
                continue
            if sig.startswith("V_") or sig.startswith("I_"):
                rail_mapping[sig] = {
                    "signal_name": sig,
                    "math": str(row.get("Math", "")),
                    "daq_slot": str(row.get("DAQ", "")),
                    "channel": str(row.get("Channel", "")),
                    "type": str(row.get("Type", "")),
                    "range": str(row.get("Range", "")),
                }
                (voltage_sigs if sig.startswith("V_") else current_sigs).append(sig)
            elif sig.startswith("P_"):
                formula = str(row.get("Math", "")).strip()
                if formula:
                    power_formulas[sig] = formula
                    bucket = soc_formulas if _is_soc_rail(sig) else plat_formulas
                    bucket[sig] = formula

        sess["power_rail_context"] = {
            "soc_rails": soc_formulas,
            "platform_rails": plat_formulas,
            "rail_formulas": power_formulas,
            "updated": True,
        }
        return {
            "success": True,
            "rail_mapping": rail_mapping,
            "total_signals": len(rail_mapping),
            "voltage_signals_count": len(voltage_sigs),
            "current_signals_count": len(current_sigs),
            "power_formulas_count": len(power_formulas),
            "soc_rail_formulas": soc_formulas,
            "soc_rails_count": len(soc_formulas),
            "platform_rail_formulas": plat_formulas,
            "platform_rails_count": len(plat_formulas),
            "message": (
                f"PACS config: {len(rail_mapping)} signals, "
                f"{len(soc_formulas)} SOC + {len(plat_formulas)} platform formulas"
            ),
        }
    except Exception as exc:
        return {"success": False, "error": f"Config parse failed: {exc}"}


# =====================================================================
# MCP TOOL 2 — Analyze power summary (PRIMARY tool, 95 % of queries)
# =====================================================================
def analyze_power_summary(
    summary_path: Annotated[str, Field(
        description="Path to the summary file (.csv or .pkl)"
    )],
    ctx: Context,
    requested_rails: Annotated[Optional[List[str]], Field(
        description="Specific rail names to retrieve. Uses fuzzy matching. "
                    "Default: priority list (~20 important rails).",
        default=None,
    )] = None,
    top_n_rails: Annotated[int, Field(
        description="Fallback: return top N rails by average power. Default 20.",
        default=20,
    )] = 20,
    rail_filter: Annotated[str, Field(
        description="Fallback: substring filter (e.g. 'CPU', 'MEMORY').",
        default="",
    )] = "",
    fuzzy_threshold: Annotated[int, Field(
        description="Minimum fuzzy-match similarity 0-100. Default 80.",
        default=80,
    )] = 80,
) -> Dict[str, Any]:
    """Analyze power summary with per-session isolation and fuzzy matching."""
    sess = _get_session(ctx)
    _debug(sess, f"analyze_power_summary: path={summary_path}, "
                 f"requested_rails={requested_rails}, top_n={top_n_rails}, filter={rail_filter}")

    try:
        # Smart default
        use_default = False
        if requested_rails is None and not rail_filter and top_n_rails > 0:
            requested_rails = list(DEFAULT_PRIORITY_RAILS)
            use_default = True
            _debug(sess, f"Using {len(DEFAULT_PRIORITY_RAILS)} default priority rails")

        safe_path = _safe_file_path(summary_path)
        is_pickle = safe_path.endswith(".pkl")

        # --- Detect FlexLogger CSV ---
        is_flexlogger_csv = False
        if not is_pickle and safe_path.endswith(".csv"):
            try:
                hdr = pd.read_csv(safe_path, nrows=0)
                cols = [str(c).strip() for c in hdr.columns]
                if "Property Name" in cols or (cols and "Property" in cols[0]):
                    is_flexlogger_csv = True
                elif len(cols) > 100:
                    is_flexlogger_csv = True
            except Exception:
                is_flexlogger_csv = summary_path.endswith("_Raw_Summary.csv")

        if is_flexlogger_csv:
            fl = _load_flexlogger_summary_csv(safe_path)
            if not fl["success"]:
                return fl
            return _classify_and_filter(
                fl["summary_stats"], "FlexLogger",
                requested_rails, use_default, rail_filter, top_n_rails, fuzzy_threshold, sess,
            )

        # --- Load DataFrame ---
        if is_pickle:
            with open(safe_path, "rb") as f:
                summary_df = pickle.load(f)
        else:
            summary_df = pd.read_csv(safe_path)

        # --- Detect FlexLogger pickle ---
        is_fl_pickle = False
        if is_pickle:
            idx_vals = [str(i) for i in summary_df.index[:10]]
            fl_markers = ["Min_Value", "Max_Value", "Total_Average", "Total_Energy", "Property Name"]
            if any(m in idx_vals for m in fl_markers):
                is_fl_pickle = True
            elif len(summary_df.columns) > 50:
                is_fl_pickle = True

        if is_fl_pickle:
            props = {}
            for pn in ("Min_Value", "Max_Value", "Total_Average", "Total_Energy"):
                if pn in summary_df.index:
                    props[pn] = summary_df.loc[pn]
            all_rails: Dict[str, Any] = {}
            for sig in summary_df.columns:
                if sig.startswith("P_"):
                    try:
                        mn = _safe_float(props.get("Min_Value", pd.Series()).get(sig)) if "Min_Value" in props else None
                        mx = _safe_float(props.get("Max_Value", pd.Series()).get(sig)) if "Max_Value" in props else None
                        av = _safe_float(props.get("Total_Average", pd.Series()).get(sig)) if "Total_Average" in props else None
                        en = _safe_float(props.get("Total_Energy", pd.Series()).get(sig)) if "Total_Energy" in props else None
                        all_rails[sig] = {"min": mn, "max": mx, "average": av, "energy": en}
                    except Exception:
                        continue
            return _classify_and_filter(
                all_rails, "FlexLogger",
                requested_rails, use_default, rail_filter, top_n_rails, fuzzy_threshold, sess,
            )

        # --- PACS format ---
        # Collect ALL P_* rails first, then route through _classify_and_filter
        # so that requested_rails / fuzzy-match / top_n all work identically
        # to the FlexLogger path.
        all_pacs_rails: Dict[str, Any] = {}
        for _, row in summary_df.iterrows():
            sig = str(row.get("Name", "")).strip()
            if not sig.startswith("P_"):
                continue
            try:
                peak = float(row.get("Peak", 0))
                avg = float(row.get("Average", 0))
                pt = float(row.get("Peak Time", 0))
            except (ValueError, TypeError):
                continue
            all_pacs_rails[sig] = {"peak": peak, "average": avg, "peak_time": pt}

        # Synthesise a P_SOC aggregate when it is not measured directly
        if "P_SOC" not in all_pacs_rails:
            soc_components = {k: v for k, v in all_pacs_rails.items() if _is_soc_rail(k)}
            if soc_components:
                all_pacs_rails["P_SOC"] = {
                    "peak": sum(v["peak"] for v in soc_components.values()),
                    "average": sum(v["average"] for v in soc_components.values()),
                    "peak_time": 0.0,
                    "computed": True,  # flag so callers know it is derived
                }
                _debug(sess, f"PACS: synthesised P_SOC from {len(soc_components)} SOC components")

        _debug(sess, f"PACS: {len(all_pacs_rails)} total P_* rails before filtering")
        result = _classify_and_filter(
            all_pacs_rails, "PACS",
            requested_rails, use_default, rail_filter, top_n_rails, fuzzy_threshold, sess,
        )
        # Attach P_SOC total so callers can surface it easily
        soc_agg = all_pacs_rails.get("P_SOC", {})
        result["total_soc_power"] = {
            "average": soc_agg.get("average", 0.0),
            "peak": soc_agg.get("peak", 0.0),
            "computed": soc_agg.get("computed", False),
        }
        return result
    except Exception as exc:
        _debug(sess, f"analyze_power_summary error: {exc}", "ERROR")
        return {"success": False, "error": f"Summary analysis failed: {exc}"}


def _classify_and_filter(
    all_power_rails: Dict[str, Any],
    fmt: str,
    requested_rails: Optional[List[str]],
    use_default: bool,
    rail_filter: str,
    top_n: int,
    threshold: int,
    sess: Dict[str, Any],
) -> Dict[str, Any]:
    """Classify rails into SOC/platform and apply fuzzy filtering."""
    soc = {k: v for k, v in all_power_rails.items() if _is_soc_rail(k)}
    plat = {k: v for k, v in all_power_rails.items() if not _is_soc_rail(k)}

    matched: Dict[str, Any] = {}
    unmatched: list = []
    details: list = []
    soc_f: Dict[str, Any] = {}
    plat_f: Dict[str, Any] = {}
    top_rails: Dict[str, Any] = {}

    if requested_rails:
        for rr in requested_rails:
            name, score = _fuzzy_match(rr, all_power_rails, threshold)
            if name:
                matched[name] = all_power_rails[name]
                details.append({"requested": rr, "matched": name, "similarity": score})
            else:
                unmatched.append(rr)
        soc_f = {k: v for k, v in matched.items() if k in soc}
        plat_f = {k: v for k, v in matched.items() if k in plat}
        top_rails = matched
    elif rail_filter:
        fu = rail_filter.upper()
        soc_f = {k: v for k, v in soc.items() if fu in k.upper()}
        plat_f = {k: v for k, v in plat.items() if fu in k.upper()}
        top_rails = {**soc_f, **plat_f}
    elif top_n > 0:
        sorted_r = sorted(all_power_rails.items(), key=lambda x: (x[1].get("average") or 0), reverse=True)
        top_rails = dict(sorted_r[:top_n])
        soc_f = {k: v for k, v in top_rails.items() if k in soc}
        plat_f = {k: v for k, v in top_rails.items() if k in plat}

    _debug(sess, f"{fmt}: {len(top_rails)} returned / {len(all_power_rails)} total")

    if requested_rails and use_default:
        msg = (f"{fmt}: {len(matched)}/{len(requested_rails)} priority rails matched "
               f"from {len(all_power_rails)} total")
    elif requested_rails:
        msg = (f"{fmt}: {len(matched)}/{len(requested_rails)} requested rails matched "
               f"from {len(all_power_rails)} total")
    elif rail_filter:
        msg = f"{fmt}: '{rail_filter}' → {len(top_rails)} rails from {len(all_power_rails)} total"
    else:
        msg = f"{fmt}: top {len(top_rails)} rails from {len(all_power_rails)} total"

    return {
        "success": True, "format": fmt,
        "soc_rails": soc_f, "platform_rails": plat_f,
        "top_rails": top_rails,
        "total_rails_in_file": len(all_power_rails),
        "total_soc_rails": len(soc), "total_platform_rails": len(plat),
        "returned_rail_count": len(top_rails),
        "requested_rails": requested_rails or [],
        "fuzzy_match_details": details,
        "matched_rail_count": len(matched) if requested_rails else 0,
        "unmatched_rails": unmatched,
        "used_default_priority_rails": use_default,
        "filter_applied": rail_filter or "none",
        "top_n_limit": top_n, "fuzzy_threshold": threshold,
        "message": msg,
    }


# =====================================================================
# MCP TOOL 3 — Analyze trace files (time-series)
# =====================================================================
def analyze_power_traces(
    ctx: Context,
    math_trace_path: Annotated[Optional[str], Field(
        description="Path to math-trace file (.csv or .pkl)", default=None
    )] = None,
    channel_trace_path: Annotated[Optional[str], Field(
        description="Path to channel-trace file (.csv or .pkl)", default=None
    )] = None,
    time_start: Annotated[Optional[float], Field(
        description="Start of time window in seconds", default=None
    )] = None,
    time_end: Annotated[Optional[float], Field(
        description="End of time window in seconds", default=None
    )] = None,
    signals: Annotated[Optional[List[str]], Field(
        description="Signal names to analyze. Default: all numeric columns.", default=None
    )] = None,
) -> Dict[str, Any]:
    """Compute per-signal statistics over an optional time window."""
    sess = _get_session(ctx)
    _debug(sess, f"analyze_power_traces: math={math_trace_path}, channel={channel_trace_path}")

    if not math_trace_path and not channel_trace_path:
        return {"success": False, "error": "Provide at least one trace file path."}

    try:
        chosen = math_trace_path or channel_trace_path
        safe = _safe_file_path(chosen)
        if safe.endswith(".pkl"):
            with open(safe, "rb") as f:
                df = pickle.load(f)
        else:
            # Optimisation: for large CSVs (e.g. 285 MB math-traces) only load
            # the Time column + the requested signals instead of the whole file.
            if signals:
                available_cols = list(pd.read_csv(safe, nrows=0).columns)
                usecols = ["Time"] + [s for s in signals if s in available_cols]
                missing = [s for s in signals if s not in available_cols]
                if missing:
                    _debug(sess, f"Trace: signals not found in file: {missing}", "INFO")
                df = pd.read_csv(safe, usecols=usecols)
            else:
                df = pd.read_csv(safe)
        trace_type = "math_trace" if math_trace_path else "channel_trace"

        if time_start is not None and time_end is not None:
            df = df[(df["Time"] >= time_start) & (df["Time"] <= time_end)]
            tr = (time_start, time_end)
        else:
            tr = (float(df["Time"].min()), float(df["Time"].max()))

        if signals is None:
            signals = [c for c in df.columns if c != "Time" and pd.api.types.is_numeric_dtype(df[c])]
        else:
            signals = [s for s in signals if s in df.columns]

        stats: Dict[str, Any] = {}
        for s in signals:
            try:
                stats[s] = {
                    "min": float(df[s].min()), "max": float(df[s].max()),
                    "mean": float(df[s].mean()), "median": float(df[s].median()),
                    "std": float(df[s].std()),
                }
            except Exception as e:
                stats[s] = {"error": str(e)}

        _debug(sess, f"Trace: {len(stats)} signals, {len(df)} points, {tr[0]:.2f}–{tr[1]:.2f}s")
        return {
            "success": True, "trace_type": trace_type,
            "time_range_analyzed": tr, "data_points": len(df),
            "signal_count": len(stats), "signals_analyzed": list(stats.keys()),
            "statistics": stats,
            "message": (f"{len(stats)} signals, {len(df)} samples, "
                        f"{tr[0]:.2f}s–{tr[1]:.2f}s"),
        }
    except Exception as exc:
        return {"success": False, "error": f"Trace analysis failed: {exc}"}


# =====================================================================
# MCP TOOL 4 — Load CSV into per-session storage
# =====================================================================
def load_power_csv(
    file_path: Annotated[str, Field(description="Path to the CSV file")],
    ctx: Context,
    dataframe_name: Annotated[str, Field(
        description="Name for this DataFrame. Default: 'df'", default="df"
    )] = "df",
) -> Dict[str, Any]:
    """Load CSV and persist a pickle reference in session state."""
    sess = _get_session(ctx)
    try:
        df = pd.read_csv(file_path)
        scratch = os.environ.get("MCP_AGENT_SCRATCH_DIR", tempfile.gettempdir())
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        uid = uuid.uuid4().hex[:8]
        pkl = os.path.join(scratch, f"power_{dataframe_name}_{ts}_{uid}.pkl")
        with open(pkl, "wb") as f:
            pickle.dump(df, f)
        sess["dataframes"][dataframe_name] = pkl
        _debug(sess, f"load_power_csv: {dataframe_name} → {pkl} ({len(df)} rows)")
        return {
            "success": True, "dataframe_name": dataframe_name,
            "shape": {"rows": len(df), "columns": len(df.columns)},
            "columns": list(df.columns),
            "dtypes": {c: str(d) for c, d in df.dtypes.items()},
            "pickle_path": pkl,
            "message": f"Loaded {len(df)} rows × {len(df.columns)} cols as '{dataframe_name}'",
        }
    except Exception as exc:
        return {"success": False, "error": f"CSV load failed: {exc}"}


# =====================================================================
# MCP TOOL 5 — Load JSON into per-session storage
# =====================================================================
def load_power_json(
    file_path: Annotated[str, Field(description="Path to the JSON file")],
    ctx: Context,
    dataframe_name: Annotated[str, Field(
        description="Name for this DataFrame. Default: 'df'", default="df"
    )] = "df",
) -> Dict[str, Any]:
    """Load JSON and persist a pickle reference in session state."""
    sess = _get_session(ctx)
    try:
        try:
            df = pd.read_json(file_path)
        except ValueError:
            # Flat dict JSON (e.g. {"rail_name": value}) — convert to two-column DataFrame
            with open(file_path, "r", encoding="utf-8") as _jf:
                _jdata = json.load(_jf)
            if isinstance(_jdata, dict):
                df = pd.DataFrame(list(_jdata.items()), columns=["name", "value"])
            else:
                raise
        scratch = os.environ.get("MCP_AGENT_SCRATCH_DIR", tempfile.gettempdir())
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        uid = uuid.uuid4().hex[:8]
        pkl = os.path.join(scratch, f"power_{dataframe_name}_{ts}_{uid}.pkl")
        with open(pkl, "wb") as f:
            pickle.dump(df, f)
        sess["dataframes"][dataframe_name] = pkl
        _debug(sess, f"load_power_json: {dataframe_name} → {pkl} ({len(df)} rows)")
        return {
            "success": True, "dataframe_name": dataframe_name,
            "shape": {"rows": len(df), "columns": len(df.columns)},
            "columns": list(df.columns),
            "dtypes": {c: str(d) for c, d in df.dtypes.items()},
            "pickle_path": pkl,
            "message": f"Loaded {len(df)} rows × {len(df.columns)} cols as '{dataframe_name}'",
        }
    except Exception as exc:
        return {"success": False, "error": f"JSON load failed: {exc}"}


# =====================================================================
# MCP TOOL 6 — Analyze a loaded DataFrame with LLM
# =====================================================================
def analyze_power_dataframe(
    dataframe_name: Annotated[str, Field(description="Name of the loaded DataFrame")],
    query: Annotated[str, Field(description="Natural language question about the data")],
    ctx: Context,
) -> Dict[str, Any]:
    """LLM-driven pandas analysis on per-session DataFrames."""
    sess = _get_session(ctx)
    store = sess["dataframes"]
    if dataframe_name not in store:
        return {
            "success": False,
            "error": f"'{dataframe_name}' not found. Available: {list(store.keys())}",
        }
    try:
        pkl = store[dataframe_name]
        with open(pkl, "rb") as f:
            df = pickle.load(f)

        # Try to use the shared utility; fall back gracefully on any failure
        try:
            from _shared_utilities.dataframe_analysis import analyze_dataframe_with_llm
            result = analyze_dataframe_with_llm(
                df=df, query=query,
                domain_context={
                    "domain": "power_management",
                    "dataframe_description": "Power monitoring data",
                },
            )
            # If the external helper returned a failure dict (e.g. LLM credential
            # unreachable), treat it as unavailable and fall through to basic stats.
            if not result.get("success"):
                raise RuntimeError(result.get("error", "LLM analysis returned failure"))
            return result
        except Exception as llm_exc:
            # LLM unavailable (import error, network error, credential error,
            # or the helper returned success=False)
            # → return basic stats as a graceful fallback
            _debug(sess, f"analyze_power_dataframe: LLM unavailable ({type(llm_exc).__name__}), using fallback", "INFO")
            return {
                "success": True,
                "note": f"LLM analysis unavailable ({type(llm_exc).__name__}); returning basic stats",
                "describe": df.describe().to_dict(),
                "head": df.head(10).to_dict(),
                "shape": {"rows": len(df), "columns": len(df.columns)},
            }
    except Exception as exc:
        return {"success": False, "error": f"DataFrame analysis failed: {exc}"}


# =====================================================================
# POWER COMPARISON PIPELINE — Helpers
# =====================================================================
# These port the PowerSocwatchDataCompiler power-CSV pipeline
# (detectPowerRailConfig → processSummaryRailsToJson →
#  create_power_output_summaries) into session-safe FastMCP tools.

# Patterns that identify the top-level SoC/CPU aggregate row in a config.
# Boards that label it P_CPU_TOTAL or P_CPU (instead of P_SOC) are common.
_SOC_NAME_RE = _re_module.compile(
    r"^(p_soc|soc|p_cpu_total|p_cpu_pch_total|p_cpu)$", _re_module.IGNORECASE
)
_TOKEN_RE = _re_module.compile(r"[A-Za-z_][A-Za-z0-9_.]*")
_SOC_FALLBACK_KW = ["soc", "core", "sa", "vnn", "io", "vdd", "prim"]
_PLATFORM_FALLBACK_KW = ["memory", "wlan", "wifi", "camera", "display", "edp",
                          "backlight", "panel", "disp", "storage", "ssd", "audio"]

_FMT_PACS = "PACS_FLEXLOGGER"
_FMT_CATAPULT = "CATAPULT_RAW"
_FMT_CATAPULT_WIDE = "CATAPULT_WIDE"
_FMT_MICROWATT = "MICROWATT"
_FMT_GENERIC = "GENERIC"

_STAT_LABELS = {
    "max_value", "min_value", "total_average", "total_energy",
    "average", "avg", "mean", "median", "max", "min",
    "rms", "std", "sum", "count",
}
_STAT_PREFERENCE = [
    "total_average", "average", "mean", "median", "max_value", "min_value",
]


def _normalize(name: str) -> str:
    return name.strip().lower()


def _canonical_rail_key(name: str) -> str:
    key = _normalize(name)
    key = _re_module.sub(r"^(p_val_|v_val_|i_val_)", "", key)
    key = _re_module.sub(r"^(p_|v_|i_)", "", key)
    key = _re_module.sub(r"_r[0-9a-z]+_[0-9]+(?:\.[0-9]+)?$", "", key)
    key = _re_module.sub(r"[^a-z0-9]", "", key)
    return key


def _build_target_rail_lookup(target_rails: List[str]) -> Dict[str, str]:
    lookup: Dict[str, str] = {}
    for rail in target_rails:
        exact = _normalize(rail)
        canonical = _canonical_rail_key(rail)
        if exact and exact not in lookup:
            lookup[exact] = rail
        if canonical and canonical not in lookup:
            lookup[canonical] = rail
    return lookup


def _match_target_rail(rail_name: str, lookup: Dict[str, str]) -> Optional[str]:
    exact = _normalize(rail_name)
    if exact in lookup:
        return lookup[exact]
    canonical = _canonical_rail_key(rail_name)
    if canonical in lookup:
        return lookup[canonical]
    return None


def _is_power_rail(name: str) -> bool:
    return _normalize(name).startswith("p_")


def _select_column(fieldnames: List[str], candidates: List[str]) -> Optional[str]:
    lower_map = {f.strip().lower(): f for f in fieldnames}
    for candidate in candidates:
        if candidate in lower_map:
            return lower_map[candidate]
    for candidate in candidates:
        for key, original in lower_map.items():
            if candidate in key:
                return original
    return None


def _parse_formula_tokens(formula: str) -> List[str]:
    if not formula:
        return []
    ordered: List[str] = []
    seen: set = set()
    for token in _TOKEN_RE.findall(formula):
        token = token.strip()
        if not token or token.upper() in {"YES", "NO"} or token.isdigit():
            continue
        key = _normalize(token)
        if key not in seen:
            seen.add(key)
            ordered.append(token)
    return ordered


def _rail_sort_priority(name: str) -> int:
    ln = name.lower()
    if "soc" in ln: return 0
    if "core" in ln: return 1
    if "sa" in ln: return 2
    if "gt" in ln: return 3
    if "vnn" in ln: return 4
    return 5


def _sort_rails(rails: List[str]) -> List[str]:
    return sorted(rails, key=lambda n: (_rail_sort_priority(n), _normalize(n)))


def _detect_power_csv_format(csv_path: str) -> Dict[str, Any]:
    """Fingerprint a power summary CSV and return parsing hints."""
    result: Dict[str, Any] = {
        "format_name": _FMT_GENERIC, "rail_col": None, "value_col": None,
        "p_multiplier": 1000.0, "v_multiplier": 1.0, "unit_source": "unknown",
        "fieldnames": [], "is_wide_format": False, "wide_stat_row": "",
    }
    try:
        with open(csv_path, "r", newline="", encoding="utf-8-sig") as fh:
            sample = fh.read(2048)
            fh.seek(0)
            try:
                delimiter = csv.Sniffer().sniff(sample).delimiter
            except Exception:
                delimiter = ","
            reader = csv.DictReader(fh, delimiter=delimiter)
            if not reader.fieldnames:
                return result
            raw_fn = [f for f in reader.fieldnames if f is not None]
            data_rows = []
            for _ in range(8):
                try:
                    data_rows.append(next(reader))
                except StopIteration:
                    break
    except Exception:
        return result

    if not raw_fn:
        return result

    result["fieldnames"] = [f.lower() for f in raw_fn]
    headers_lower = " ".join(result["fieldnames"])

    # Wide-format detection
    first_col = raw_fn[0]
    first_vals = [(_normalize(row.get(first_col) or "")) for row in data_rows]
    stat_hits = [v for v in first_vals if v in _STAT_LABELS]

    if len(stat_hits) >= 2:
        result["is_wide_format"] = True
        result["format_name"] = _FMT_CATAPULT_WIDE
        chosen = ""
        for pref in _STAT_PREFERENCE:
            if pref in stat_hits:
                for row in data_rows:
                    raw_label = (row.get(first_col) or "").strip()
                    if raw_label.lower() == pref:
                        chosen = raw_label
                        break
            if chosen:
                break
        if not chosen and stat_hits:
            for row in data_rows:
                raw_label = (row.get(first_col) or "").strip()
                if raw_label.lower() in _STAT_LABELS:
                    chosen = raw_label
                    break
        result["wide_stat_row"] = chosen
        result["rail_col"] = first_col
        result["value_col"] = raw_fn[1] if len(raw_fn) > 1 else raw_fn[0]
    else:
        result["rail_col"] = _select_column(raw_fn, [
            "signal name", "signal", "rail", "name", "channel", "label", "measurement",
        ]) or raw_fn[0]
        result["value_col"] = _select_column(raw_fn, [
            "average (mw)", "average(mw)", "avg (mw)", "avg(mw)",
            "average (w)", "average(w)", "avg (w)", "avg(w)",
            "average", "avg", "mean", "value", "total", "sum", "power", "watt", "mw", "_w",
        ]) or (raw_fn[1] if len(raw_fn) > 1 else raw_fn[0])

    has_mw = "(mw)" in headers_lower or "milliwatt" in headers_lower
    has_w = "(w)" in headers_lower and not has_mw
    has_uw = "(uw)" in headers_lower or "(µw)" in headers_lower or "microwatt" in headers_lower
    is_cat = bool(_re_module.match(r"^\d{8}t\d{6}", Path(csv_path).stem.lower()))

    # PACS files use  Name | Peak | Average | Peak Time  with no unit annotation.
    # Detect this pattern: rail_col resolved to 'name'/'signal name' AND data rows
    # have values whose first column starts with P_ or V_.
    _rail_col_lower = (result["rail_col"] or "").strip().lower()
    _is_pacs_name_avg = (
        not result["is_wide_format"]
        and not has_mw and not has_w and not has_uw and not is_cat
        and _rail_col_lower in {"name", "signal name", "signal", "rail"}
        and any(
            (row.get(result["rail_col"]) or "").strip().upper().startswith(("P_", "V_"))
            for row in data_rows
        )
    )

    if not result["is_wide_format"]:
        if has_uw:
            result.update(format_name=_FMT_MICROWATT, p_multiplier=0.001, v_multiplier=0.001, unit_source="µW")
        elif has_mw:
            result.update(format_name=_FMT_PACS if not is_cat else _FMT_CATAPULT, p_multiplier=1.0, v_multiplier=1.0, unit_source="mW")
        elif has_w or is_cat:
            result.update(format_name=_FMT_CATAPULT if is_cat else _FMT_PACS, p_multiplier=1000.0, v_multiplier=1.0, unit_source="W")
        elif _is_pacs_name_avg:
            # PACS summary without unit annotation in headers (Name/Peak/Average columns).
            # PACS tools write values in Watts → multiply ×1000 to normalise to mW,
            # consistent with explicit-mW PACS files (p_mult=1.0 after (mW) detection).
            result.update(format_name=_FMT_PACS, p_multiplier=1000.0, v_multiplier=1.0, unit_source="W (inferred-PACS)")
        else:
            result.update(format_name=_FMT_GENERIC, p_multiplier=1000.0, v_multiplier=1.0, unit_source="unknown (assumed W)")
    else:
        if has_uw:
            result.update(p_multiplier=0.001, v_multiplier=0.001, unit_source="µW")
        elif has_mw:
            result.update(p_multiplier=1.0, v_multiplier=1.0, unit_source="mW")
        else:
            result.update(p_multiplier=1000.0, v_multiplier=1.0, unit_source="W")
    return result


def _parse_summary_for_target_rails(summary_csv: str, target_lookup: Dict[str, str]) -> Dict[str, float]:
    """Parse a summary CSV and extract values for target rails with unit conversion."""
    extracted: Dict[str, float] = {}
    fmt = _detect_power_csv_format(summary_csv)
    p_mult = fmt["p_multiplier"]
    v_mult = fmt["v_multiplier"]

    try:
        with open(summary_csv, "r", newline="", encoding="utf-8-sig") as fh:
            sample = fh.read(1024)
            fh.seek(0)
            try:
                delim = csv.Sniffer().sniff(sample).delimiter
            except Exception:
                delim = ","
            reader = csv.DictReader(fh, delimiter=delim)
            if not reader.fieldnames:
                return extracted
            raw_fn = [f for f in reader.fieldnames if f is not None]

            if fmt.get("is_wide_format"):
                stat_col = raw_fn[0]
                rail_cols = raw_fn[1:]
                target_stat = fmt.get("wide_stat_row", "")
                for row in reader:
                    if (row.get(stat_col) or "").strip() != target_stat:
                        continue
                    for rc in rail_cols:
                        matched = _match_target_rail(rc, target_lookup)
                        if not matched:
                            continue
                        try:
                            val = float((row.get(rc) or "").strip())
                        except ValueError:
                            continue
                        mult = p_mult if matched.upper().startswith("P_") else v_mult
                        extracted[matched] = round(val * mult, 3)
                    break
            else:
                rail_col = fmt["rail_col"]
                value_col = fmt["value_col"]
                if rail_col not in raw_fn:
                    rail_col = _select_column(raw_fn, ["signal name", "signal", "rail", "name"]) or raw_fn[0]
                if value_col not in raw_fn:
                    value_col = _select_column(raw_fn, ["average", "avg", "mean", "value"]) or (raw_fn[1] if len(raw_fn) > 1 else raw_fn[0])
                for row in reader:
                    rn = (row.get(rail_col) or "").strip()
                    if not rn:
                        continue
                    matched = _match_target_rail(rn, target_lookup)
                    if not matched:
                        continue
                    try:
                        val = float((row.get(value_col) or "").strip())
                    except ValueError:
                        continue
                    mult = p_mult if matched.upper().startswith("P_") else v_mult
                    extracted[matched] = round(val * mult, 3)
    except Exception:
        pass
    return extracted


def _load_rail_map(config_csv: str) -> Dict[str, tuple]:
    """Load Signal Name → (original, math) map from configuration.csv."""
    rail_map: Dict[str, tuple] = {}
    with open(config_csv, "r", newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            return rail_map
        sig_col = math_col = None
        for f in reader.fieldnames:
            fl = f.strip().lower()
            if fl == "signal name":
                sig_col = f
            elif fl == "math":
                math_col = f
        if not sig_col or not math_col:
            return rail_map
        for row in reader:
            sig = (row.get(sig_col) or "").strip()
            math_expr = (row.get(math_col) or "").strip()
            if sig:
                rail_map[_normalize(sig)] = (sig, math_expr)
    return rail_map


def _find_soc_root(rail_map: Dict[str, tuple]) -> str:
    """Return the rail_map key for the top-level SOC/CPU aggregate.

    Search order (most-specific first):
      P_SOC, SOC                   ← standard PACS naming
      P_CPU_TOTAL                  ← boards where CPU die == SOC budget
      P_CPU_PCH_TOTAL, P_CPU       ← other CPU aggregate patterns
    Falls back to any key whose original name matches _SOC_NAME_RE.
    """
    for pref in ["p_soc", "soc", "p_cpu_total", "p_cpu_pch_total", "p_cpu"]:
        if pref in rail_map:
            return pref
    for key, (orig, _) in rail_map.items():
        if _SOC_NAME_RE.match(orig.strip()):
            return key
    raise ValueError("Could not find SOC row (P_SOC / SOC / P_CPU_TOTAL)")


def _expand_deps(key: str, rail_map: Dict[str, tuple], visited: set, expanded: list):
    if key in visited:
        return
    visited.add(key)
    _, math_expr = rail_map.get(key, (key, ""))
    for comp in _parse_formula_tokens(math_expr):
        comp_key = _normalize(comp)
        if _is_power_rail(comp) and comp not in expanded:
            expanded.append(comp)
        if comp_key in rail_map:
            _expand_deps(comp_key, rail_map, visited, expanded)


def _is_alias_or_sum(rail_name: str, rail_map: Dict[str, tuple]) -> bool:
    key = _normalize(rail_name)
    if key not in rail_map:
        return False
    _, math_expr = rail_map[key]
    for token in _parse_formula_tokens(math_expr):
        tk = _normalize(token)
        if tk == key:
            continue
        if _is_power_rail(token) and tk in rail_map:
            return True
    return False


def _get_platform_rails(rail_map: Dict[str, tuple]) -> List[str]:
    kw = ["vbata", "audio", "camera", "display", "panel", "edp",
          "backlight", "memory", "ssd", "storage", "wlan", "wifi"]
    out: List[str] = []
    for _, (orig, _) in rail_map.items():
        if not _is_power_rail(orig):
            continue
        ln = orig.lower()
        if any(k in ln for k in kw) and orig not in out:
            out.append(orig)
    return out


def _get_voltage_rails(rail_map: Dict[str, tuple]) -> List[str]:
    kw = ["core", "ecore", "vcccore", "vccsa", "vccgt", "vnnaon"]
    out: List[str] = []
    for _, (orig, _) in rail_map.items():
        ln = orig.lower()
        if not ln.startswith("v_"):
            continue
        if any(k in ln for k in kw) and orig not in out:
            out.append(orig)
    return out


def _is_power_summary_csv(fp: "Path") -> bool:
    """Return True when a CSV file looks like a power summary via content scan.

    Primary check: delegates to _detect_power_csv_format which already fingerprints
    PACS (Signal Name + Average (mW) headers), FlexLogger/Catapult wide format
    (stat-label first column), and Catapult raw ((W) headers).

    Fallback: counts P_/V_ prefixed values in the first column for PACS files
    that omit unit annotations from their headers.
    """
    try:
        fmt = _detect_power_csv_format(str(fp))
        if fmt.get("format_name", _FMT_GENERIC) != _FMT_GENERIC:
            return True
        # Fallback: PACS files with no (mW)/(W) header annotation
        with open(fp, "r", newline="", encoding="utf-8-sig", errors="replace") as fh:
            reader = csv.reader(fh)
            next(reader, None)  # skip header row
            p_count = 0
            for i, row in enumerate(reader):
                if i >= 50:
                    break
                if row and (row[0].strip().upper().startswith("P_") or
                            row[0].strip().upper().startswith("V_")):
                    p_count += 1
                    if p_count >= 2:
                        return True
    except Exception:
        pass
    return False


def _can_read_file(path) -> bool:
    """Return True if the file can be opened and at least 4 KB read."""
    try:
        with open(str(path), "rb") as _f:
            _f.read(4096)
        return True
    except (PermissionError, OSError):
        return False


def _stage_files_to_temp(
    source_folder: str,
    staging_root: str = _DEFAULT_STAGING_ROOT,
) -> Dict[str, Any]:
    """Copy data files from *source_folder* to *staging_root/<tag>* and return the new path.

    Only files whose extension is in _STAGING_COPY_EXTENSIONS are copied.
    Output artefacts already written to Analysis/ sub-folders are skipped.
    The destination sub-folder name is derived from the last two path
    components of *source_folder* so each source maps to a unique staging slot.
    """
    src = Path(source_folder)
    # Build a safe sub-folder name from the last two path components
    parts = [p for p in src.parts if p.strip("/\\")]
    tag = "_".join(parts[-2:]) if len(parts) >= 2 else (parts[-1] if parts else "staged")
    tag = _re_module.sub(r'[<>:"/\\|?*]', "_", tag)
    dst = Path(staging_root) / tag
    try:
        dst.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        return {"success": False, "error": f"Cannot create staging folder '{dst}': {exc}"}

    copied, failed = [], []
    for fp in src.rglob("*"):
        # Skip anything already inside an Analysis/ output tree
        if any(p.name == "Analysis" for p in fp.parents):
            continue
        if fp.is_file() and fp.suffix.lower() in _STAGING_COPY_EXTENSIONS:
            rel = fp.relative_to(src)
            dest_file = dst / rel
            try:
                dest_file.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(str(fp), str(dest_file))
                copied.append(str(rel))
            except Exception as exc:
                failed.append(f"{rel}: {exc}")

    if not copied and failed:
        return {
            "success": False,
            "error": f"All {len(failed)} file(s) failed to copy. First error: {failed[0]}",
            "failures": failed,
        }

    return {
        "success": True,
        "staging_folder": str(dst),
        "source_folder": source_folder,
        "files_copied": len(copied),
        "files_failed": len(failed),
        "failures": failed,
        "message": (
            f"Staged {len(copied)} file(s) from '{source_folder}' → '{dst}'."
            + (f" {len(failed)} file(s) failed to copy." if failed else "")
            + " Use staging_folder as the path for find_power_summary_files and compile_power_data."
        ),
    }


def _find_summary_csvs(folder: str) -> List[str]:
    """Discover power summary CSVs under *folder* using two strategies:

    1. **Filename match** – files whose stem ends with ``_summary`` or
       ``-summary`` (case-insensitive), covering variants like
       ``pacs-summary``, ``pacs_summary``, ``Raw_Summary``, etc.
    2. **Content scan** – any remaining ``.csv`` not matched by name but
       containing recognisable power-rail headers or PACS ``P_/V_`` rail
       rows.  This catches files like ``4Tab-v2_r12_pacs-summary.csv``
       that use a hyphen separator instead of underscore.

    NOTE: This function is only called on a first-time compile or when
    force_reparse=True.  Subsequent calls return from the manifest cache
    without invoking this scan at all.
    """
    name_matched: List[str] = []
    candidates: List["Path"] = []
    for fp in Path(folder).rglob("*.csv"):
        stem_lower = fp.stem.lower()
        if stem_lower.endswith("_summary") or stem_lower.endswith("-summary"):
            name_matched.append(str(fp))
        else:
            candidates.append(fp)

    # Files that are known config/metadata — never treat as summaries
    _CONFIG_NAMES = {"configuration.csv", "powerrailconfig.txt"}

    # Content-based fallback for files that didn't match by name
    name_matched_set = set(name_matched)
    content_matched: List[str] = []
    for fp in candidates:
        if fp.name.lower() in _CONFIG_NAMES:
            continue
        if str(fp) not in name_matched_set and _is_power_summary_csv(fp):
            content_matched.append(str(fp))

    return sorted(name_matched + content_matched)


def _extract_rails_from_summaries_fallback(summary_files: List[str]) -> Dict[str, List[str]]:
    """Fallback rail detection from summary files when no configuration.csv."""
    soc, plat, volt = [], [], []
    seen_s, seen_p, seen_v = set(), set(), set()
    for sf in summary_files:
        try:
            fmt = _detect_power_csv_format(sf)
            with open(sf, "r", newline="", encoding="utf-8-sig") as fh:
                reader = csv.DictReader(fh)
                if not reader.fieldnames:
                    continue
                fns = [f for f in reader.fieldnames if f is not None]
                if fmt.get("is_wide_format"):
                    rails_here = fns[1:]
                else:
                    name_col = fmt.get("rail_col") or fns[0]
                    if name_col not in fns:
                        name_col = fns[0]
                    rails_here = [(row.get(name_col) or "").strip() for row in reader]
                for rn in rails_here:
                    if not rn:
                        continue
                    ln = rn.lower()
                    if ln.startswith("p_"):
                        if any(k in ln for k in _SOC_FALLBACK_KW):
                            key = _normalize(rn)
                            if key not in seen_s:
                                seen_s.add(key)
                                soc.append(rn)
                        if any(k in ln for k in _PLATFORM_FALLBACK_KW):
                            key = _normalize(rn)
                            if key not in seen_p:
                                seen_p.add(key)
                                plat.append(rn)
                    elif ln.startswith("v_"):
                        if any(k in ln for k in _SOC_FALLBACK_KW):
                            key = _normalize(rn)
                            if key not in seen_v:
                                seen_v.add(key)
                                volt.append(rn)
        except Exception:
            continue
    return {"soc": _sort_rails(soc), "platform": _sort_rails(plat), "voltage": _sort_rails(volt)}


def _extract_test_name(filename: str) -> str:
    stem = Path(filename).stem
    # Strip _summary / -summary FIRST so the run-number pattern can match at end-of-string
    stem = _re_module.sub(r"[-_]summary$", "", stem, flags=_re_module.IGNORECASE)
    stem = _re_module.sub(r"_R\d+$", "", stem)
    return stem


def _extract_kpi_and_run(filename: str) -> tuple:
    stem = Path(filename).stem
    # Strip _summary / -summary FIRST so the run-number pattern can match at end-of-string
    stem = _re_module.sub(r"[-_]summary$", "", stem, flags=_re_module.IGNORECASE)
    m = _re_module.search(r"_R(\d+)$", stem)
    if m:
        run = int(m.group(1))
        kpi = stem[:m.start()]
    else:
        m2 = _re_module.search(r"(\d+)$", stem)
        if m2:
            run = int(m2.group(0))
            kpi = stem[:m2.start()].rstrip("_")
        else:
            kpi = stem
            run = 1
    return kpi, run


def _select_run_by_soc(runs: List[Dict[str, float]], mode: str, soc_keys: List[str]):
    if not runs:
        return None
    if len(runs) == 1:
        return runs[0]

    def soc_val(run_dict):
        for k in soc_keys:
            if k in run_dict:
                return run_dict[k]
        return 0.0

    vals = [(soc_val(r), r) for r in runs]
    vals.sort(key=lambda x: x[0])
    mode_l = mode.lower()
    if mode_l == "min":
        return vals[0][1]
    if mode_l == "max":
        return vals[-1][1]
    if mode_l == "average":
        avg_run: Dict[str, float] = {}
        keys_all = {k for r in runs for k in r if k != "run_number"}
        for k in keys_all:
            v = [r[k] for r in runs if k in r]
            avg_run[k] = sum(v) / len(v) if v else 0.0
        return avg_run
    # median
    mid = len(vals) // 2
    return vals[mid][1]


def _write_matrix_csv(output_csv: str, row_order: List[str], matrix: Dict[str, Dict[str, float]]):
    columns = sorted({col for vals in matrix.values() for col in vals})
    with open(output_csv, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Rail", *columns])
        for rail in row_order:
            vals = matrix.get(rail, {})
            writer.writerow([rail] + [vals.get(c, "") for c in columns])


def _write_matrix_xlsx(output_xlsx: str, row_order: List[str], matrix: Dict[str, Dict[str, float]]):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        # If openpyxl is not available, just write CSV instead
        _write_matrix_csv(output_xlsx.replace(".xlsx", ".csv"), row_order, matrix)
        return

    columns = sorted({col for vals in matrix.values() for col in vals})
    wb = Workbook()
    ws = wb.active
    ws.title = "All_Runs"
    header = ["Rail", *columns]
    ws.append(header)
    hfont = Font(bold=True, size=11)
    hfill = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
    for ci in range(1, len(header) + 1):
        c = ws.cell(row=1, column=ci)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for rail in row_order:
        vals = matrix.get(rail, {})
        ws.append([rail] + [vals.get(c, "") for c in columns])
    # Highlight SOC rows
    bold14 = Font(bold=True, size=14)
    for ri in range(2, ws.max_row + 1):
        rn = _re_module.sub(r"[\s_\-]+", "", str(ws.cell(row=ri, column=1).value or "").lower())
        if rn in {"soc", "socpower", "psoc", "vbata", "vbatapower", "pvbata"}:
            for ci in range(1, ws.max_column + 1):
                ws.cell(row=ri, column=ci).font = bold14
    for ci in range(1, ws.max_column + 1):
        col_l = get_column_letter(ci)
        mx = max((len(str(ws.cell(row=r, column=ci).value or "")) for r in range(1, ws.max_row + 1)), default=0)
        ws.column_dimensions[col_l].width = min(max(mx + 2, 12), 60)
    ws.freeze_panes = "B2"
    wb.save(output_xlsx)


def _matrix_to_markdown(row_order: List[str], matrix: Dict[str, Dict[str, float]],
                         rails_filter: Optional[List[str]] = None,
                         kpis_filter: Optional[List[str]] = None) -> str:
    """Render matrix as markdown table, optionally filtering rows (rails) and columns (KPIs)."""
    if rails_filter:
        filtered = [r for r in row_order if r in rails_filter]
        for r in rails_filter:
            if r not in filtered and r in matrix:
                filtered.append(r)
        row_order = filtered
    columns = sorted({col for vals in matrix.values() for col in vals})
    if kpis_filter:
        # Keep only requested KPI columns (case-insensitive prefix match supported)
        kpi_lower = [k.lower() for k in kpis_filter]
        columns = [c for c in columns if any(c.lower().startswith(k) or c.lower() == k for k in kpi_lower)]
    if not columns:
        return "No data to display."
    lines = []
    header = "| Rail | " + " | ".join(columns) + " |"
    sep = "|------|" + "|".join(["------"] * len(columns)) + "|"
    lines.extend([header, sep])
    for rail in row_order:
        vals = matrix.get(rail, {})
        cells = []
        for c in columns:
            v = vals.get(c, "")
            cells.append(f"{v:.3f}" if isinstance(v, (int, float)) else str(v))
        lines.append(f"| {rail} | " + " | ".join(cells) + " |")
    return "\n".join(lines)


# =====================================================================
# MCP TOOL 7 — Detect power rail config (Step 1 of pipeline)
# =====================================================================
def detect_power_rail_config(
    result_folder: Annotated[str, Field(
        description="Path to the result folder containing configuration.csv and/or *_summary.csv files"
    )],
    ctx: Context,
) -> Dict[str, Any]:
    """Generate PowerRailConfig.txt from a result folder."""
    sess = _get_session(ctx)
    _debug(sess, f"detect_power_rail_config: {result_folder}")

    try:
        folder = Path(result_folder)
        if not folder.exists():
            return {"success": False, "error": f"Folder not found: {result_folder}"}

        # Look for configuration.csv
        config_matches = sorted(folder.rglob("configuration.csv"))
        config_csv = config_matches[0] if config_matches else None

        if config_csv:
            _debug(sess, f"Found configuration.csv: {config_csv}")
            rail_map = _load_rail_map(str(config_csv))
            try:
                soc_root_key = _find_soc_root(rail_map)
                soc_name, soc_math = rail_map[soc_root_key]
                direct = _parse_formula_tokens(soc_math)
                expanded: List[str] = []
                visited: set = set()
                for r in direct:
                    if _is_power_rail(r) and r not in expanded:
                        expanded.append(r)
                for r in direct:
                    rk = _normalize(r)
                    if rk in rail_map:
                        _expand_deps(rk, rail_map, visited, expanded)
                final = [r for r in expanded if not _is_alias_or_sum(r, rail_map)]
                soc_rails = [soc_name] + [r for r in final if r != soc_name]
            except ValueError:
                # No SOC root found — use fallback
                soc_rails = [orig for _, (orig, _) in rail_map.items()
                             if _is_power_rail(orig) and any(k in orig.lower() for k in _SOC_FALLBACK_KW)]
            platform_rails = _get_platform_rails(rail_map)
            voltage_rails = _get_voltage_rails(rail_map)
            soc_rails = _sort_rails(soc_rails)
            platform_rails = _sort_rails(platform_rails)
            voltage_rails = _sort_rails(voltage_rails)
        else:
            _debug(sess, "No configuration.csv found — using summary fallback")
            summary_files = _find_summary_csvs(result_folder)
            fb = _extract_rails_from_summaries_fallback(summary_files)
            soc_rails = fb["soc"]
            platform_rails = fb["platform"]
            voltage_rails = fb["voltage"]

        output_path = folder / "PowerRailConfig.txt"
        with output_path.open("w", encoding="utf-8") as out:
            out.write("SOC Power rails\n")
            for r in soc_rails:
                out.write(f"{r}\n")
            out.write("\nPlatform power rails\n")
            for r in platform_rails:
                out.write(f"{r}\n")
            out.write("\nVoltage rails\n")
            for r in voltage_rails:
                out.write(f"{r}\n")

        _debug(sess, f"PowerRailConfig.txt: {len(soc_rails)} SOC, {len(platform_rails)} platform, {len(voltage_rails)} voltage")
        return {
            "success": True,
            "config_path": str(output_path),
            "soc_rails": soc_rails,
            "platform_rails": platform_rails,
            "voltage_rails": voltage_rails,
            "used_configuration_csv": config_csv is not None,
            "message": (
                f"PowerRailConfig.txt generated: {len(soc_rails)} SOC + "
                f"{len(platform_rails)} platform + {len(voltage_rails)} voltage rails"
            ),
        }
    except Exception as exc:
        return {"success": False, "error": f"detect_power_rail_config failed: {exc}"}


# =====================================================================
# MCP TOOL 8 — Process summary rails to JSON (Step 2 of pipeline)
# =====================================================================
def process_summary_rails_to_json(
    result_folder: Annotated[str, Field(
        description="Base result folder (same as Step 1)"
    )],
    power_config_path: Annotated[str, Field(
        description="Path to PowerRailConfig.txt from Step 1"
    )],
    ctx: Context,
) -> Dict[str, Any]:
    """Copy summaries and create filtered JSON files."""
    sess = _get_session(ctx)
    _debug(sess, f"process_summary_rails_to_json: folder={result_folder}, config={power_config_path}")

    try:
        folder = Path(result_folder)
        config_path = Path(power_config_path)
        if not config_path.exists():
            return {"success": False, "error": f"PowerRailConfig.txt not found: {power_config_path}"}

        output_dir = folder / "Analysis" / "power_output"
        output_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(config_path), str(output_dir / "PowerRailConfig.txt"))

        # Read target rails from config
        target_rails: List[str] = []
        seen: set = set()
        section_headers = {"soc power rails", "platform power rails", "voltage rails"}
        with config_path.open("r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line or line.lower() in section_headers:
                    continue
                key = _normalize(line)
                if key not in seen:
                    seen.add(key)
                    target_rails.append(line)

        target_lookup = _build_target_rail_lookup(target_rails)

        summary_files = [
            f for f in _find_summary_csvs(result_folder)
            if str(output_dir) not in str(f)
            and not Path(f).name.lower().startswith("power_output_summary")
        ]

        json_files_created = []
        for sf in summary_files:
            extracted = _parse_summary_for_target_rails(sf, target_lookup)
            json_path = output_dir / f"{Path(sf).stem}.json"
            with json_path.open("w", encoding="utf-8") as jf:
                json.dump(extracted, jf, indent=2)
            json_files_created.append({
                "csv": Path(sf).name, "json": json_path.name,
                "rails_extracted": len(extracted),
            })

        _debug(sess, f"Step 2: {len(json_files_created)} JSON files created in {output_dir}")
        return {
            "success": True,
            "output_dir": str(output_dir),
            "target_rails_count": len(target_rails),
            "summary_files_processed": len(json_files_created),
            "files": json_files_created,
            "message": (
                f"Analysis/power_output created: {len(json_files_created)} summary files → JSON, "
                f"{len(target_rails)} target rails"
            ),
        }
    except Exception as exc:
        return {"success": False, "error": f"process_summary_rails_to_json failed: {exc}"}


# =====================================================================
# MCP TOOL 9 — Create power comparison matrix (Step 3 of pipeline)
# =====================================================================
def create_power_comparison_matrix(
    power_output_dir: Annotated[str, Field(
        description="Path to Analysis/power_output folder from Step 2"
    )],
    ctx: Context,
    format_name: Annotated[str, Field(
        description="Run selection mode: 'Median', 'Average', 'Min', or 'Max'. Default: 'Median'.",
        default="Median",
    )] = "Median",
) -> Dict[str, Any]:
    """Build all power summary deliverables."""
    sess = _get_session(ctx)
    _debug(sess, f"create_power_comparison_matrix: dir={power_output_dir}, mode={format_name}")

    try:
        pod = Path(power_output_dir)
        if not pod.exists():
            return {"success": False, "error": f"Folder not found: {power_output_dir}"}

        config_path = pod / "PowerRailConfig.txt"
        if not config_path.exists():
            return {"success": False, "error": "PowerRailConfig.txt not found in output dir"}

        # Read sections
        section_rails: Dict[str, List[str]] = {"soc": [], "platform": [], "voltage": []}
        current_section = None
        with config_path.open("r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                ll = line.lower()
                if "soc power rails" in ll:
                    current_section = "soc"; continue
                if "platform power rails" in ll:
                    current_section = "platform"; continue
                if "voltage rails" in ll:
                    current_section = "voltage"; continue
                if current_section:
                    section_rails[current_section].append(line)

        rail_order: List[str] = []
        for r in section_rails["soc"] + section_rails["platform"] + section_rails["voltage"]:
            if r not in rail_order:
                rail_order.append(r)

        target_lookup = _build_target_rail_lookup(rail_order)

        # Source CSVs are never copied into the output folder — scan the original
        # result folder (grandparent of pod: <folder>/Analysis/power_output → <folder>)
        source_folder = pod.parent.parent
        summary_files = sorted([
            Path(f) for f in _find_summary_csvs(str(source_folder))
            if str(pod) not in f
            and not Path(f).name.lower().startswith("power_output_summary")
        ])

        all_runs_matrix: Dict[str, Dict[str, float]] = {}
        kpi_runs: Dict[str, List[Dict[str, float]]] = {}

        def _parse_one_summary(sf):
            test_name = _extract_test_name(sf.name)
            kpi_name, run_number = _extract_kpi_and_run(sf.name)
            parsed = _parse_summary_for_target_rails(str(sf), target_lookup)
            return sf, test_name, kpi_name, run_number, parsed

        max_workers = min(len(summary_files), 8) if summary_files else 1
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            parse_results = list(pool.map(_parse_one_summary, summary_files))

        for _sf, test_name, kpi_name, run_number, parsed in parse_results:
            run_data: Dict[str, float] = {"run_number": float(run_number)}
            run_data.update(parsed)
            kpi_runs.setdefault(kpi_name, []).append(run_data)
            for rail, value in parsed.items():
                all_runs_matrix.setdefault(rail, {})[test_name] = value

        # Write all-runs xlsx
        xlsx_path = pod / "Power_output_summary.xlsx"
        _write_matrix_xlsx(str(xlsx_path), rail_order, all_runs_matrix)

        # Build final (selected run) matrix
        final_matrix: Dict[str, Dict[str, float]] = {}
        # Prefer exact top-level aggregates; fall back to any rail whose name
        # contains 'soc' or 'cpu_total' so CPU-named boards work correctly.
        soc_keys = (
            [r for r in section_rails["soc"] if _SOC_NAME_RE.match(r)]
            or [r for r in section_rails["soc"] if "soc" in r.lower() or "cpu_total" in r.lower() or "cpu_pch" in r.lower()]
            or section_rails["soc"][:1]  # last resort: first SOC rail
        )
        for kpi_name, runs in kpi_runs.items():
            selected = _select_run_by_soc(runs, format_name, soc_keys)
            if not selected:
                continue
            for rail in rail_order:
                if rail in selected:
                    final_matrix.setdefault(rail, {})[kpi_name] = selected[rail]

        output_matrix = final_matrix if final_matrix else all_runs_matrix
        fallback_note = ""
        if not final_matrix and all_runs_matrix:
            fallback_note = ("\n\n> **Note**: Median-selection produced no matches — "
                             "showing all runs instead.")

        if not output_matrix:
            return {
                "success": False,
                "error": "No power data could be extracted from the summary files.",
                "summary_files_found": len(summary_files),
                "rail_order_first_10": rail_order[:10],
            }

        final_csv = pod / "Power_output_summary_final.csv"
        _write_matrix_csv(str(final_csv), rail_order, output_matrix)

        md_path = pod / "Power_output_summary_final_markdown.txt"
        md_text = _matrix_to_markdown(rail_order, output_matrix) + fallback_note
        with md_path.open("w", encoding="utf-8") as fh:
            fh.write(md_text)

        kpi_names = sorted(kpi_runs.keys())
        _debug(sess, f"Step 3: {len(output_matrix)} rails × {len(kpi_names)} KPIs → xlsx+csv+md")
        return {
            "success": True,
            "excel_path": str(xlsx_path),
            "csv_path": str(final_csv),
            "markdown_path": str(md_path),
            "rails_in_output": len(output_matrix),
            "summary_files_processed": len(summary_files),
            "kpi_groups": len(kpi_names),
            "kpi_names": kpi_names,
            "selection_mode": format_name,
            "used_fallback": not bool(final_matrix),
            "markdown_preview": md_text[:2000],
            "message": (
                f"Power comparison: {len(output_matrix)} rails × "
                f"{len(kpi_names)} KPIs → Excel + CSV + Markdown."
                f" KPIs: {kpi_names}"
            ),
        }
    except Exception as exc:
        return {"success": False, "error": f"create_power_comparison_matrix failed: {exc}"}


# ---------------------------------------------------------------------------
# Duck-typed context — lets us call internal functions without a real MCP session
# ---------------------------------------------------------------------------
class _PowerCtx:
    session_id = "_internal"


# =====================================================================
# MCP TOOL 1 — Find power summary files (discovery)
# =====================================================================
@mcp.tool(
    description=(
        "Find all power summary CSV files in a folder tree — matched by filename "
        "(*_summary.csv, *-summary.csv, *pacs-summary.csv, etc.) AND by content scan "
        "(PACS P_/V_ rail headers, Average (mW)/(W) columns). "
        "Checks if results already compiled (Analysis/power_output/ exists). "
        "Call this first, then compile_power_data() to build the full comparison matrix."
    ),
    tags={"power", "discovery", "summary"},
)
@async_tool
@embed_if_large(threshold=3000)
def find_power_summary_files(
    parent_folder: Annotated[str, Field(description="Root folder to scan for power summary CSV files (filename or content matched)")],
    force_reparse: Annotated[bool, Field(
        description="Re-discover even if already compiled. Default: False.",
        default=False,
    )] = False,
) -> Dict[str, Any]:
    """Discover power summary CSV files and check for cached compilation."""
    folder = Path(parent_folder)
    if not folder.exists():
        return {
            "found": False, "file_count": 0, "file_names": [], "file_paths": [],
            "already_compiled": False, "message": f"Folder not found: {parent_folder}",
        }
    output_dir = folder / "Analysis" / "power_output"
    md_cache = output_dir / "Power_output_summary_final_markdown.txt"
    _manifest_p = output_dir / "power_compile_manifest.json"

    # Fast path: if markdown + manifest both exist, return the file list from the
    # manifest directly — skip the (slow) _find_summary_csvs scan entirely.
    # The expensive scan only runs on first call or when force_reparse=True.
    if (not force_reparse) and md_cache.exists() and md_cache.stat().st_size > 0 and _manifest_p.exists():
        try:
            import json as _json_fpsf
            with open(_manifest_p, "r", encoding="utf-8") as _mf:
                _man = _json_fpsf.load(_mf)
            summary_files = _man.get("compiled_files", [])
            if summary_files:
                sample_names = [Path(f).name for f in summary_files[:5]]
                more = len(summary_files) - 5
                return {
                    "found": True,
                    "file_count": len(summary_files),
                    "file_paths": summary_files,
                    "file_names": sample_names + ([f"... and {more} more"] if more > 0 else []),
                    "already_compiled": True,
                    "can_read": True,
                    "staging_hint": None,
                    "output_dir": str(output_dir),
                    "message": (
                        "Already compiled — Analysis/power_output/ exists. "
                        "Call compile_power_data() to run/cache the pipeline, then query_power_matrix() to browse results."
                    ),
                }
        except Exception:
            pass  # manifest corrupt → fall through to full scan

    # Full scan — only runs on first call or when force_reparse=True.
    summary_files = [f for f in _find_summary_csvs(str(folder)) if str(output_dir) not in f]
    if not summary_files:
        return {
            "found": False, "file_count": 0, "file_names": [], "file_paths": [],
            "already_compiled": False,
            "message": (
                f"No power summary CSV files found in {parent_folder}. "
                "Searched for *_summary.csv/*-summary.csv (filename) and PACS "
                "P_/V_ rail content (content scan)."
            ),
        }

    # Quick read-test on the first file so the agent knows upfront whether staging is needed
    can_read = _can_read_file(summary_files[0]) if summary_files else False
    staging_hint = (
        f"Files were found but cannot be read (network permission issue). "
        f"Call stage_power_files_to_temp('{parent_folder}') to copy them to "
        f"'{_DEFAULT_STAGING_ROOT}', then re-run from the returned staging_folder."
    ) if (summary_files and not can_read) else None

    sample_names = [Path(f).name for f in summary_files[:5]]
    more = len(summary_files) - 5
    file_names_display = sample_names + ([f"... and {more} more"] if more > 0 else [])
    return {
        "found": True,
        "file_count": len(summary_files),
        "file_paths": summary_files,
        "file_names": file_names_display,
        "already_compiled": False,
        "can_read": can_read,
        "staging_hint": staging_hint,
        "output_dir": str(output_dir),
        "message": (
            staging_hint if staging_hint else
            f"Found {len(summary_files)} summary file(s). "
            f"Call compile_power_data('{parent_folder}') to compile, then query_power_matrix() to browse."
        ),
    }


# =====================================================================
# MCP TOOL 2 — Compile power data (full pipeline + full markdown)
# =====================================================================
@mcp.tool(
    description=(
        "Compile all *_summary.csv files in a folder into a power comparison matrix. "
        "Runs the full 3-step pipeline once and writes Excel/CSV/Markdown to disk. "
        "Returns compact metadata only (rail_count, kpi_names, file paths). "
        "Call this ONCE then use query_power_matrix() for all user-facing data queries."
    ),
    tags={"power", "pipeline", "markdown", "excel", "compilation"},
)
@async_tool
@embed_if_large(threshold=3000)
def compile_power_data(
    parent_folder: Annotated[str, Field(
        description="Root folder containing *_summary.csv files (same folder as find_power_summary_files)"
    )],
    format_name: Annotated[str, Field(
        description="Run selection mode: Median, Average, Min, or Max. Default: Median.",
        default="Median",
    )] = "Median",
    force_recompile: Annotated[bool, Field(
        description="Recompile even if Analysis/power_output/ already exists. Default: False.",
        default=False,
    )] = False,
) -> Dict[str, Any]:
    """Run the full 3-step power pipeline and write Excel/CSV/Markdown to disk.

    Returns compact metadata (rail_count, kpi_names, file paths).
    Use query_power_matrix() — NOT this tool — for any user-facing data tables.
    """

    folder = Path(parent_folder)
    if not folder.exists():
        return {"success": False, "error": f"Folder not found: {parent_folder}"}

    output_dir = folder / "Analysis" / "power_output"
    md_path = output_dir / "Power_output_summary_final_markdown.txt"
    csv_final = output_dir / "Power_output_summary_final.csv"

    def _build_tiered_response(cached: bool, r1=None, r2=None, r3=None) -> Dict[str, Any]:
        """Return compact metadata only — no full table, no large lists."""
        # Derive grouped KPI names from the CSV columns (same grouping as query_power_matrix)
        kpi_groups_grouped: List[str] = []
        rail_count = 0
        if csv_final.exists():
            try:
                with open(csv_final, "r", newline="", encoding="utf-8-sig") as fh:
                    reader = csv.DictReader(fh)
                    fns = reader.fieldnames or []
                    col_names = [c for c in fns[1:] if c]
                    rail_count = sum(1 for _ in reader)
                groups: Dict[str, List[str]] = {}
                for col in col_names:
                    m = _re_module.match(r"\d{8}T\d{6}-(.*)", col)
                    base = m.group(1) if m else col
                    base = _re_module.sub(r"_R\d+[A-Za-z]*$", "", base)
                    groups.setdefault(base, []).append(col)
                kpi_groups_grouped = sorted(groups.keys())
            except Exception:
                pass

        files_processed = r2.get("summary_files_processed", 0) if r2 else 0
        soc_rails = r1.get("soc_rails", []) if r1 else []
        platform_rails = r1.get("platform_rails", []) if r1 else []

        return {
            "success": True,
            "cached": cached,
            "rail_count": rail_count,
            "kpi_names": kpi_groups_grouped,        # grouped names: ['CQP','Netflix','Teams',...]
            "kpi_groups": len(kpi_groups_grouped),
            "soc_rails": soc_rails[:10],            # first 10 only — avoid long list
            "platform_rails": platform_rails[:10],
            "summary_files_processed": files_processed,
            "full_table_path": str(md_path),
            "excel_path": str(output_dir / "Power_output_summary.xlsx"),
            "csv_path": str(csv_final),
            "message": (
                f"Pipeline {'loaded from cache' if cached else 'compiled'}. "
                f"{rail_count} rails × {len(kpi_groups_grouped)} KPI groups. "
                f"KPIs: {kpi_groups_grouped}. "
                f"Excel: {output_dir / 'Power_output_summary.xlsx'}. "
                "Use query_power_matrix(parent_folder, rails=[...], kpis=[...]) to browse results."
            ),
        }

    # Return cached result if available.
    # Fast path: if markdown + manifest both exist, trust the cache without
    # re-scanning the (potentially slow UNC) folder. A full re-scan only
    # happens when force_recompile=True.
    if not force_recompile and md_path.exists() and md_path.stat().st_size > 0:
        _manifest_p = output_dir / "power_compile_manifest.json"
        if _manifest_p.exists():
            return _build_tiered_response(cached=True)
        # No manifest → old-style cache → fall through and re-compile

    # Early read-test: if the first summary CSV can't be opened, tell the agent to stage files first.
    _probe_files = [f for f in _find_summary_csvs(str(folder)) if str(output_dir) not in f]
    if _probe_files and not _can_read_file(_probe_files[0]):
        return {
            "success": False,
            "error": (
                f"Cannot read files in '{parent_folder}' — likely a network permission issue. "
                "Files were found but cannot be opened for processing. "
                f"Call stage_power_files_to_temp(source_folder='{parent_folder}') to copy them to "
                f"'{_DEFAULT_STAGING_ROOT}', then re-run compile_power_data from the returned staging_folder."
            ),
            "files_found": _probe_files,
            "staging_hint": True,
            "default_staging_root": _DEFAULT_STAGING_ROOT,
        }

    _ctx = _PowerCtx()

    # Step 1 — detect rail config
    r1 = detect_power_rail_config(str(folder), _ctx)
    if not r1["success"]:
        return r1

    # Step 2 — extract rails to JSON
    r2 = process_summary_rails_to_json(str(folder), r1["config_path"], _ctx)
    if not r2["success"]:
        return r2

    # Step 3 — build comparison matrix
    r3 = create_power_comparison_matrix(str(output_dir), _ctx, format_name)
    if not r3["success"]:
        return r3

    # Write manifest so future cache checks know which files this compile covered.
    try:
        import json as _json_w
        _manifest_out = output_dir / "power_compile_manifest.json"
        _compiled = sorted([f for f in _find_summary_csvs(str(folder)) if str(output_dir) not in f])
        with open(_manifest_out, "w", encoding="utf-8") as _mfw:
            _json_w.dump({"compiled_files": _compiled}, _mfw, indent=2)
    except Exception:
        pass

    return _build_tiered_response(cached=False, r1=r1, r2=r2, r3=r3)


# =====================================================================
# MCP TOOL 2b — Stage network files to accessible temp location
# =====================================================================
@mcp.tool(
    description=(
        "Copy power-relevant data files from a network/UNC path to an accessible staging area. "
        "Use this when find_power_summary_files returns can_read=False or compile_power_data "
        "returns staging_hint=True (e.g. files found but permission denied on the network share). "
        "Returns staging_folder — pass that path to find_power_summary_files and "
        "compile_power_data instead of the original network path."
    ),
    tags={"power", "staging", "network", "copy"},
)
@async_tool
def stage_power_files_to_temp(
    source_folder: Annotated[str, Field(
        description="Original source folder (UNC or network path) where files were found but cannot be read."
    )],
    staging_root: Annotated[str, Field(
        description=(
            f"Base staging folder to copy files into. "
            f"Defaults to '{_DEFAULT_STAGING_ROOT}'."
        ),
        default=_DEFAULT_STAGING_ROOT,
    )] = _DEFAULT_STAGING_ROOT,
) -> Dict[str, Any]:
    """Stage unreadable network files to an accessible path for power pipeline processing.

    Copies all .csv / .txt / .xlsx / .json files (excluding Analysis/ output trees)
    to staging_root/<derived_tag>/ preserving relative folder structure.
    Returns the staging_folder path ready for find_power_summary_files / compile_power_data.
    """
    return _stage_files_to_temp(source_folder, staging_root)


# =====================================================================
# MCP TOOL 3 — Query power matrix (filtered + averaged, low token cost)
# =====================================================================
@mcp.tool(
    description=(
        "Query the compiled power matrix CSV with automatic multi-run averaging and "
        "compact filtered output. Use this for ALL user-facing browsing after compile_power_data "
        "has run. Returns a small markdown table (< 2 KB) with averaged iteration columns. "
        "Never use compile_power_data repeatedly for filtering — use this tool instead."
    ),
    tags={"power", "query", "matrix", "filter", "average"},
)
@async_tool
def query_power_matrix(
    parent_folder: Annotated[str, Field(
        description="Root folder — same as used for compile_power_data."
    )],
    rails: Annotated[Optional[List[str]], Field(
        description=(
            "Rail names to show. Prefix-match supported: 'P_SOC' or 'P_MEM'. "
            "Default: P_SOC, P_VCCCORE, P_VCC_LP_ECORE, P_VCCSA, P_BACKLIGHT, "
            "P_DISPLAY, P_MEMORY, P_SSD, P_WLAN."
        ),
        default=None,
    )] = None,
    kpis: Annotated[Optional[List[str]], Field(
        description=(
            "KPI workload names to show. Prefix-match: 'Teams' matches Teams, Teams_R1, Teams_ETL. "
            "Default: all KPI groups. Multiple iterations are automatically averaged into one column."
        ),
        default=None,
    )] = None,
    show_individual_runs: Annotated[bool, Field(
        description="If True, show each iteration as its own column instead of averaging. Default: False.",
        default=False,
    )] = False,
) -> Dict[str, Any]:
    """Query the compiled power matrix with averaging and filtering. Low token cost."""
    _PRIORITY = ["P_SOC", "P_VCCCORE", "P_VCC_LP_ECORE", "P_VCCSA",
                 "P_BACKLIGHT", "P_DISPLAY", "P_MEMORY", "P_SSD", "P_WLAN"]

    csv_final = Path(parent_folder) / "Analysis" / "power_output" / "Power_output_summary_final.csv"
    if not csv_final.exists():
        return {
            "success": False,
            "error": "Analysis/power_output/Power_output_summary_final.csv not found. "
                     "Call compile_power_data first.",
        }

    # Read matrix from CSV
    matrix: Dict[str, Dict[str, float]] = {}
    row_order: List[str] = []
    col_names: List[str] = []
    try:
        with open(csv_final, "r", newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            fns = reader.fieldnames or []
            rail_col = fns[0] if fns else "Rail"
            col_names = [c for c in fns[1:] if c]
            for row in reader:
                rail = (row.get(rail_col) or "").strip()
                if not rail:
                    continue
                row_order.append(rail)
                vals: Dict[str, float] = {}
                for c in col_names:
                    v = (row.get(c) or "").strip()
                    try:
                        vals[c] = float(v)
                    except ValueError:
                        pass
                matrix[rail] = vals
    except Exception as e:
        return {"success": False, "error": f"Failed to read matrix CSV: {e}"}

    # Group columns: strip timestamp prefix and trailing _R<digits>[letter] run suffix
    def _col_group(col: str) -> str:
        m = _re_module.match(r"\d{8}T\d{6}-(.*)", col)
        if m:
            col = m.group(1)
        # Strip _R<digits><optional_letters> at end — handles R1, R2, R21, R41, R3
        col = _re_module.sub(r"_R\d+[A-Za-z]*$", "", col)
        return col

    groups: Dict[str, List[str]] = {}
    for col in col_names:
        g = _col_group(col)
        groups.setdefault(g, []).append(col)
    all_group_names = sorted(groups.keys())

    # Filter groups by kpis
    if kpis:
        kpi_lower = [k.lower() for k in kpis]
        selected_groups = {
            g: cols for g, cols in groups.items()
            if any(g.lower().startswith(k) or g.lower() == k for k in kpi_lower)
        }
        if not selected_groups:
            return {
                "success": False,
                "error": f"No KPI groups matched {kpis}.",
                "all_kpi_groups": all_group_names,
            }
    else:
        selected_groups = groups

    # Filter rails
    if rails:
        rails_lower = [r.lower() for r in rails]
        display_rails = [
            r for r in row_order
            if any(r.lower().startswith(rl) or r.lower() == rl for rl in rails_lower)
        ]
        if not display_rails:
            display_rails = row_order[:8]
    else:
        display_rails = [r for r in row_order if r in _PRIORITY]
        if not display_rails:
            display_rails = row_order[:8]

    # Build table (averaged or individual)
    iteration_notes: List[str] = []
    if show_individual_runs:
        out_cols = [c for g_cols in selected_groups.values() for c in g_cols]
        table_data = {r: {c: matrix[r].get(c, "") for c in out_cols}
                      for r in display_rails if r in matrix}
        col_headers = out_cols
    else:
        out_grps = sorted(selected_groups.keys())
        table_data = {}
        for rail in display_rails:
            if rail not in matrix:
                continue
            row_avg: Dict[str, float] = {}
            for grp in out_grps:
                raw_cols = selected_groups[grp]
                vals_for_grp = [matrix[rail][c] for c in raw_cols if c in matrix[rail]]
                if vals_for_grp:
                    row_avg[grp] = round(sum(vals_for_grp) / len(vals_for_grp), 3)
            table_data[rail] = row_avg
        # Build column headers with run count annotation
        col_headers = out_grps
        for g in out_grps:
            n = len(selected_groups[g])
            if n > 1:
                iteration_notes.append(
                    f"{g}: avg of {n} runs ({', '.join(selected_groups[g])})"
                )

    # Render markdown
    if show_individual_runs:
        col_labels = col_headers
    else:
        col_labels = [
            f"{g} (avg {len(selected_groups[g])})" if len(selected_groups[g]) > 1 else g
            for g in col_headers
        ]
    lines = [
        "| Rail | " + " | ".join(col_labels) + " |",
        "|------|" + "|".join(["------"] * len(col_labels)) + "|",
    ]
    for rail in display_rails:
        if rail not in table_data:
            continue
        cells = []
        for c in col_headers:
            v = table_data[rail].get(c, "")
            cells.append(f"{v:.2f}" if isinstance(v, float) else str(v))
        lines.append(f"| {rail} | " + " | ".join(cells) + " |")
    table_md = "\n".join(lines)

    return {
        "success": True,
        "table": table_md,
        "rails_shown": display_rails,
        "kpi_groups_shown": list(selected_groups.keys()),
        "all_kpi_groups": all_group_names,
        "iteration_notes": iteration_notes,
        "matrix_path": str(csv_final),
        "message": (
            f"Showing {len(display_rails)} rail(s) × {len(selected_groups)} KPI group(s). "
            + (f"Averaged: {'; '.join(iteration_notes)}. " if iteration_notes else "")
            + f"All KPI groups: {all_group_names}. "
            "Use rails=[...] and/or kpis=[...] to filter, show_individual_runs=True to see each run."
        ),
    }
