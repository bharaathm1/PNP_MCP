"""
SocWatch Tools - MCP tools for parsing Intel SocWatch CSV output files.

Two entry-point MCP tools:
  find_socwatch_files  - Discover and copy SocWatch CSVs from a result folder tree.
  parse_socwatch_data  - Parse the copied CSVs and produce Excel + Markdown summary.

All other functions are internal helpers (no @mcp.tool decorator).
"""

import os
import re
import sys
import json
import shutil
import logging
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Annotated, Dict, List, Any, Optional, Tuple, Set
from collections import Counter

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pydantic import Field
from concurrent.futures import ThreadPoolExecutor

from app import mcp
from utils.decorators import embed_if_large, async_tool

logging.basicConfig(level=logging.DEBUG, filename="debug_socwatch.log", filemode="w")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Regex patterns (shared across parsers)
# ---------------------------------------------------------------------------
SOC_NAME_PATTERN = re.compile(r"^(p_soc|soc)$", re.IGNORECASE)
TOKEN_PATTERN = re.compile(r"[A-Za-z_][A-Za-z0-9_.]*")
SOC_FALLBACK_KEYWORDS = ["soc", "core", "sa", "vnn", "io", "vdd", "prim"]
PLATFORM_FALLBACK_KEYWORDS = [
    "memory", "wlan", "wifi", "camera", "display", "edp",
    "backlight", "panel", "disp", "storage", "ssd", "audio"
]

# Global log-file path (one file per process lifetime, reset via reset_log_session)
_log_file_path = None


# === _smart_read_text ===
# Hardcoded utf-8 opens fail on those files.  This helper detects the BOM and
# picks the correct codec automatically, with latin-1 as a final no-fail fallback.
# ---------------------------------------------------------------------------
def _smart_read_text(path: str) -> str:
    """Read a text file, auto-detecting UTF-16 / UTF-8-BOM / UTF-8 / latin-1 encoding."""
    with open(path, 'rb') as _f:
        raw = _f.read()
    # BOM detection
    if raw.startswith(b'\xff\xfe\x00\x00') or raw.startswith(b'\x00\x00\xfe\xff'):
        return raw.decode('utf-32')
    if raw.startswith(b'\xff\xfe'):
        return raw.decode('utf-16-le').lstrip('\ufeff')
    if raw.startswith(b'\xfe\xff'):
        return raw.decode('utf-16-be').lstrip('\ufeff')
    if raw.startswith(b'\xef\xbb\xbf'):
        return raw[3:].decode('utf-8')
    # No BOM — try UTF-8, fall back to latin-1
    try:
        return raw.decode('utf-8')
    except UnicodeDecodeError:
        return raw.decode('latin-1')

# === safe_print + reset_log_session ===
def safe_print(text):
    """
    Print text with Unicode character replacement for compatibility and log to file.
    Creates one log file per session in the project logs/ folder with timestamp.
    """
    global _log_file_path
    
    # Initialize log file path once per session
    if _log_file_path is None:
        # Create log directory if it doesn't exist
        _LOG_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "logs")
        log_dir = _LOG_DIR
        os.makedirs(log_dir, exist_ok=True)
        
        # Generate log filename with current date and time
        current_time = datetime.now()
        log_filename = f"PowerSocwatchCompiler_Log_{current_time.strftime('%Y%m%d_%H%M%S')}.txt"
        _log_file_path = os.path.join(log_dir, log_filename)
        
        # Write session start header
        try:
            with open(_log_file_path, 'w', encoding='utf-8') as log_file:
                log_file.write(f"=== PowerSocwatch Compiler Log Session Started ===\n")
                log_file.write(f"Session Start Time: {current_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                log_file.write(f"{'='*60}\n\n")
        except Exception as e:
            print(f"Warning: Could not initialize log file: {str(e)}", file=sys.stderr)
            _log_file_path = None  # Reset to prevent further logging attempts
    
    # Process text — write to stderr only (stdout = JSON-RPC wire in stdio MCP)
    try:
        safe_text = text
        print(text, file=sys.stderr)
    except UnicodeEncodeError:
        # Replace problematic Unicode characters
        safe_text = text.replace('\u2713', 'OK').replace('\u2717', 'FAIL')
        print(safe_text.encode('ascii', 'ignore').decode('ascii'), file=sys.stderr)
    
    # Write to log file if path is available
    if _log_file_path:
        try:
            with open(_log_file_path, 'a', encoding='utf-8') as log_file:
                # Add timestamp for each log entry
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                log_file.write(f"[{timestamp}] {safe_text}\n")
        except Exception as e:
            # If logging fails, write to stderr — never stdout
            print(f"Warning: Could not write to log file: {str(e)}", file=sys.stderr)

def reset_log_session():
    """
    Reset the log session to create a new log file.
    Call this if you want to start a new log file within the same program execution.
    """
    global _log_file_path
    _log_file_path = None
    



# === save_to_json ===
def save_to_json(data: Dict[str, Any], output_file: str) -> None:
    """
    Save the parsed data to a JSON file.
    
    Args:
        data: Parsed data dictionary
        output_file: Output JSON file path
    """
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# === excel_helper_functions ===
def parse_content_to_excel_format(content: str) -> List[List]:
    """
    Parse the content string into Excel-ready format.
    Handles both tables and text content.
    """
    safe_print("Parsing content to Excel format...")
    
    lines = content.strip().split('\n')
    excel_data = []
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Skip empty lines
        if not line:
            excel_data.append([])
            i += 1
            continue
        
        # Check if this line looks like a table header (contains tabs or multiple columns)
        if '\t' in line or is_table_header(line):
            # Parse table
            table_data = parse_table_from_lines(lines, i)
            excel_data.extend(table_data['rows'])
            i = table_data['next_index']
        else:
            # Regular text line
            excel_data.append([line])
            i += 1
    
    safe_print(f"Parsed {len(excel_data)} rows from content")
    return excel_data


def is_table_header(line: str) -> bool:
    """
    Check if a line looks like a table header.
    """
    # Common table header indicators
    table_indicators = [
        'Metric\t',
        'Section\t',
        'Observations\t',
        'Comments\t',
        'Value\t',
        'Change\t',
        'Impact\t'
    ]
    
    # Check for tab-separated values
    if '\t' in line:
        return True
    
    # Check for common table headers
    for indicator in table_indicators:
        if indicator in line:
            return True
    
    # Check if line has multiple words that could be column headers
    words = line.split()
    if len(words) >= 2 and any(word in ['Metric', 'Section', 'Observations', 'Comments', 'Value', 'Change', 'Impact'] for word in words):
        return True
    
    return False


def parse_table_from_lines(lines: List[str], start_index: int) -> Dict:
    """
    Parse a table starting from the given index.
    Returns the table rows and the next index to continue parsing.
    """
    table_rows = []
    current_index = start_index
    
    # Parse header
    header_line = lines[current_index].strip()
    if '\t' in header_line:
        # Tab-separated
        header_cols = [col.strip() for col in header_line.split('\t')]
    else:
        # Space-separated (try to split intelligently)
        header_cols = smart_split_line(header_line)
    
    table_rows.append(header_cols)
    current_index += 1
    
    # Parse data rows
    while current_index < len(lines):
        line = lines[current_index].strip()
        
        # Stop if we hit an empty line or non-table content
        if not line:
            break
        
        # Check if this looks like a table row
        if '\t' in line:
            # Tab-separated
            row_cols = [col.strip() for col in line.split('\t')]
            table_rows.append(row_cols)
        elif looks_like_table_row(line, len(header_cols)):
            # Try to parse as table row
            row_cols = smart_split_line(line)
            # Pad or trim to match header length
            while len(row_cols) < len(header_cols):
                row_cols.append('')
            if len(row_cols) > len(header_cols):
                # Merge extra columns into the last column
                merged_last = ' '.join(row_cols[len(header_cols)-1:])
                row_cols = row_cols[:len(header_cols)-1] + [merged_last]
            table_rows.append(row_cols)
        else:
            # Not a table row, stop parsing table
            break
        
        current_index += 1
    
    return {
        'rows': table_rows,
        'next_index': current_index
    }


def smart_split_line(line: str) -> List[str]:
    """
    Intelligently split a line into columns.
    Handles various formats like space-separated, pipe-separated, etc.
    """
    # Try different separators
    if '|' in line:
        # Pipe-separated
        cols = [col.strip() for col in line.split('|') if col.strip()]
    elif '\t' in line:
        # Tab-separated
        cols = [col.strip() for col in line.split('\t')]
    else:
        # Space-separated - be more intelligent about this
        # Look for patterns like "word word    word word    word"
        import re
        # Split on multiple spaces (2 or more)
        cols = re.split(r'  +', line.strip())
        if len(cols) == 1:
            # Fallback to single space split
            cols = line.strip().split()
    
    return [col.strip() for col in cols if col.strip()]


def looks_like_table_row(line: str, expected_cols: int) -> bool:
    """
    Check if a line looks like it could be a table row.
    """
    # If it has tabs, it's likely a table row
    if '\t' in line:
        return True
    
    # If it has pipes, it's likely a table row
    if '|' in line:
        return True
    
    # Check if splitting gives reasonable number of columns
    words = line.split()
    if len(words) >= 2 and len(words) <= expected_cols * 2:  # Allow some flexibility
        return True
    
    # Check for numeric patterns (common in data tables)
    import re
    if re.search(r'\d+\.?\d*', line):
        return True
    
    return False


def write_parsed_data_to_sheet(sheet, excel_data: List[List]):
    """
    Write parsed data to Excel sheet.
    """
    safe_print(f"Writing {len(excel_data)} rows to sheet...")
    
    for row_idx, row_data in enumerate(excel_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            if value:  # Only write non-empty values
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)


def format_content_sheet(sheet):
    """
    Apply formatting to the content sheet.
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Define styles
    title_font = Font(size=14, bold=True, color="1F4E79")
    header_font = Font(size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    data_font = Font(size=10)
    comment_font = Font(size=10, italic=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply formatting
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                cell_value = str(cell.value)
                
                # Title formatting (first row or lines with "Summary", "Observations", etc.)
                if (cell.row == 1 or 
                    any(keyword in cell_value for keyword in ["Summary", "Tabulated", "Observations", "Analysis"])):
                    cell.font = title_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Table headers (look for common header words)
                elif any(header in cell_value for header in 
                        ["Metric", "Section", "Observations", "Comments", "Value", "Change", "Impact", "PSR Disabled", "PSR Enabled"]):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Comments or descriptive text
                elif (len(cell_value) > 50 or 
                      any(word in cell_value.lower() for word in ["comments:", "analysis", "results", "optimization"])):
                    cell.font = comment_font
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Data cells
                else:
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    
                    # Right-align numeric values
                    try:
                        float(cell_value.replace('%', '').replace(',', ''))
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    except ValueError:
                        pass
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set reasonable width limits
        adjusted_width = min(max(max_length + 2, 15), 80)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Set row heights for better readability
    for row in range(1, sheet.max_row + 1):
        # Taller rows for content with long text
        sheet.row_dimensions[row].height = 25


# Helper function to format content before passing to the function
def format_content_for_excel(content: str) -> str:
    """
    Helper function to format content string for better Excel parsing.
    Use this to prepare your content before calling write_summary_to_xlsx.
    """
    # Ensure tables are tab-separated for better parsing
    lines = content.split('\n')
    formatted_lines = []
    
    for line in lines:
        # Convert multiple spaces to tabs for table-like content
        if any(header in line for header in ["Metric", "Section", "Observations", "Comments", "Value"]):
            # This looks like a table header or row
            import re
            # Replace multiple spaces with tabs
            formatted_line = re.sub(r'  +', '\t', line.strip())
            formatted_lines.append(formatted_line)
        else:
            formatted_lines.append(line)
    
    return '\n'.join(formatted_lines)




# === parse_functions ===
def parse_package_c_state_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract Package C-State Summary residency percentages only.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing Package C-State Summary with residency percentages
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    # Define valid Package C-State patterns
    package_c_state_patterns = [
        r'^PC\d+$',        # PC0, PC2, PC6, PC10, etc.
    ]
    
    def is_valid_package_c_state(c_state: str) -> bool:
        """Check if the C-State is a valid Package C-State"""
        for pattern in package_c_state_patterns:
            if re.match(pattern, c_state):
                return True
        return False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Summary:",
            "C-State Summary:",
            "Package C-State Summary:",
            "Core C-State Summary:",
            "CPU Frequency Summary:",
            "Power Summary:",
            "Temperature Summary:"
        ]
        
        for indicator in section_indicators:
            if indicator in line and "Package C-State Summary: Residency" not in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header
        if "Package C-State Summary: Residency (Percentage and Time)" in line:
            current_section = "Package C-State Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 2:
                    if "C-State" in parts[0] or any("Residency" in part for part in parts):
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 2:
                    c_state = parts[0].strip()
                    
                    # Skip if c_state is empty or contains only dashes
                    if not c_state or re.match(r'^-+$', c_state):
                        continue
                    
                    # Only include valid Package C-States
                    if not is_valid_package_c_state(c_state):
                        continue
                    
                    try:
                        # Find the residency percentage column (usually the first numeric column)
                        percentage = None
                        for j in range(1, len(parts)):
                            try:
                                percentage = float(parts[j]) if parts[j] and parts[j] != '' else None
                                if percentage is not None:
                                    break
                            except ValueError:
                                continue
                        
                        if percentage is not None:
                            result[c_state] = percentage
                            data_started = True
                        
                    except ValueError:
                        # Skip lines with invalid numeric data
                        continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_package_c_state_os_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract Package C-State (OS) Summary residency percentages only.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing Package C-State (OS) Summary with residency percentages
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    # Define valid Package C-State (OS) patterns
    package_c_state_patterns = [
        r'^ACPI C\d+$'     # ACPI C0, ACPI C1, ACPI C2, ACPI C3, etc.
    ]
    
    def is_valid_package_c_state(c_state: str) -> bool:
        """Check if the C-State is a valid Package C-State (OS)"""
        for pattern in package_c_state_patterns:
            if re.match(pattern, c_state):
                return True
        return False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Summary:",
            "C-State Summary:",
            "Package C-State Summary:",
            "Core C-State Summary:",
            "CPU Frequency Summary:",
            "Power Summary:",
            "Temperature Summary:"
        ]
        
        for indicator in section_indicators:
            if indicator in line and "Package C-State (OS) Summary: Residency" not in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header
        if "Package C-State (OS) Summary: Residency (Percentage and Time)" in line:
            current_section = "Package C-State (OS) Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 2:
                    if "C-State" in parts[0] or any("Residency" in part for part in parts):
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 2:
                    c_state = parts[0].strip()
                    
                    # Skip if c_state is empty or contains only dashes
                    if not c_state or re.match(r'^-+$', c_state):
                        continue
                    
                    # Only include valid Package C-States (OS)
                    if not is_valid_package_c_state(c_state):
                        continue
                    
                    try:
                        # Find the residency percentage column (usually the first numeric column)
                        percentage = None
                        for j in range(1, len(parts)):
                            try:
                                percentage = float(parts[j]) if parts[j] and parts[j] != '' else None
                                if percentage is not None:
                                    break
                            except ValueError:
                                continue
                        
                        if percentage is not None:
                            result[c_state] = percentage
                            data_started = True
                        
                    except ValueError:
                        # Skip lines with invalid numeric data
                        continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_core_c_state_summary(file_path: str) -> Dict[str, Any]:
    """
    Parse Intel Socwatch CSV file and extract Core C-State Summary CC0 residency percentages
    for all cores/packages under columns ending with "Residency (%)".
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing CC0 residency percentages for each individual core with calculated summary
    """
    
    cc0_data = {}
    
    content = _smart_read_text(file_path)
    
    # Find the Core C-State Summary section
    core_section_start = content.find("Core C-State Summary: Residency (Percentage and Time)")
    if core_section_start == -1:
        return {
            "Core C-State Summary CC0": {},
            "calculated_Cstate_summary": 0.0
        }
    
    # Find the next section to know where this section ends
    remaining_content = content[core_section_start:]
    next_section_patterns = [
        "\nCPU Frequency Summary:",
        "\nPower Summary:", 
        "\nTemperature Summary:",
        "\nPackage C-State Summary:",
        "\nCore C-State (OS) Summary:",
        "\nSummary:"
    ]
    
    core_section_end = len(remaining_content)
    for pattern in next_section_patterns:
        pos = remaining_content.find(pattern, 100)  # Start search after 100 chars
        if pos != -1 and pos < core_section_end:
            core_section_end = pos
    
    # Extract just the Core C-State section
    core_section = remaining_content[:core_section_end]
    
    # Split into lines and process
    lines = core_section.split('\n')
    
    # Look for CC0 data and corresponding headers
    cc0_line_index = -1
    header_lines = []
    
    # Find CC0 line
    for i, line in enumerate(lines):
        stripped_line = line.strip()
        if stripped_line.startswith("CC0"):
            cc0_line_index = i
            break
    
    if cc0_line_index == -1:
        return {
            "Core C-State Summary CC0": {},
            "calculated_Cstate_summary": 0.0
        }
    
    # Look backwards from CC0 line to find headers
    for i in range(cc0_line_index - 1, max(0, cc0_line_index - 10), -1):
        line = lines[i].strip()
        if line and not re.match(r'^-+$', line.replace(',', '').replace(' ', '')):
            header_lines.insert(0, line)
    
    # Process the CC0 line
    cc0_line = lines[cc0_line_index].strip()
    cc0_parts = [part.strip() for part in cc0_line.split(',')]
    
    # Try to find core names from header lines
    core_names = []
    for header_line in header_lines:
        parts = [part.strip() for part in header_line.split(',')]
        if any("CPU/Package" in part for part in parts):
            core_names = parts[1:]  # Skip first column
            break
    
    # If we found core names and CC0 data, combine them
    if core_names and len(cc0_parts) > 1:
        for i in range(1, min(len(core_names) + 1, len(cc0_parts))):
            core_name = core_names[i-1] if i-1 < len(core_names) else f"Core_{i}"
            cc0_value = cc0_parts[i]
            
            if core_name and cc0_value and cc0_value != '':
                try:
                    percentage = float(cc0_value)
                    
                    # Clean up the core name and format properly
                    # Remove any existing "Residency (%)" or "Residency (msec)" from core_name
                    clean_core_name = core_name.replace(" Residency (%)", "").replace(" Residency (msec)", "")
                    
                    # Only include entries that originally had "Residency (%)" in the name
                    # Skip entries that had "Residency (msec)" in the name
                    if "Residency (msec)" not in core_name:
                        formatted_name = f"{clean_core_name} Residency (%)"
                        cc0_data[formatted_name] = percentage
                        
                except ValueError:
                    pass
    
    # Calculate the sum of all core percentages
    total_percentage = sum(cc0_data.values()) if cc0_data else 0.0
    
    # Return the nested structure
    result = {
        "Core C-State Summary CC0": cc0_data,
        "calculated_Cstate_summary": round(total_percentage, 2)
    }
    
    return result

def parse_cpu_pstate_summary(file_path: str) -> Dict[str, Any]:
    """
    Parse Intel Socwatch CSV file and extract CPU P-State Average Frequency data.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing CPU P-State average frequencies for each core/thread with calculated summary
        and total frequency residency from the full table
    """
    
    pstate_data = {}
    total_logical_processors = 0
    total_p_cores = 0
    total_e_cores = 0
    p_core_type = None
    e_core_type = None
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Summary:",
            "C-State Summary:",
            "Package C-State Summary:",
            "Core C-State Summary:",
            "Power Summary:",
            "Temperature Summary:",
            "================="
        ]
        
        for indicator in section_indicators:
            if indicator in line and "CPU P-State Average Frequency" not in line and "CPU P-State/Frequency Summary" not in line:
                return True
        return False
    
    def parse_pcore_ecore_residency(lines, p_core_type, e_core_type, core_type_mapping):
        """Parse P-Core and E-Core frequency residency from the CSV"""
        #print(f"\n=== DEBUG: Starting residency parsing ===")
        #print(f"P-Core type: {p_core_type}")
        #print(f"E-Core type: {e_core_type}")
        #print(f"Core type mapping: {core_type_mapping}")
        
        in_residency_section = False
        p_core_residency = 0.0
        e_core_residency = 0.0
        p_core_columns = []  # Column indices for P-Cores
        e_core_columns = []  # Column indices for E-Cores
        header_processed = False
        
        for line in lines:
            line = line.strip()
            
            # Check if we're entering the residency section
            if "CPU P-State/Frequency Summary: Residency (Percentage and Time)" in line:
                in_residency_section = True
                header_processed = False
                #print(f"DEBUG: Found residency section")
                continue
            
            # Check if we're leaving the residency section
            if in_residency_section and is_new_section_start(line):
                #print(f"DEBUG: Leaving residency section")
                break
            
            # Skip separator lines with dashes
            if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
                continue
            
            if in_residency_section and line:
                try:
                    parts = [part.strip() for part in line.split(',')]
                    
                    # Look for the header line with CPU/Package_X/Core_X/Thread_X information
                    if not header_processed and len(parts) > 2:
                        # Check if this looks like the header row
                        if "CPU/Package_0/Core_" in parts[2] and "Residency (%)" in parts[2]:
                            #print(f"DEBUG: Found header row with {len(parts)} columns")
                            # Map columns to core types based on core number
                            for i, part in enumerate(parts):
                                if "CPU/Package_0/Core_" in part and "Residency (%)" in part:
                                    # Extract core number from CPU/Package_0/Core_X/Thread_X

                                    core_match = re.search(r'Core_(\d+)', part)
                                    if core_match:
                                        core_num = int(core_match.group(1))
                                        # Look up the core type from our mapping
                                        if core_num in core_type_mapping:
                                            core_type = core_type_mapping[core_num]
                                            # print(f"DEBUG: Column {i} -> Core_{core_num} -> {core_type}")
                                            if core_type == p_core_type:
                                                p_core_columns.append(i)
                                            elif core_type == e_core_type:
                                                e_core_columns.append(i)
                            
                            header_processed = True
                            #print(f"DEBUG: P-Core columns: {p_core_columns}")
                            #print(f"DEBUG: E-Core columns: {e_core_columns}")
                            continue
                    
                    # Process frequency residency data rows (P0, P1, P2, etc.)
                    if header_processed and len(parts) > 1:
                        # Check if first column is a frequency state (P0, P1, etc.)
                        if parts[0].startswith('P') and len(parts[0]) > 1:
                            try:
                                # Verify it's a valid P-state (P followed by number)
                                pstate_num = int(parts[0][1:])
                                #print(f"DEBUG: Processing P-state: {parts[0]}")
                            except ValueError:
                                continue
                            
                            # Sum residency values for P-Cores
                            p_core_sum_this_row = 0.0
                            for col_idx in p_core_columns:
                                if col_idx < len(parts) and parts[col_idx]:
                                    try:
                                        # Remove percentage sign and convert to float
                                        residency_str = parts[col_idx].replace('%', '').strip()
                                        if residency_str:
                                            residency_value = float(residency_str)
                                            p_core_residency += residency_value
                                            p_core_sum_this_row += residency_value
                                    except ValueError:
                                        continue
                            
                            # Sum residency values for E-Cores
                            e_core_sum_this_row = 0.0
                            for col_idx in e_core_columns:
                                if col_idx < len(parts) and parts[col_idx]:
                                    try:
                                        # Remove percentage sign and convert to float
                                        residency_str = parts[col_idx].replace('%', '').strip()
                                        if residency_str:
                                            residency_value = float(residency_str)
                                            e_core_residency += residency_value
                                            e_core_sum_this_row += residency_value
                                    except ValueError:
                                        continue
                            
                            #print(f"DEBUG: {parts[0]} - P-Cores: {p_core_sum_this_row:.2f}, E-Cores: {e_core_sum_this_row:.2f}")
                
                except Exception as e:
                    continue
        
        #print(f"DEBUG: Final totals - P-Core: {p_core_residency:.2f}, E-Core: {e_core_residency:.2f}")
        return p_core_residency, e_core_residency
    
    # First pass: Extract system information
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Extract total logical processors
        if "Total # of logical processors" in line:
            try:
                # Look for the number after the colon or equals sign
                if ':' in line:
                    total_logical_processors = int(line.split(':')[1].strip())
                elif '=' in line:
                    total_logical_processors = int(line.split('=')[1].strip())
            except (ValueError, IndexError):
                pass
    
    # Second pass: Extract P-Core and E-Core information from CPU native model section
    in_cpu_native_model = False
    core_types = {}  # Dictionary to count different core types
    core_type_mapping = {}  # Map core number to core type
    total_logical_cores_found = 0
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Check if we're entering CPU native model section
        if "CPU native model:" in line:
            in_cpu_native_model = True
            continue
        
        # Check if we're leaving CPU native model section
        if in_cpu_native_model and (line.startswith("===") or 
                                   any(indicator in line for indicator in [
                                       "Summary:", "C-State Summary:", "Package C-State Summary:",
                                       "Core C-State Summary:", "Power Summary:", "Temperature Summary:"
                                   ])):
            break
        
        # Process CPU native model entries
        if in_cpu_native_model and line.startswith("Package_0/Core_"):
            total_logical_cores_found += 1
            try:
                # Extract core number and core type
                if '=' in line:
                    core_part = line.split('=')[0].strip()
                    core_name = line.split('=')[1].strip()
                    
                    # Extract core number
                    core_match = re.search(r'Core_(\d+)', core_part)
                    if core_match:
                        core_num = int(core_match.group(1))
                        core_type_mapping[core_num] = core_name
                    
                    # Count occurrences of each core type
                    if core_name in core_types:
                        core_types[core_name] += 1
                    else:
                        core_types[core_name] = 1
            except (IndexError, ValueError):
                pass
    
    # Determine P-Cores and E-Cores based on core types
    if core_types:
        # Sort core types by count (ascending order)
        sorted_core_types = sorted(core_types.items(), key=lambda x: x[1])
        
        # The core type with fewer instances is likely P-Core
        # The core type with more instances is likely E-Core
        if len(sorted_core_types) >= 2:
            p_core_type = sorted_core_types[0][0]  # Less numerous type
            total_p_cores = sorted_core_types[0][1]
            e_core_type = sorted_core_types[1][0]  # More numerous type
            total_e_cores = sorted_core_types[1][1]
        elif len(sorted_core_types) == 1:
            # Only one core type found
            core_type, count = sorted_core_types[0]
            if count <= 8:  # Typical P-Core count threshold
                p_core_type = core_type
                total_p_cores = count
                total_e_cores = 0
                e_core_type = None
            else:
                p_core_type = None
                total_p_cores = 0
                e_core_type = core_type
                total_e_cores = count
    else:
        total_p_cores = 0
        total_e_cores = 0
        p_core_type = None
        e_core_type = None
    
    # Third pass: Extract P-State frequency data - FIXED VERSION
    in_pstate_avg_section = False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check for our target section header
        if "CPU P-State Average Frequency (excluding CPU idle time)" in line:
            in_pstate_avg_section = True
            header_found = False
            data_started = False
            #print(f"DEBUG: Found CPU P-State Average Frequency section at line {i+1}")
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if in_pstate_avg_section and is_new_section_start(line):
            #print(f"DEBUG: Leaving CPU P-State Average Frequency section at line {i+1}: {line}")
            break
        
        # Also check for specific section endings that might not match is_new_section_start
        if in_pstate_avg_section and ("CPU P-State/Frequency Summary" in line or 
                                     "Total Samples Received" in line or
                                     line.startswith("=====")):
            #print(f"DEBUG: Ending CPU P-State Average Frequency section at line {i+1}: {line}")
            break
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if in_pstate_avg_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 2:
                    if "CPU ID" in parts[0] and "Average (MHz)" in parts[1]:
                        header_found = True
                        #print(f"DEBUG: Found header line at {i+1}: {line}")
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 2:
                    cpu_id = parts[0].strip()
                    avg_freq = parts[1].strip()
                    
                    # Skip if cpu_id is empty or contains only dashes
                    if not cpu_id or re.match(r'^-+$', cpu_id):
                        continue
                    
                    # Check if this is a valid CPU/Core/Thread entry
                    if cpu_id.startswith("CPU/Package_") and ("Core_" in cpu_id or "Thread_" in cpu_id):
                        try:
                            frequency = float(avg_freq) if avg_freq and avg_freq != '' else 0.0
                            
                            # Store the CPU P-State data
                            pstate_data[cpu_id] = frequency
                            data_started = True
                            #print(f"DEBUG: Added {cpu_id}: {frequency} MHz")
                            
                        except ValueError:
                            # Skip lines with invalid numeric data
                            #print(f"DEBUG: Skipping invalid frequency data: {avg_freq}")
                            continue
                        
            except Exception as e:
                # Skip problematic lines
                #print(f"DEBUG: Error processing line {i+1}: {e}")
                continue
    
    #print(f"DEBUG: Total P-State entries found: {len(pstate_data)}")
    
    # Calculate P-Core and E-Core frequency residency
    p_core_residency, e_core_residency = parse_pcore_ecore_residency(
        lines, p_core_type, e_core_type, core_type_mapping
    )
    
    # Calculate summary statistics
    if pstate_data:
        frequencies = list(pstate_data.values())
        avg_frequency = sum(frequencies) / len(frequencies)
        max_frequency = max(frequencies)
        min_frequency = min(frequencies)
        
        calculated_summary = {
            "average_frequency_mhz": round(avg_frequency, 2),
            "max_frequency_mhz": max_frequency,
            "min_frequency_mhz": min_frequency,
            "total_cores_threads": len(pstate_data),
            "total_logical_processors": total_logical_processors,
            "total_P_Cores": total_p_cores,
            "total_E_Cores": total_e_cores
        }
    else:
        calculated_summary = {
            "average_frequency_mhz": 0.0,
            "max_frequency_mhz": 0.0,
            "min_frequency_mhz": 0.0,
            "total_cores_threads": 0,
            "total_logical_processors": total_logical_processors,
            "total_P_Cores": total_p_cores,
            "total_E_Cores": total_e_cores
        }
    
    # Call parse_cpu_pstate_fulltable to get the total frequency residency
    fulltable_data = parse_cpu_pstate_fulltable(lines)
    if fulltable_data is not None:
        calculated_totalfreq_residency = fulltable_data.get("calculated_totalfreq_residency", 0.0)
    else:
        calculated_totalfreq_residency = 0.0
    
    # Return the nested structure with all frequency residency data
    result = {
        "CPU P-State Average Frequencies": pstate_data,
        "calculated_pstate_summary": calculated_summary,
        "calculated_totalfreq_residency": calculated_totalfreq_residency,
        "calculated_PCore_totalfreq_residency": round(p_core_residency, 2),
        "calculated_ECore_totalfreq_residency": round(e_core_residency, 2)
    }
    
    return result
    
def parse_cpu_pstate_fulltable(lines):
    """
    Parse CPU P-State/Frequency Summary: Residency (Percentage and Time) table
    Only consider columns ending with "Residency (%)" and exclude "CPU Idle" row
    """
    # Find the CPU P-State Full Table section
    start_line = None
    for i, line in enumerate(lines):
        if 'CPU P-State/Frequency Summary: Residency (Percentage and Time)' in line:
            start_line = i
            break
    
    if start_line is None:
        # Return a default structure instead of None
        return {
            'frequency_residency_per_core': {},
            'core_totals': {},
            'calculated_totalfreq_residency': 0.0,
            'total_cores': 0
        }
    
    # Initialize data structures
    frequency_residency_per_core = {}
    core_headers = []
    header_found = False
    
    # Process lines starting from the section header
    i = start_line + 1
    while i < len(lines):
        line = lines[i].strip()
        
        # Check if we've reached the end of this section
        if (line.startswith('CPU P-State/Frequency Summary: Total Samples Received') or
            (line and not line.startswith(',') and not line.startswith('P') and 
             not line.startswith('-') and not line.startswith('CPU Idle') and
             'Residency' not in line and len(line.split(',')) < 10)):
            break
        
        # Skip empty lines
        if not line or line.replace(',', '').strip() == '':
            i += 1
            continue
        
        # Parse CSV
        parts = [part.strip() for part in line.split(',')]
        
        # Look for header line
        if not header_found and len(parts) > 2 and 'Residency (%)' in line:
            # Extract core headers (only those ending with "Residency (%)")
            for j, part in enumerate(parts[2:], 2):  # Skip first two columns
                if part.endswith('Residency (%)'):
                    core_name = part.replace('CPU/Package_0/', '').replace(' Residency (%)', '')
                    core_headers.append(core_name)
                    frequency_residency_per_core[core_name] = {}
            
            header_found = True
            i += 1
            continue
        
        # Skip separator lines
        if line.startswith('---') or all(c in '-,' for c in line.replace(' ', '')):
            i += 1
            continue
        
        # Process data lines
        if header_found and len(parts) >= 2:
            pstate_label = parts[0].strip()  # P11, P12, etc.
            freq_range = parts[1].strip()    # Actual frequency range like "2201 -- 2300"
            
            # Skip empty or invalid frequency ranges
            if not freq_range or freq_range == '-' or freq_range == '':
                i += 1
                continue
            
            # IMPORTANT: Skip "CPU Idle" row as per requirements
            if pstate_label == 'CPU Idle':
                i += 1
                continue
            
            # Extract residency values (only from percentage columns)
            residency_values = parts[2:2+len(core_headers)]  # Only take percentage columns
            
            # Store residency data for each core
            for core_idx, core_name in enumerate(core_headers):
                if core_idx < len(residency_values):
                    try:
                        residency = float(residency_values[core_idx])
                        if residency > 0:  # Only store non-zero values
                            frequency_residency_per_core[core_name][freq_range] = residency
                    except (ValueError, IndexError):
                        pass  # Skip invalid values
        
        i += 1
    
    # Calculate totals per core (excluding CPU Idle)
    core_totals = {}
    total_frequency_residency = 0
    
    for core_name, frequencies in frequency_residency_per_core.items():
        core_total = sum(frequencies.values())
        core_totals[core_name] = core_total
        total_frequency_residency += core_total
    
    result = {
        'frequency_residency_per_core': frequency_residency_per_core,
        'core_totals': core_totals,
        'calculated_totalfreq_residency': total_frequency_residency,
        'total_cores': len(core_totals)
    }
    
    return result

def parse_media_p_state_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract Media P-State Summary frequency distribution.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing Media P-State frequency distribution percentages
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Media P-State Summary - Sampled: Counts",
            "Media P-State Summary - Sampled: Total Samples",
            "Media C-State Residency",
            "================="
        ]
        
        for indicator in section_indicators:
            if indicator in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header - exact match
        if "Media P-State Summary - Sampled: Approximated Residency (Percentage)" in line:
            current_section = "Media P-State Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 2:
                    if "Frequency" in parts[0] and "MEDIA" in parts[1]:
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 2:
                    frequency_str = parts[0].strip()
                    percentage_str = parts[1].strip()
                    
                    # Skip if frequency is empty or contains only dashes
                    if not frequency_str or re.match(r'^-+$', frequency_str):
                        continue
                    
                    try:
                        # Convert frequency from float to int (e.g., "100.0" -> 100)
                        frequency_float = float(frequency_str)
                        frequency = int(frequency_float)  # Convert to int for clean display
                        percentage = float(percentage_str)
                        
                        # Format frequency key
                        frequency_key = f"{frequency} MHz"
                        result[frequency_key] = percentage
                        data_started = True
                        
                    except ValueError:
                        # Skip lines with invalid numeric data
                        continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_media_c_state_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract Media C-State Residency Summary.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing Media C-State residency percentages
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Summary:",
            "C-State Summary:",
            "Package C-State Summary:",
            "Core C-State Summary:",
            "CPU Frequency Summary:",
            "Power Summary:",
            "Temperature Summary:",
            "Media P-State Summary",  # This will catch any Media P-State section
            "P-State Summary"
        ]
        
        for indicator in section_indicators:
            if indicator in line and "Media C-State Residency Summary" not in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header - exact match
        if "Media C-State Residency Summary: Residency (Percentage and Time)" in line:
            current_section = "Media C-State Residency Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 2:
                    if "C-State" in parts[0] or any("Residency" in part for part in parts):
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 2:
                    c_state = parts[0].strip()
                    
                    # Skip if c_state is empty or contains only dashes
                    if not c_state or re.match(r'^-+$', c_state):
                        continue
                    
                    # Only include entries that start with "Media" and are valid C-States
                    if c_state.startswith("Media") and ("C0" in c_state or "C6" in c_state or "C1" in c_state or "C3" in c_state):
                        try:
                            # Find the residency percentage column (usually the first numeric column)
                            percentage = None
                            for j in range(1, len(parts)):
                                try:
                                    percentage = float(parts[j]) if parts[j] and parts[j] != '' else None
                                    if percentage is not None:
                                        break
                                except ValueError:
                                    continue
                            
                            if percentage is not None:
                                # Clean up the c_state name
                                clean_c_state = c_state.rstrip(',').strip()
                                result[clean_c_state] = percentage
                                data_started = True
                            
                        except ValueError:
                            # Skip lines with invalid numeric data
                            continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_gfx_c_state_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract Integrated Graphics C-State Summary residency percentages.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing GFX C-State residency percentages
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    # Define valid GFX C-State patterns
    gfx_c_state_patterns = [
        r'^RC\d+$',        # RC0, RC6, etc.
    ]
    
    def is_valid_gfx_c_state(c_state: str) -> bool:
        """Check if the C-State is a valid GFX C-State"""
        for pattern in gfx_c_state_patterns:
            if re.match(pattern, c_state):
                return True
        return False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Summary:",
            "C-State Summary:",
            "P-State/Frequency Summary:",
            "Package C-State Summary:",
            "Core C-State Summary:",
            "CPU Frequency Summary:",
            "Power Summary:",
            "Temperature Summary:",
            "Media P-State Summary:",
            "Media C-State Summary:",
            "Total Samples Received"  # This catches the end of the section
        ]
        
        for indicator in section_indicators:
            if indicator in line and "Integrated Graphics C-State" not in line and "Residency (Percentage and Time)" not in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header - handle extra spaces
        if "Integrated Graphics C-State" in line and "Summary: Residency (Percentage and Time)" in line:
            current_section = "Integrated Graphics C-State Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 2:
                    if "C-State" in parts[0] or any("iGPU" in part or "Graphics" in part for part in parts):
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 2:
                    c_state = parts[0].strip()
                    
                    # Skip if c_state is empty or contains only dashes
                    if not c_state or re.match(r'^-+$', c_state):
                        continue
                    
                    # Only include valid GFX C-States
                    if not is_valid_gfx_c_state(c_state):
                        continue
                    
                    try:
                        # Find the residency percentage column (usually the first numeric column)
                        percentage = None
                        for j in range(1, len(parts)):
                            try:
                                percentage = float(parts[j]) if parts[j] and parts[j] != '' else None
                                if percentage is not None:
                                    break
                            except ValueError:
                                continue
                        
                        if percentage is not None:
                            result[c_state] = percentage
                            data_started = True
                        
                    except ValueError:
                        # Skip lines with invalid numeric data
                        continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_gfx_p_state_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract Integrated Graphics P-State/Frequency Summary.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing GFX P-State frequency distribution percentages
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Integrated Graphics P-State/Frequency Summary - Sampled: Counts",  # Stop before Counts section
            "Summary:",
            "C-State Summary:",
            "Package C-State Summary:",
            "Core C-State Summary:",
            "CPU Frequency Summary:",
            "Power Summary:",
            "Temperature Summary:",
            "Media P-State Summary:",
            "Media C-State Summary:",
            "Integrated Graphics C-State"
        ]
        
        for indicator in section_indicators:
            if indicator in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header
        if "Integrated Graphics P-State/Frequency Summary - Sampled: Approximated Residency (Percentage)" in line:
            current_section = "Integrated Graphics P-State Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 2:
                    if "Frequency" in parts[0] and "IGFX" in parts[1]:
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 2:
                    frequency_str = parts[0].strip()
                    percentage_str = parts[1].strip()
                    
                    # Skip if frequency is empty or contains only dashes
                    if not frequency_str or re.match(r'^-+$', frequency_str):
                        continue
                    
                    try:
                        # Convert frequency from float to int (e.g., "400.0" -> 400)
                        frequency_float = float(frequency_str)
                        frequency = int(frequency_float)  # Convert to int for clean display
                        percentage = float(percentage_str)
                        
                        # Format frequency key
                        frequency_key = f"{frequency} MHz"
                        result[frequency_key] = percentage
                        data_started = True
                        
                    except ValueError:
                        # Skip lines with invalid numeric data
                        continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_memss_p_state_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract Memory Subsystem (MEMSS) P-State Summary frequency distribution.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing MEMSS P-State frequency distribution percentages
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Memory Subsystem (MEMSS) P-State Summary - Sampled: Counts",
            "Memory Subsystem (MEMSS) P-State Summary - Sampled: Total Samples",
            "Summary:",
            "================="
        ]
        
        for indicator in section_indicators:
            if indicator in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header
        if "Memory Subsystem (MEMSS) P-State Summary - Sampled: Approximated Residency (Percentage)" in line:
            current_section = "Memory Subsystem (MEMSS) P-State Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 2:
                    if "Frequency" in parts[0] and "MEMSS" in parts[1]:
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 2:
                    frequency_str = parts[0].strip()
                    percentage_str = parts[1].strip()
                    
                    # Skip if frequency is empty or contains only dashes
                    if not frequency_str or re.match(r'^-+$', frequency_str):
                        continue
                    
                    try:
                        # Convert frequency from float to int (e.g., "594.0" -> 594)
                        frequency_float = float(frequency_str)
                        frequency = int(frequency_float)  # Convert to int for clean display
                        percentage = float(percentage_str)
                        
                        # Format frequency key
                        frequency_key = f"{frequency} MHz"
                        result[frequency_key] = percentage
                        data_started = True
                        
                    except ValueError:
                        # Skip lines with invalid numeric data
                        continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_ipu_c_state_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract IPU C-State Residency Summary.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing IPU C-State residency percentages for both IS and PS subsystems
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Image Processing Unit (IPU) C-State Residency Summary: Total Samples Received",
            "Summary:",
            "C-State Summary:",
            "P-State Summary:",
            "======================="
        ]
        
        for indicator in section_indicators:
            if indicator in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header
        if "Image Processing Unit (IPU) C-State Residency Summary: Residency (Percentage and Time)" in line:
            current_section = "IPU C-State Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 3:
                    if "State" in parts[0] and "IS Residency" in parts[1] and "PS Residency" in parts[2]:
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 3:
                    state = parts[0].strip()
                    is_residency = parts[1].strip()
                    ps_residency = parts[2].strip()
                    
                    # Check if this is a valid C-State (C0, C6, etc.)
                    if state and re.match(r'^C\d+$', state):
                        try:
                            is_percentage = float(is_residency) if is_residency and is_residency != '' else None
                            ps_percentage = float(ps_residency) if ps_residency and ps_residency != '' else None
                            
                            # Store both IS and PS residencies
                            if is_percentage is not None:
                                result[f"IPU {state} IS"] = is_percentage
                            
                            if ps_percentage is not None:
                                result[f"IPU {state} PS"] = ps_percentage
                            
                            data_started = True
                            
                        except ValueError:
                            # Skip lines with invalid numeric data
                            continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_ipu_p_state_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract IPU P-State Summary frequency distribution.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing IPU P-State frequency distribution percentages for both IS and PS subsystems
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Image Processing Unit (IPU) P-State Summary - Sampled: Counts",
            "Image Processing Unit (IPU) P-State Summary - Sampled: Total Samples",
            "Summary:",
            "================="
        ]
        
        for indicator in section_indicators:
            if indicator in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header
        if "Image Processing Unit (IPU) P-State Summary - Sampled: Approximated Residency (Percentage)" in line:
            current_section = "IPU P-State Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 3:
                    if "Frequency" in parts[0] and "IS-FREQ" in parts[1] and "PS-FREQ" in parts[2]:
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 3:
                    frequency_str = parts[0].strip()
                    is_freq_percentage = parts[1].strip()
                    ps_freq_percentage = parts[2].strip()
                    
                    # Check if this is a valid frequency
                    if frequency_str and frequency_str != '':
                        try:
                            # Convert frequency from string to int (e.g., "0" -> 0, "400" -> 400)
                            frequency = int(float(frequency_str))
                            
                            # Remove % sign and convert to float
                            is_percentage = None
                            ps_percentage = None
                            
                            if is_freq_percentage and is_freq_percentage != '':
                                # Remove % sign if present and convert to float
                                is_clean = is_freq_percentage.replace('%', '').strip()
                                if is_clean:
                                    is_percentage = float(is_clean)
                            
                            if ps_freq_percentage and ps_freq_percentage != '':
                                # Remove % sign if present and convert to float
                                ps_clean = ps_freq_percentage.replace('%', '').strip()
                                if ps_clean:
                                    ps_percentage = float(ps_clean)
                            
                            # Store both IS and PS frequencies
                            if is_percentage is not None:
                                result[f"IPU {frequency} MHz IS"] = is_percentage
                            
                            if ps_percentage is not None:
                                result[f"IPU {frequency} MHz PS"] = ps_percentage
                            
                            data_started = True
                            
                        except ValueError:
                            # Skip lines with invalid numeric data
                            continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_npu_p_state_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract NPU P-State Summary frequency distribution.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing NPU P-State frequency distribution percentages
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Neural Processing Unit (NPU) P-State Summary - Sampled: Counts",
            "Neural Processing Unit (NPU) P-State Summary - Sampled: Total Samples",
            "Summary:",
            "================="
        ]
        
        for indicator in section_indicators:
            if indicator in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header
        if "Neural Processing Unit (NPU) P-State Summary - Sampled: Approximated Residency (Percentage)" in line:
            current_section = "NPU P-State Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 2:
                    if "Frequency" in parts[0] and "NPU" in parts[1]:
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 2:
                    frequency_str = parts[0].strip()
                    npu_percentage = parts[1].strip()
                    
                    # Skip if frequency is empty or contains only dashes
                    if not frequency_str or re.match(r'^-+$', frequency_str):
                        continue
                    
                    try:
                        # Convert frequency from string to int
                        frequency = int(float(frequency_str))
                        
                        # Remove % sign if present and convert to float
                        if npu_percentage and npu_percentage != '':
                            npu_clean = npu_percentage.replace('%', '').strip()
                            if npu_clean:
                                percentage = float(npu_clean)
                                
                                # Format frequency key
                                frequency_key = f"NPU {frequency} MHz"
                                result[frequency_key] = percentage
                                data_started = True
                        
                    except ValueError:
                        # Skip lines with invalid numeric data
                        continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_npu_d_state_summary(file_path: str) -> Dict[str, float]:
    """
    Parse Intel Socwatch CSV file and extract NPU D-State Residency Summary.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing NPU D-State residency percentages
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Neural Processing Unit (NPU) D-State Residency Summary: Total Samples Received",
            "Neural Processing Unit (NPU) D-State Entrance Count",
            "Summary:",
            "================="
        ]
        
        for indicator in section_indicators:
            if indicator in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line):
            break
            
        # Check for our target section header
        if "Neural Processing Unit (NPU) D-State Residency Summary: Residency (Percentage and Time)" in line:
            current_section = "NPU D-State Summary"
            header_found = False
            data_started = False
            continue
        
        # Skip separator lines with dashes
        if re.match(r'^-+', line.replace(',', '').replace(' ', '')):
            continue
            
        # Process CSV data only if we're in the target section
        if current_section and line:
            try:
                # Split by comma and clean up the data
                parts = [part.strip() for part in line.split(',')]
                
                # Check if this is a header line
                if not header_found and len(parts) >= 2:
                    if "State" in parts[0] and "Residency (%)" in parts[1]:
                        header_found = True
                        continue
                
                # Process data lines
                if header_found and len(parts) >= 2:
                    state = parts[0].strip()
                    residency_percentage = parts[1].strip()
                    
                    # Skip if state is empty or contains only dashes
                    if not state or re.match(r'^-+$', state):
                        continue
                    
                    try:
                        # Convert residency percentage to float
                        if residency_percentage and residency_percentage != '':
                            percentage = float(residency_percentage)
                            
                            # Clean up the state name and store
                            clean_state = state.rstrip(',').strip()
                            result[f"NPU {clean_state}"] = percentage
                            data_started = True
                        
                    except ValueError:
                        # Skip lines with invalid numeric data
                        continue
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def parse_thread_wakeups_os_summary(file_path: str) -> Dict[str, Any]:
    """
    Parse Intel Socwatch CSV file and extract Thread Wakeups (OS) summary data.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing Overall Platform Activity and process-wise CPU % summary
    """
    
    result = {
        "Overall Platform Activity": 0.0,
        "process_wise_summary": {}
    }
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    current_section = None
    header_found = False
    data_started = False
    
    def is_new_section_start(line: str) -> bool:
        """Check if line indicates start of a new section"""
        section_indicators = [
            "Summary:",
            "C-State Summary:",
            "P-State Summary:",
            "Bandwidth Summary:",
            "Temperature Summary:",
            "Power Summary:",
            "Legend:",
            "================="
        ]
        
        for indicator in section_indicators:
            if indicator in line and "Processes by Platform" not in line:
                return True
        return False
    
    for i, line in enumerate(lines):
        line_stripped = line.strip()
        
        # Skip empty lines
        if not line_stripped:
            continue
        
        # Check if we've hit a new section (stop parsing current section)
        if current_section and data_started and is_new_section_start(line_stripped):
            break
            
        # Check for our target section header
        if "Processes by Platform Busy Duration" in line_stripped and "Legend" not in line_stripped:
            current_section = "Processes by Platform Busy Duration"
            header_found = False
            data_started = False
            continue
        
        # Process data only if we're in the target section
        if current_section and line_stripped:
            try:
                # Look for the table header
                if not header_found and 'Process Name (PID)' in line_stripped and 'CPU % (Platform)' in line_stripped:
                    header_found = True
                    continue
                
                # Skip separator lines with dashes
                if line_stripped.startswith('---') or all(c in '-,' for c in line_stripped.replace(' ', '')):
                    continue
                
                # Process data lines
                if header_found:
                    # Split by comma and clean up the data
                    parts = [part.strip() for part in line_stripped.split(',')]
                    
                    if len(parts) >= 3:  # Need at least Rank, Process Name, CPU %
                        rank = parts[0].strip()
                        process_name_raw = parts[1].strip()
                        cpu_platform_raw = parts[2].strip()
                        
                        # Skip header-like lines or empty data
                        if not process_name_raw or 'Process Name' in process_name_raw:
                            continue
                        
                        # Handle "Overall Platform Activity" special case
                        if "Overall Platform Activity" in process_name_raw:
                            try:
                                result["Overall Platform Activity"] = float(cpu_platform_raw)
                                data_started = True
                            except (ValueError, TypeError):
                                pass
                        else:
                            # Extract process name (remove PID if present)
                            process_name = process_name_raw
                            
                            # Clean up process name - remove rank numbers and extract actual process name
                            if '(' in process_name and ')' in process_name:
                                # Extract process name before the PID
                                process_name = process_name.split('(')[0].strip()
                            
                            # Skip if process name is empty after cleaning or is just a number (rank)
                            if not process_name or process_name.isdigit():
                                continue
                            
                            # Parse CPU percentage
                            try:
                                cpu_percent = float(cpu_platform_raw)
                                if cpu_percent > 0:  # Only store non-zero values
                                    result["process_wise_summary"][process_name] = cpu_percent
                                    data_started = True
                            except (ValueError, TypeError):
                                pass
                        
            except Exception as e:
                # Skip problematic lines
                continue
    
    return result

def save_to_json(data: Dict[str, Any], output_file: str) -> None:
    """
    Save the parsed data to a JSON file.
    
    Args:
        data: Parsed data dictionary
        output_file: Output JSON file path
    """
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def parse_ddr_bandwidth_summary(file_path: str) -> Dict[str, Any]:
    """
    Parse Intel Socwatch CSV file and extract DDR Bandwidth Summary data including individual IPs.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing DDR bandwidth data, individual channels, and individual IP bandwidth
    """
    
    result = {
        "individual_channels": {},
        "total_bandwidth": 0.0,
        "calculated_summary": {
            "total_read_rate_mbps": 0.0,
            "total_write_rate_mbps": 0.0,
            "combined_rate_mbps": 0.0,
            "read_channels_count": 0,
            "write_channels_count": 0
        },
        "individual_IP": {}
    }
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    # Find DDR Bandwidth section
    ddr_section_start = None
    ddr_section_end = None
    
    for i, line in enumerate(lines):
        if "DDR Bandwidth Requests by Component Summary: Average Rate and Total" in line:
            ddr_section_start = i
            break
    
    if ddr_section_start is None:
        return result
    
    # Find the end of DDR section
    for i in range(ddr_section_start + 1, len(lines)):
        line = lines[i].strip()
        if line.startswith("=") or (line == "" and i > ddr_section_start + 20):
            ddr_section_end = i
            break
    
    if ddr_section_end is None:
        ddr_section_end = len(lines)
    
    # Parse DDR bandwidth data
    total_read_rate = 0.0
    total_write_rate = 0.0
    read_count = 0
    write_count = 0
    
    for i in range(ddr_section_start, ddr_section_end):
        line = lines[i].strip()
        
        # Skip header and separator lines
        if "Device, Event" in line or line.startswith("-----") or line == "":
            continue
            
        # Check if we hit the Total line
        if line.startswith("Total"):
            parts = [p.strip() for p in line.split(',')]
            if len(parts) >= 3:
                try:
                    total_bandwidth = float(parts[2])
                    result["total_bandwidth"] = total_bandwidth
                except ValueError:
                    pass
            break
        
        # Parse individual DDR channel data
        if line.startswith("DDR"):
            parts = [p.strip() for p in line.split(',')]
            if len(parts) >= 4:
                try:
                    device = parts[0]
                    event = parts[1]
                    rate_mbps = float(parts[2])
                    total_bytes = float(parts[3])
                    
                    # Store individual channel data - simplified format (event: rate)
                    result["individual_channels"][event] = rate_mbps
                    
                    # Accumulate totals for verification
                    if "READS" in event:
                        total_read_rate += rate_mbps
                        read_count += 1
                    elif "WRITES" in event:
                        total_write_rate += rate_mbps
                        write_count += 1
                    
                except (ValueError, IndexError):
                    continue
    
    # Update calculated summary
    result["calculated_summary"]["total_read_rate_mbps"] = total_read_rate
    result["calculated_summary"]["total_write_rate_mbps"] = total_write_rate
    result["calculated_summary"]["combined_rate_mbps"] = total_read_rate + total_write_rate
    result["calculated_summary"]["read_channels_count"] = read_count
    result["calculated_summary"]["write_channels_count"] = write_count
    
    # Parse other IP bandwidth sections - store individual channels directly in individual_IP
    ip_sections = [
        ("Display VC1 Bandwidth Summary: Average Rate and Total"),
        ("Neural Processing Unit (NPU) to Memory Bandwidth Summary: Average Rate and Total"),
        ("Media to Network on Chip (NoC) Bandwidth Summary: Average Rate and Total"),
        ("Image Processing Unit (IPU) to Network on Chip (NoC) Bandwidth Summary: Average Rate and Total"),
        ("CCE to Network on Chip (NoC) Bandwidth Summary: Average Rate and Total"),
        ("Network on a Chip GT Bandwidth Summary: Average Rate and Total"),
        ("Network on a Chip Die to Die Bandwidth Summary: Average Rate and Total"),
        ("Cluster1 Cores Bandwidth Summary: Average Rate and Total"),
        ("System Cache Bandwidth Summary: Average Rate and Total"),
        ("Home Agent Box (HBO) Bandwidth Summary: Average Rate and Total"),
        ("Cluster0 Cores Bandwidth Summary: Average Rate and Total"),
        ("Network on a Chip IO Bandwidth Summary: Average Rate and Total")
    ]
    
    for section_header in ip_sections:
        # Find the section
        section_start = None
        for i, line in enumerate(lines):
            if section_header in line:
                section_start = i
                break
        
        if section_start is None:
            continue
        
        # Look for data lines after the header
        for i in range(section_start + 1, min(section_start + 20, len(lines))):
            line = lines[i].strip()
            
            # Skip header and separator lines
            if "Device, Event" in line or line.startswith("-----") or line == "":
                continue
            
            # Stop at Total line
            if line.startswith("Total"):
                break
            
            # Parse individual channel data and add directly to individual_IP
            if line and not line.startswith("Total"):
                parts = [p.strip() for p in line.split(',')]
                if len(parts) >= 3:
                    try:
                        event = parts[1]
                        rate_mbps = float(parts[2])
                        
                        # Only store if it's a valid event name (not empty and not numeric)
                        if event and not event.replace('.', '').isdigit():
                            result["individual_IP"][event] = rate_mbps
                        
                    except (ValueError, IndexError):
                        continue
    
    return result

def parse_ltr_residency_summary(file_path: str) -> Dict[str, Any]:
    """
    Parse Intel Socwatch CSV file and extract PCD Platform LTR Snoop Residency data from histogram.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing LTR residency data with component count and residency sums
    """
    
    result = {
        "no_of_components": 0,
        "40us": 0.0,
        "60us": 0.0,
        "110us": 0.0,
        "150us": 0.0,
        "250us": 0.0,
        "350us": 0.0,
        "500us": 0.0,
        "750us": 0.0,
        "1ms": 0.0,
        "3ms": 0.0,
        "3ms_greater": 0.0,
        "no_req": 0.0
    }
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    # Find the LTR Histogram section
    histogram_section_start = None
    histogram_section_end = None
    
    for i, line in enumerate(lines):
        if "PCD Platform LTR Snoop Summary - Sampled: Histogram" in line:
            histogram_section_start = i
            break
    
    if histogram_section_start is None:
        return result
    
    # Find the end of histogram section (next section starts with "Total Samples Received")
    for i in range(histogram_section_start + 1, len(lines)):
        line = lines[i].strip()
        if "PCD Platform LTR Snoop Summary - Sampled: Total Samples Received" in line or line == "":
            histogram_section_end = i
            break
    
    if histogram_section_end is None:
        histogram_section_end = len(lines)
    
    # Parse histogram data
    header_found = False
    component_count = 0
    
    for i in range(histogram_section_start, histogram_section_end):
        line = lines[i].strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Skip separator lines with dashes
        if line.startswith("---") or line.replace("-", "").replace(" ", "") == "":
            continue
            
        # Look for header line
        if not header_found and "Component Name" in line and "<=40us(%)" in line:
            header_found = True
            continue
        
        # Process data lines after header is found
        if header_found and line:
            # Split by comma
            parts = [part.strip() for part in line.split(',')]
            
            if len(parts) >= 13:  # Should have 13 columns based on your image
                try:
                    component_name = parts[0].strip()
                    
                    # Skip if component name is empty or contains only dashes
                    if not component_name or component_name.startswith("---"):
                        continue
                    
                    # Count valid components
                    component_count += 1
                    
                    # Extract residency values and add to totals
                    result["40us"] += float(parts[1]) if parts[1] else 0.0        # <=40us(%)
                    result["60us"] += float(parts[2]) if parts[2] else 0.0        # <=60us(%)
                    result["110us"] += float(parts[3]) if parts[3] else 0.0       # <=110us(%)
                    result["150us"] += float(parts[4]) if parts[4] else 0.0       # <=150us(%)
                    result["250us"] += float(parts[5]) if parts[5] else 0.0       # <=250us(%)
                    result["350us"] += float(parts[6]) if parts[6] else 0.0       # <=350us(%)
                    result["500us"] += float(parts[7]) if parts[7] else 0.0       # <=500us(%)
                    result["750us"] += float(parts[8]) if parts[8] else 0.0       # <=750us(%)
                    result["1ms"] += float(parts[9]) if parts[9] else 0.0         # <=1ms(%)
                    result["3ms"] += float(parts[10]) if parts[10] else 0.0       # <=3ms(%)
                    result["3ms_greater"] += float(parts[11]) if parts[11] else 0.0  # 3ms<(%)
                    result["no_req"] += float(parts[12]) if parts[12] else 0.0    # No Req(%)
                    
                except (ValueError, IndexError):
                    continue
    
    # Set the component count
    result["no_of_components"] = component_count
    
    # Round all values to 2 decimal places
    for key in result:
        if key != "no_of_components" and isinstance(result[key], float):
            result[key] = round(result[key], 2)
    
    return result

def parse_psr_residency_summary(file_path: str) -> Dict[str, Any]:
    """
    Parse Intel Socwatch CSV file and extract Panel Self-Refresh (PSR2) residency data.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing PSR residency data with state percentages
    """
    
    result = {
        "display_on_not_psr": 0.0,
        "psr_inactive": 0.0,
        "psr_capture_frame": 0.0,
        "psr_active": 0.0,
        "psr_deep_sleep": 0.0,
        "display_off_pwr_gated": 0.0
    }
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    # Find the PSR2 Summary section
    psr_section_start = None
    psr_section_end = None
    
    for i, line in enumerate(lines):
        if "Panel Self-Refresh (PSR2) Summary" in line and "Approximated Residency" in line:
            psr_section_start = i
            break
    
    if psr_section_start is None:
        return result
    
    # Find the end of PSR section
    for i in range(psr_section_start + 1, len(lines)):
        line = lines[i].strip()
        if line == "" or any(keyword in line for keyword in ["Summary", "---", "Report"]):
            # Check if this is actually the start of a new section
            if any(section in line for section in ["Summary", "Report"]) and "PSR" not in line:
                psr_section_end = i
                break
    
    if psr_section_end is None:
        psr_section_end = len(lines)
    
    # Parse PSR data
    header_found = False
    
    for i in range(psr_section_start, psr_section_end):
        line = lines[i].strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Skip separator lines with dashes
        if line.startswith("---") or line.replace("-", "").replace(" ", "") == "":
            continue
            
        # Look for header line
        if not header_found and "State" in line and "PSR (%)" in line:
            header_found = True
            continue
        
        # Process data lines after header is found
        if header_found and line:
            # Split by comma
            parts = [part.strip() for part in line.split(',')]
            
            if len(parts) >= 2:
                try:
                    state_name = parts[0].strip()
                    psr_percentage = float(parts[1]) if parts[1] and parts[1] != '-------' else 0.0
                    
                    # Map state names to result keys
                    if "Display On (Not PSR)" in state_name:
                        result["display_on_not_psr"] = psr_percentage
                    elif "PSR Inactive" in state_name:
                        result["psr_inactive"] = psr_percentage
                    elif "PSR Capture Frame" in state_name:
                        result["psr_capture_frame"] = psr_percentage
                    elif "PSR Active" in state_name:
                        result["psr_active"] = psr_percentage
                    elif "PSR Deep Sleep" in state_name:
                        result["psr_deep_sleep"] = psr_percentage
                    elif "Display Off (Pwr Gated)" in state_name:
                        result["display_off_pwr_gated"] = psr_percentage
                    
                except (ValueError, IndexError):
                    continue
    
    # Round all values to 2 decimal places
    for key in result:
        if isinstance(result[key], float):
            result[key] = round(result[key], 2)
    
    return result

def parse_display_refresh_rate_residency(file_path: str) -> Dict[str, Any]:
    """
    Parse Intel Socwatch CSV file and extract Display Refresh Rate residency data.
    Only returns non-zero residency values.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing only non-zero refresh rate residencies
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    # Find the Display Refresh Rate Histogram section
    refresh_section_start = None
    
    for i, line in enumerate(lines):
        if "Display Refresh Rate Summary - Sampled: Histogram" in line:
            refresh_section_start = i
            break
    
    if refresh_section_start is None:
        return result
    
    # Parse the histogram data
    header_found = False
    refresh_rate_columns = []
    
    for i in range(refresh_section_start, min(refresh_section_start + 10, len(lines))):
        line = lines[i].strip()
        
        # Skip empty lines and separator lines
        if not line or line.startswith("---") or line.replace("-", "").replace(",", "").replace(" ", "") == "":
            continue
        
        # Look for the header line with refresh rates
        if not header_found and "Hz(%)" in line:
            header_found = True
            parts = [part.strip() for part in line.split(',')]
            
            # Extract refresh rate column headers (skip first column which is "Display")
            for j, part in enumerate(parts[1:], 1):  # Start from index 1, skip "Display" column
                if "Hz(%)" in part:
                    # Extract just the frequency part (e.g., "60Hz" from "60Hz(%)")
                    refresh_rate = part.replace("(%)", "").strip()
                    refresh_rate_columns.append((j, refresh_rate))
            continue
        
        # Process data lines after header is found
        if header_found and line and "," in line:
            parts = [part.strip() for part in line.split(',')]
            
            if len(parts) > 1:
                display_name = parts[0].strip()
                
                # Skip if display name is empty or just dashes
                if not display_name or display_name.replace("-", "") == "":
                    continue
                
                # Check each refresh rate column for non-zero values
                for col_index, refresh_rate in refresh_rate_columns:
                    if col_index < len(parts):
                        try:
                            residency_value = float(parts[col_index]) if parts[col_index] and parts[col_index] != '-------' else 0.0
                            
                            # Only add non-zero residencies
                            if residency_value > 0:
                                result[refresh_rate] = round(residency_value, 2)
                        except (ValueError, IndexError):
                            continue
            
            # Stop processing after the first data row (since we only expect one display)
            break
    
    return result

def parse_ltr_snoop_summary(file_path: str) -> Dict[str, Any]:
    """
    Parse Intel Socwatch CSV file and extract PCD Platform LTR Snoop Summary data.
    Only reads from the exact table we identified in the debug.
    """
    
    result = {
        "ltr_components": {}
    }
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    # Find the exact LTR section header
    ltr_section_start = None
    
    for i, line in enumerate(lines):
        if "PCD Platform LTR Snoop Summary - Sampled: Ignore/Min/Max" in line:
            ltr_section_start = i
            break
    
    if ltr_section_start is None:
        return result
    
    # Find the header line with column names
    header_line_index = None
    min_ltr_column_index = None
    
    for i in range(ltr_section_start + 1, min(ltr_section_start + 5, len(lines))):
        line = lines[i].strip()
        if "Component Name" in line and "Min LTR(ns)" in line:
            header_line_index = i
            parts = [part.strip() for part in line.split(',')]
            for j, part in enumerate(parts):
                if "Min LTR(ns)" in part:
                    min_ltr_column_index = j
                    break
            break
    
    if header_line_index is None or min_ltr_column_index is None:
        return result
    
    # Process ONLY the lines immediately following the header until we hit the next section
    # Based on debug, we know the data starts right after the header
    data_start = header_line_index + 1
    
    # Look for the end of this specific table
    table_end = None
    for i in range(data_start, min(data_start + 100, len(lines))):
        line = lines[i].strip()
        
        # Stop when we hit the next section (look for specific patterns)
        if (line.startswith("PCD Platform LTR Snoop Summary - Sampled: Histogram") or
            line.startswith("PCD Platform LTR Snoop Summary - Sampled: Total Samples") or
            (line == "" and i > data_start + 20)):  # Empty line after reasonable data
            table_end = i
            break
    
    if table_end is None:
        table_end = min(data_start + 50, len(lines))  # Reasonable limit
    
    # Process ONLY the data lines within this specific table range
    for i in range(data_start, table_end):
        line = lines[i].strip()
        
        # Skip separator lines
        if (not line or 
            line.startswith("---") or 
            line.replace("-", "").replace(",", "").replace(" ", "") == ""):
            continue
        
        # Process CSV data lines
        if "," in line:
            parts = [part.strip() for part in line.split(',')]
            
            # Make sure we have enough columns
            if len(parts) > min_ltr_column_index:
                component_name = parts[0].strip()
                min_ltr_value = parts[min_ltr_column_index].strip()
                
                # Skip if component name is empty or invalid
                if not component_name or component_name.startswith("---"):
                    continue
                
                # Only capture components with numeric Min LTR values (not "No LTR")
                if (min_ltr_value and 
                    min_ltr_value != "No LTR" and 
                    min_ltr_value != "N/A" and 
                    min_ltr_value != ""):
                    
                    try:
                        min_ltr_numeric = float(min_ltr_value)
                        if min_ltr_numeric > 0:
                            result["ltr_components"][component_name] = min_ltr_numeric
                    except ValueError:
                        continue
    
    return result


    """
    Debug function to find all LTR-related content in the file.
    """
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    print("=== DEBUG: Searching for LTR content ===")
    
    ltr_lines = []
    for i, line in enumerate(lines):
        if "LTR" in line.upper():
            ltr_lines.append((i, line.strip()))
    
    print(f"Found {len(ltr_lines)} lines containing 'LTR':")
    for line_num, content in ltr_lines[:20]:  # Show first 20 matches
        print(f"Line {line_num}: {content}")
    
    if len(ltr_lines) > 20:
        print(f"... and {len(ltr_lines) - 20} more lines")
    
    # Also search for component names from your table
    component_names = [
        "AGGREGATE-SUBSYSTEM",
        "C2P2-CNVI-WIFI", 
        "CUR-PLT",
        "PCIE-CONTROLLER-B",
        "PMC"
    ]
    
    print("\n=== DEBUG: Searching for specific component names ===")
    for component in component_names:
        for i, line in enumerate(lines):
            if component in line:
                print(f"Line {i}: {line.strip()}")
                break
        else:
            print(f"Component '{component}' not found")

def parse_ltr_snoop_histogram(file_path: str) -> Dict[str, Any]:
    """
    Parse Intel Socwatch CSV file and extract PCD Platform LTR Snoop Histogram data.
    Bucketizes residency into <=150us, <=1ms, and >3ms for components with LTR values.
    
    Args:
        file_path (str): Path to the CSV file
        
    Returns:
        Dict containing bucketized LTR histogram data
    """
    
    result = {}
    
    lines = _smart_read_text(file_path).splitlines(True)
    
    # First, get the list of components that have LTR values
    ltr_components = parse_ltr_snoop_summary(file_path).get("ltr_components", {})
    
    if not ltr_components:
        return result
    
    # Find the LTR Histogram section
    histogram_section_start = None
    
    for i, line in enumerate(lines):
        if "PCD Platform LTR Snoop Summary - Sampled: Histogram" in line:
            histogram_section_start = i
            break
    
    if histogram_section_start is None:
        return result
    
    # Find the header line with column names
    header_line_index = None
    column_indices = {}
    
    for i in range(histogram_section_start + 1, min(histogram_section_start + 10, len(lines))):
        line = lines[i].strip()
        if "Component Name" in line and "us(%)" in line:
            header_line_index = i
            parts = [part.strip() for part in line.split(',')]
            
            # Map column headers to indices
            for j, part in enumerate(parts):
                if "<=40us(%)" in part:
                    column_indices["40us"] = j
                elif "<=60us(%)" in part:
                    column_indices["60us"] = j
                elif "<=110us(%)" in part:
                    column_indices["110us"] = j
                elif "<=150us(%)" in part:
                    column_indices["150us"] = j
                elif "<=250us(%)" in part:
                    column_indices["250us"] = j
                elif "<=350us(%)" in part:
                    column_indices["350us"] = j
                elif "<=500us(%)" in part:
                    column_indices["500us"] = j
                elif "<=750us(%)" in part:
                    column_indices["750us"] = j
                elif "<=1ms(%)" in part:
                    column_indices["1ms"] = j
                elif "<=3ms(%)" in part:
                    column_indices["3ms"] = j
                elif "3ms<(%)" in part:
                    column_indices["3ms_greater"] = j
            break
    
    if header_line_index is None:
        return result
    
    # Find the end of the histogram table
    table_end = None
    for i in range(header_line_index + 1, min(header_line_index + 100, len(lines))):
        line = lines[i].strip()
        if (line.startswith("PCD Platform LTR Snoop Summary - Sampled: Total Samples") or
            (line == "" and i > header_line_index + 30)):
            table_end = i
            break
    
    if table_end is None:
        table_end = min(header_line_index + 50, len(lines))
    
    # Process histogram data for components that have LTR values
    for i in range(header_line_index + 1, table_end):
        line = lines[i].strip()
        
        # Skip separator lines
        if (not line or 
            line.startswith("---") or 
            line.replace("-", "").replace(",", "").replace(" ", "") == ""):
            continue
        
        if "," in line:
            parts = [part.strip() for part in line.split(',')]
            
            if len(parts) > 1:
                component_name = parts[0].strip()
                
                # Only process components that have LTR values
                if component_name in ltr_components:
                    component_histogram = {
                        "ltr_value_ns": ltr_components[component_name],
                        "buckets": {
                            "<=150us": 0.0,
                            "<=1ms": 0.0,
                            ">3ms": 0.0
                        }
                    }
                    
                    try:
                        # Calculate <=150us bucket (sum of <=40us, <=60us, <=110us, <=150us)
                        bucket_150us = 0.0
                        for bucket in ["40us", "60us", "110us", "150us"]:
                            if bucket in column_indices and column_indices[bucket] < len(parts):
                                value = parts[column_indices[bucket]].strip()
                                if value and value != "0":
                                    bucket_150us += float(value)
                        
                        # Calculate <=1ms bucket (sum of <=250us, <=350us, <=500us, <=750us, <=1ms)
                        bucket_1ms = 0.0
                        for bucket in ["250us", "350us", "500us", "750us", "1ms"]:
                            if bucket in column_indices and column_indices[bucket] < len(parts):
                                value = parts[column_indices[bucket]].strip()
                                if value and value != "0":
                                    bucket_1ms += float(value)
                        
                        # Calculate >3ms bucket (3ms< column)
                        bucket_3ms_greater = 0.0
                        if "3ms_greater" in column_indices and column_indices["3ms_greater"] < len(parts):
                            value = parts[column_indices["3ms_greater"]].strip()
                            if value and value != "0":
                                bucket_3ms_greater = float(value)
                        
                        component_histogram["buckets"]["<=150us"] = round(bucket_150us, 2)
                        component_histogram["buckets"]["<=1ms"] = round(bucket_1ms, 2)
                        component_histogram["buckets"][">3ms"] = round(bucket_3ms_greater, 2)
                        
                        # Only add if at least one bucket has non-zero value
                        if any(v > 0 for v in component_histogram["buckets"].values()):
                            result[component_name] = component_histogram
                            
                    except (ValueError, IndexError):
                        continue
    
    return result
def parse_psr_link_state_residency(file_path: str) -> Dict[str, float]:
    """
    More flexible version that can handle slight variations in state names
    """
    result = {}
    lines = _smart_read_text(file_path).splitlines(True)
    
    section_found = False
    for i, line in enumerate(lines):
        if "Panel Self-Refresh (PSR2) (PSR Link States) Summary - Sampled: Approximated Residency (Percentage)" in line:
            section_found = True
            continue
            
        if section_found:
            if line.strip().startswith("-----") or line.strip().startswith("="):
                continue
                
            if line.strip() == "":
                break
                
            parts = [part.strip() for part in line.split(',')]
            
            # More flexible matching - check if the first part contains key state indicators
            if len(parts) >= 3:
                state_name = parts[0]
                if any(keyword in state_name for keyword in ["Full Off", "Full On", "Standby", "Display Off"]):
                    try:
                        result[state_name] = float(parts[2])
                        print(f"Found state: '{state_name}' = {parts[2]}")  # Debug output
                    except (ValueError, IndexError):
                        continue
                        
            # Check if we've hit the next section
            if "Panel Self-Refresh (PSR2) (PSR Link States) Summary - Sampled: Total Samples Received" in line:
                break
                
    return result

def parse_display_state_residency(file_path: str) -> Dict[str, int]:
    """
    Fixed version: Parse Intel Socwatch CSV file and extract Display State residency counts.
    """
    result = {}
    lines = _smart_read_text(file_path).splitlines(True)
    
    section_found = False
    for i, line in enumerate(lines):
        if "Display State Entry Summary: Total" in line:
            section_found = True
            continue
            
        if section_found:
            if "Display State" in line and "Total (Count)" in line:
                continue  # Skip header
                
            if line.strip().startswith("-----"):
                continue  # Skip separator
                
            if line.strip() == "" or line.strip().startswith("Total"):
                break  # End of section
                
            parts = [part.strip() for part in line.split(',')]
            
            # Based on your debug output: "Display, DC5, 0, ..."
            # So DC5 is in parts[1], and the count is in parts[2]
            if len(parts) >= 3 and parts[1] == "DC5":
                try:
                    result["DC5"] = int(parts[2])
                except (ValueError, IndexError):
                    continue
                    
            # You might also want to capture other display states
            elif len(parts) >= 3 and parts[1] in ["DC6", "DC0", "DC1", "DC2", "DC3", "DC4"]:
                try:
                    result[parts[1]] = int(parts[2])
                except (ValueError, IndexError):
                    continue
                    
    return result


def debug_file_structure(file_path: str) -> None:
    """
    Debug function to analyze the file structure around sections.
    """
    
    content = _smart_read_text(file_path)
    
    # Find all section headers
    sections = []
    lines = content.split('\n')
    
    for i, line in enumerate(lines):
        if "Summary:" in line:
            sections.append((i, line.strip()))
    
    #print("=== DEBUG: All sections found ===")
    #for line_num, section in sections:
        #print(f"Line {line_num}: {section}")


# === parse_socwatch_csv ===
def parse_socwatch_csv(file_path: str, json_output_dir: str = None) -> Dict[str, Any]:
    """
    Parse Intel Socwatch CSV file and extract all summaries including CPU P-State.
    Creates JSON file and returns the parsed data dictionary.

    Args:
        file_path (str): Path to the CSV file
        json_output_dir (str): Directory where the JSON output should be written.
                                Defaults to the same directory as the CSV file.

    Returns:
        Dict containing all summaries
    """
    
    # Parse each section independently — a missing/unsupported section returns {} instead of
    # aborting the whole file.
    _section_parsers = [
        ("Package C-State Summary",          parse_package_c_state_summary),
        ("Package C-State (OS) Summary",      parse_package_c_state_os_summary),
        ("Core C-State Summary",              parse_core_c_state_summary),
        ("CPU P-State Summary",               parse_cpu_pstate_summary),
        ("Thread Wakeups (OS) Summary",       parse_thread_wakeups_os_summary),
        ("Media P-State Summary",             parse_media_p_state_summary),
        ("Media C-State Summary",             parse_media_c_state_summary),
        ("GFX P-State Summary",               parse_gfx_p_state_summary),
        ("GFX C-State Summary",               parse_gfx_c_state_summary),
        ("MEMSS P-State Summary",             parse_memss_p_state_summary),
        ("IPU C-State Summary",               parse_ipu_c_state_summary),
        ("IPU P-State Summary",               parse_ipu_p_state_summary),
        ("NPU P-State Summary",               parse_npu_p_state_summary),
        ("NPU D-State Summary",               parse_npu_d_state_summary),
        ("DDR Bandwidth Summary",             parse_ddr_bandwidth_summary),
        ("LTR Snoop Summary",                 parse_ltr_snoop_summary),
        ("LTR Snoop Histogram",               parse_ltr_snoop_histogram),
        ("PSR Residency Summary",             parse_psr_residency_summary),
        ("Display Refresh Rate Residency",    parse_display_refresh_rate_residency),
        ("PSR Link state Summary",            parse_psr_link_state_residency),
        ("Display Cstate Residency",          parse_display_state_residency),
    ]

    result = {}
    section_errors = []
    for section_name, parser_fn in _section_parsers:
        try:
            result[section_name] = parser_fn(file_path)
        except Exception as sec_err:
            result[section_name] = {}  # empty — section not present in this file
            section_errors.append(f"  [{section_name}]: {sec_err}")

    if section_errors:
        safe_print(f"  Note: {len(section_errors)} section(s) not found / skipped:")
        for msg in section_errors:
            safe_print(msg)
    safe_print(f"\nParsing Socwatch file DONE ({len(result) - len(section_errors)}/{len(_section_parsers)} sections parsed)")
    # Create JSON file before returning the result
    file_key = os.path.splitext(os.path.basename(file_path))[0]
    _json_dir = json_output_dir if json_output_dir else os.path.dirname(file_path)
    os.makedirs(_json_dir, exist_ok=True)
    json_output_path = os.path.join(_json_dir, f"{file_key}.json")
    save_to_json(result, json_output_path)
    safe_print(f"\nCreating Json here: {json_output_path} ")
    
    #Hint: if excel summary is asked, send folder_path as argument to create_excel_comparison().
    #Hint: Please pass the data "result" as argument "all_data" to create_excel_comparison()
    return result

# Global variable to store the log file path for the session
_log_file_path = None


# === create_excel_comparison ===
def create_excel_comparison(all_data: Dict[str, Dict], output_path: str):
    """
    Create an Excel file with a single sheet in pivot format with section headings.
    Generic version that works with any data structure without hardcoding.
    Maintains original format: Metric column + one column per file.
    
    Args:
        all_data: Dictionary with file names as keys and parsed data as values
                 OR single file data (will be auto-wrapped)
        output_path: Path for output Excel file
    """
    safe_print(f"\nIn create_excel_comparison, processing {len(all_data)} data sources")
    safe_print(f"Output path: {output_path}")
    
    # **Handle both formats - single file data or multi-file data**
    if all_data and not any(isinstance(v, dict) for v in all_data.values()):
        # This looks like single file data (sections at top level)
        filename = os.path.splitext(os.path.basename(output_path))[0]
        if filename == "summary":
            filename = "socwatch_data"
        all_data = {filename: all_data}
        safe_print(f"Auto-wrapped single file data with key: {filename}")
    
    # Get sorted file names for consistent column order
    file_names = sorted(all_data.keys())
    safe_print(f"Processing files: {file_names}")

    # Prepare data for DataFrame
    df_data = []
    
    # **GENERIC APPROACH: Discover all sections dynamically**
    all_sections = set()
    for file_name, data in all_data.items():
        if isinstance(data, dict):
            all_sections.update(data.keys())
    
    # Sort sections for consistent ordering
    sorted_sections = sorted(all_sections)
    safe_print(f"Found {len(sorted_sections)} sections: {sorted_sections}")
    
    sections_processed = 0
    
    for section_key in sorted_sections:
        # Check if any file has data for this section
        section_has_data = False
        section_metrics = {}
        
        for file_name, data in all_data.items():
            if not isinstance(data, dict):
                continue
                
            section_data = data.get(section_key, {})
            if section_data:
                section_has_data = True
                safe_print(f"Processing section '{section_key}' for file '{file_name}'")
                
                # **GENERIC PROCESSING: Handle different data structures dynamically**
                processed_metrics = process_section_generically(section_data, section_key)
                
                # Merge metrics into section_metrics
                for metric_name, value in processed_metrics.items():
                    if metric_name not in section_metrics:
                        section_metrics[metric_name] = {}
                    section_metrics[metric_name][file_name] = value
        
        # Add section to DataFrame if it has data
        if section_has_data and section_metrics:
            sections_processed += 1
            safe_print(f"Adding section '{section_key}' with {len(section_metrics)} metrics")
            
            # Add section heading row (SAME FORMAT AS ORIGINAL)
            section_heading = format_section_heading(section_key)
            heading_row = {"Metric": section_heading}
            for file_name in file_names:
                heading_row[file_name] = ""
            df_data.append(heading_row)
            
            # **MAINTAIN ORIGINAL GROUPING LOGIC**
            grouped_metrics = group_metrics_intelligently(section_metrics, section_key)
            
            # Add metrics to DataFrame (SAME FORMAT AS ORIGINAL)
            for group_info in grouped_metrics:
                if group_info["type"] == "subheading":
                    # Add sub-heading
                    subheading_row = {"Metric": f"  {group_info['name']}"}
                    for file_name in file_names:
                        subheading_row[file_name] = ""
                    df_data.append(subheading_row)
                    
                    # Add metrics under this subheading
                    for metric_name in group_info["metrics"]:
                        row = {"Metric": f"    {metric_name}"}
                        for file_name in file_names:
                            value = section_metrics[metric_name].get(file_name, 0)
                            row[file_name] = value
                        df_data.append(row)
                
                elif group_info["type"] == "direct":
                    # Add metrics directly under section
                    for metric_name in group_info["metrics"]:
                        row = {"Metric": f"  {metric_name}"}
                        for file_name in file_names:
                            value = section_metrics[metric_name].get(file_name, 0)
                            row[file_name] = value
                        df_data.append(row)
            
            # Add empty row after each section for spacing (SAME AS ORIGINAL)
            empty_row = {"Metric": ""}
            for file_name in file_names:
                empty_row[file_name] = ""
            df_data.append(empty_row)
    
    safe_print(f"Total sections processed: {sections_processed}")
    safe_print(f"Total DataFrame rows: {len(df_data)}")
    
    if not df_data:
        safe_print("ERROR: No data to write to Excel!")
        return
    
    # Create DataFrame (SAME AS ORIGINAL)
    df_pivot = pd.DataFrame(df_data)
    safe_print(f"DataFrame created with shape: {df_pivot.shape}")
    
    # Create Excel file with single sheet (SAME AS ORIGINAL)
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_pivot.to_excel(writer, sheet_name="Socwatch Comparison", index=False)
            
            # Format the worksheet (SAME AS ORIGINAL)
            worksheet = writer.sheets["Socwatch Comparison"]
            
            # Auto-adjust column widths (SAME AS ORIGINAL)
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format cells (SAME AS ORIGINAL)
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Style section headings (SAME AS ORIGINAL)
            section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            subsection_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            bold_font = Font(bold=True)
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                metric_cell = row[0]
                metric_value = str(metric_cell.value) if metric_cell.value else ""
                
                # Main section heading (no indentation) - SAME AS ORIGINAL
                if metric_value and not metric_value.startswith("  ") and metric_value != "":
                    for cell in row:
                        cell.font = bold_font
                        cell.fill = section_fill
                        if cell.column == 1:
                            cell.alignment = Alignment(horizontal='left')
                
                # Sub-section heading (2 spaces) - SAME AS ORIGINAL
                elif metric_value.startswith("  ") and not metric_value.startswith("    "):
                    # Check if it's a sub-heading
                    if is_subheading(metric_value):
                        for cell in row:
                            cell.font = Font(bold=True, italic=True)
                            cell.fill = subsection_fill
                            if cell.column == 1:
                                cell.alignment = Alignment(horizontal='left')
                    else:
                        # Regular metric - SAME AS ORIGINAL
                        for col_num, cell in enumerate(row):
                            if col_num == 0:
                                cell.alignment = Alignment(horizontal='left')
                            else:
                                if isinstance(cell.value, (int, float)) and cell.value != 0:
                                    cell.number_format = '0.00'
                                    cell.alignment = Alignment(horizontal='right')
                                elif cell.value == "No LTR":
                                    cell.alignment = Alignment(horizontal='center')
                
                # Sub-metric (4 spaces indentation) - SAME AS ORIGINAL
                elif metric_value.startswith("    "):
                    for col_num, cell in enumerate(row):
                        if col_num == 0:
                            cell.alignment = Alignment(horizontal='left')
                        else:
                            if isinstance(cell.value, (int, float)) and cell.value != 0:
                                # Don't format count metrics as decimals
                                if any(word in metric_value for word in ["Total", "Count"]) and not "%" in metric_value:
                                    cell.number_format = '0'
                                else:
                                    cell.number_format = '0.00'
                                cell.alignment = Alignment(horizontal='right')
                
                # Empty rows or other cases - SAME AS ORIGINAL
                else:
                    for col_num, cell in enumerate(row):
                        if col_num == 0:
                            cell.alignment = Alignment(horizontal='left')
                        else:
                            if isinstance(cell.value, (int, float)) and cell.value != 0:
                                cell.number_format = '0.00'
                                cell.alignment = Alignment(horizontal='right')
        
        safe_print(f"Excel file created successfully: {output_path}")
        
    except Exception as e:
        safe_print(f"ERROR creating Excel file: {str(e)}")
        return
    
    # Create markdown file automatically after Excel creation (SAME AS ORIGINAL)
    try:
        base_path = os.path.splitext(output_path)[0]
        markdown_path = f"{base_path}.md"
        
        excel_to_markdown_content = excel_to_markdown(output_path)
        
        if excel_to_markdown_content:
            with open(markdown_path, 'w', encoding='utf-8') as f:
                f.write(excel_to_markdown_content)
            print(f"  Markdown file created: {os.path.basename(markdown_path)}")
        
    except Exception as e:
        print(f"  Warning: Could not create markdown file: {str(e)}")


def process_section_generically(section_data: Dict, section_key: str) -> Dict[str, any]:
    """
    Generic function to process any section data structure.
    Returns a flat dictionary of metric_name -> value.
    """
    metrics = {}
    
    def flatten_data(data, prefix="", max_depth=4, current_depth=0):
        """Recursively flatten nested dictionaries"""
        if current_depth >= max_depth:
            return {prefix.rstrip("_"): str(data)}
        
        if isinstance(data, dict):
            result = {}
            for key, value in data.items():
                new_prefix = f"{prefix}{key}_" if prefix else f"{key}_"
                
                if isinstance(value, dict):
                    # Recursively process nested dictionaries
                    nested_result = flatten_data(value, new_prefix, max_depth, current_depth + 1)
                    result.update(nested_result)
                elif isinstance(value, (list, tuple)):
                    # Handle lists/tuples
                    for i, item in enumerate(value):
                        item_prefix = f"{new_prefix}item_{i}_"
                        if isinstance(item, dict):
                            nested_result = flatten_data(item, item_prefix, max_depth, current_depth + 1)
                            result.update(nested_result)
                        else:
                            result[f"{item_prefix.rstrip('_')}"] = item
                else:
                    # Direct value
                    result[new_prefix.rstrip("_")] = value
            return result
        else:
            return {prefix.rstrip("_"): data}
    
    # Flatten the section data
    flattened = flatten_data(section_data)
    
    # Clean up metric names
    for key, value in flattened.items():
        clean_key = clean_metric_name(key, section_key)
        metrics[clean_key] = value
    
    return metrics


def clean_metric_name(metric_name: str, section_key: str) -> str:
    """Clean up metric names for better readability"""
    # Remove redundant section prefixes
    clean_name = metric_name
    
    # Replace underscores with spaces
    clean_name = clean_name.replace("_", " ")
    
    # Handle special abbreviations
    replacements = {
        " cpu ": " CPU ",
        " ddr ": " DDR ",
        " ltr ": " LTR ",
        " psr ": " PSR ",
        " gfx ": " GFX ",
        " npu ": " NPU ",
        " ipu ": " IPU ",
        " memss ": " MEMSS ",
        " mhz": " MHz",
        " mbps": " MB/s",
        "pcore": "P-Core",
        "ecore": "E-Core"
    }
    
    clean_name_lower = clean_name.lower()
    for old, new in replacements.items():
        clean_name_lower = clean_name_lower.replace(old, new)
    
    # Capitalize first letter of each word
    clean_name = " ".join(word.capitalize() if word.lower() not in ["mhz", "mb/s", "cpu", "ddr", "ltr", "psr", "gfx", "npu", "ipu", "memss"] 
                         else word.upper() if word.lower() in ["cpu", "ddr", "ltr", "psr", "gfx", "npu", "ipu", "memss"]
                         else word for word in clean_name_lower.split())
    
    return clean_name


def format_section_heading(section_key: str) -> str:
    """Format section key into a readable heading - maintains original format"""
    # Remove "Summary" suffix and convert to uppercase
    heading = section_key.replace(" Summary", "").replace("Summary", "")
    return heading.upper()


def group_metrics_intelligently(section_metrics: Dict[str, Dict], section_key: str) -> List[Dict]:
    """
    Intelligently group metrics based on patterns, maintaining original grouping logic.
    Returns list of group info dictionaries.
    """
    groups = []
    
    # Define intelligent grouping patterns based on common metric types
    grouping_rules = {
        # Frequency-related groupings
        "Frequency Statistics": {
            "keywords": ["frequency", "mhz", "average", "max", "min"],
            "priority": 1
        },
        "Core/Thread Information": {
            "keywords": ["total", "cores", "threads", "processors", "logical"],
            "priority": 2
        },
        "Frequency Residency": {
            "keywords": ["residency", "frequency"],
            "priority": 3
        },
        # Activity-related groupings
        "Platform Activity": {
            "keywords": ["platform", "activity", "overall"],
            "priority": 1
        },
        "Process-wise Activity": {
            "keywords": ["process"],
            "priority": 2
        },
        # Bandwidth-related groupings
        "Individual DDR Channels": {
            "keywords": ["channel", "subch", "mc0", "mc1", "reads", "writes"],
            "priority": 2
        },
        "Individual IP Channels": {
            "keywords": ["display", "npu", "noc", "idi", "mufasa", "hbo", "santa"],
            "priority": 3
        }
    }
    
    # Categorize metrics
    categorized_metrics = {}
    uncategorized_metrics = []
    
    for metric_name in section_metrics.keys():
        metric_lower = metric_name.lower()
        assigned = False
        
        for group_name, rule in grouping_rules.items():
            if any(keyword in metric_lower for keyword in rule["keywords"]):
                if group_name not in categorized_metrics:
                    categorized_metrics[group_name] = {
                        "metrics": [],
                        "priority": rule["priority"]
                    }
                categorized_metrics[group_name]["metrics"].append(metric_name)
                assigned = True
                break
        
        if not assigned:
            uncategorized_metrics.append(metric_name)
    
    # Sort groups by priority
    sorted_groups = sorted(categorized_metrics.items(), key=lambda x: x[1]["priority"])
    
    # Add uncategorized metrics first (direct under section)
    if uncategorized_metrics:
        groups.append({
            "type": "direct",
            "metrics": sorted(uncategorized_metrics)
        })
    
    # Add categorized groups as subheadings
    for group_name, group_info in sorted_groups:
        groups.append({
            "type": "subheading",
            "name": group_name,
            "metrics": sorted(group_info["metrics"])
        })
    
    # If no intelligent grouping found, return all as direct
    if not groups:
        groups.append({
            "type": "direct",
            "metrics": sorted(section_metrics.keys())
        })
    
    return groups


def is_subheading(metric_value: str) -> bool:
    """Check if a metric value is a subheading"""
    subheading_keywords = [
        "Individual", "Frequency Statistics", "Core/Thread Information", 
        "Frequency Residency", "Platform Activity", "Process-wise Activity",
        "DDR Channels", "IP Channels", "Statistics", "Information", "Activity"
    ]
    
    clean_value = metric_value.strip()
    return any(keyword in clean_value for keyword in subheading_keywords)

      
def excel_to_markdown(excel_file) -> str:
    """
    Convert Excel file to markdown table format and save as .txt file.    
    Args:
        excel_file (str): Path to Excel file (.xlsx or .xls)    
    Returns:
        str: Path to the created .txt file
    """
    
    try:
        # Check if file exists
        if not os.path.exists(excel_file):
            #print(f"Error: File '{excel_file}' not found")
            return None
        
        # Read Excel file (first sheet)
        df = pd.read_excel(excel_file)
        #print(f"  Read Excel file: {os.path.basename(excel_file)}")
        #print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")
        
        # Handle empty file
        if df.empty:
            #print("Warning: Excel file is empty")
            return None
        
        # Fill NaN values with empty strings
        df = df.fillna('')
        
        # Convert to markdown table
        markdown_table = create_markdown_table(df)
        
        # Create output filename (same directory, same name with .txt extension)
        base_name = os.path.splitext(excel_file)[0]
        output_file = f"{base_name}_markdown.txt"
        
        # Write to .txt file
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(markdown_table)
        
        #print(f" Markdown table saved to: {os.path.basename(output_file)}")
        return markdown_table  # Return content, not file path
        
    except Exception as e:
        #print(f"Error: {str(e)}")
        return None

def create_markdown_table(df):
    """
    Create markdown table from DataFrame.    
    Args:
        df (pd.DataFrame): DataFrame to convert    
    Returns:
        str: Markdown table string
    """
    
    # Get column names
    columns = [str(col) for col in df.columns]
    
    # Clean column names (escape pipes)
    clean_columns = [col.replace('|', '\\|') for col in columns]
    
    # Start building markdown
    lines = []
    
    # Header row
    header = '| ' + ' | '.join(clean_columns) + ' |'
    lines.append(header)
    
    # Separator row (center-aligned)
    separator = '| ' + ' | '.join(['---'] * len(clean_columns)) + ' |'
    lines.append(separator)
    
    # Data rows
    for _, row in df.iterrows():
        # Convert all values to strings and clean them
        row_values = []
        for val in row.values:
            clean_val = str(val).replace('|', '\\|').replace('\n', ' ')
            row_values.append(clean_val)
        
        # Create row
        data_row = '| ' + ' | '.join(row_values) + ' |'
        lines.append(data_row)
    
    return '\n'.join(lines)

# Keep the compatibility functions

# ---------------------------------------------------------------------------
# Stubs – parse_socwatch_data only processes socwatch CSVs (find_socwatch_files
# guarantees that), so these branches are never reached in normal usage.
# ---------------------------------------------------------------------------
def parse_power_summary_csv(file_path: str) -> Dict[str, Any]:
    """Stub – not used when called via find_socwatch_files + parse_socwatch_data."""
    return parse_socwatch_csv(file_path)

def parse_generic_csv(file_path: str) -> Dict[str, Any]:
    """Stub – not used when called via find_socwatch_files + parse_socwatch_data."""
    return parse_socwatch_csv(file_path)


# ---------------------------------------------------------------------------
# Module-level helper — shared by parse_socwatch_data and query_socwatch_data
# ---------------------------------------------------------------------------
def _extract_sections_from_md(md_text: str) -> Dict[str, str]:
    """Split markdown into {section_name: content}.

    Handles two formats:
    1. Headed format  — lines starting with '## ' or '# ' become section keys.
    2. Flat-table format — the output of excel_to_markdown(), where section headings
       appear as pipe-table rows whose first column is non-empty with no leading spaces
       and all other columns are empty (e.g. ``| PACKAGE C-STATE |  |  |``).
    """
    # --- Format 1: headed markdown -------------------------------------------
    sections_map: Dict[str, str] = {}
    current_name: Optional[str] = None
    current_lines: List[str] = []
    for line in md_text.splitlines(keepends=True):
        if line.startswith("## "):
            if current_name is not None:
                sections_map[current_name] = "".join(current_lines)
            current_name = line[3:].strip()
            current_lines = [line]
        elif line.startswith("# "):
            if current_name is not None:
                sections_map[current_name] = "".join(current_lines)
            current_name = line[2:].strip()
            current_lines = [line]
        else:
            current_lines.append(line)
    if current_name is not None:
        sections_map[current_name] = "".join(current_lines)

    if sections_map:
        return sections_map

    # --- Format 2: flat pipe-table (excel_to_markdown output) ----------------
    # Rows look like:  | PACKAGE C-STATE |  |  |  |
    # Section detector: first cell non-empty, no leading spaces, all other cells empty.
    all_rows = md_text.splitlines()
    header_row: Optional[str] = None
    sep_row: Optional[str] = None
    data_rows: List[str] = []

    for row in all_rows:
        stripped = row.strip()
        if not stripped.startswith("|"):
            continue
        if header_row is None:
            header_row = row
            continue
        if sep_row is None and set(stripped.replace("|", "").replace("-", "").replace(" ", "")) == set():
            sep_row = row
            continue
        data_rows.append(row)

    if header_row is None or not data_rows:
        return {}

    def _parse_pipe_row_raw(r: str) -> List[str]:
        """Return raw (unstripped) cell values from a pipe-delimited row."""
        cells = r.split("|")
        if cells and cells[0].strip() == "":
            cells = cells[1:]
        if cells and cells[-1].strip() == "":
            cells = cells[:-1]
        return cells  # raw, not stripped

    def _parse_pipe_row(r: str) -> List[str]:
        return [c.strip() for c in _parse_pipe_row_raw(r)]

    # Parse header to know column count
    header_cells = _parse_pipe_row(header_row)
    n_cols = len(header_cells)

    section_rows: Dict[str, List[str]] = {}
    section_order: List[str] = []
    current_section: Optional[str] = None
    table_header = header_row + "\n" + (sep_row or "") + "\n"

    for row in data_rows:
        raw_cells = _parse_pipe_row_raw(row)
        cells = [c.strip() for c in raw_cells]
        if not cells:
            continue
        first_raw = raw_cells[0] if raw_cells else ""
        first = cells[0]
        rest = cells[1:n_cols] if len(cells) >= n_cols else cells[1:]
        all_rest_empty = all(c == "" for c in rest)

        # Section heading vs sub-heading detection:
        # In markdown pipe tables, every cell has 1 space of padding after |.
        # Excel section headings have NO extra indent → raw first cell = " VALUE "  (1 leading space)
        # Excel sub-headings have 2-space indent     → raw first cell = "   VALUE " (3 leading spaces)
        # Excel metrics have 2-4 space indent        → raw first cell = "   VALUE " (3+ leading spaces)
        # Rule: leading spaces ≤ 1 (just pipe padding) = section heading
        leading_sp = len(first_raw) - len(first_raw.lstrip())
        is_section_heading = (
            first               # non-empty after strip
            and leading_sp <= 1  # only standard pipe-table padding, no extra indent
            and all_rest_empty
        )

        if is_section_heading:
            current_section = first
            if current_section not in section_rows:
                section_rows[current_section] = []
                section_order.append(current_section)
        elif current_section is not None:
            section_rows[current_section].append(row)

    for sec in section_order:
        rows = [r for r in section_rows[sec] if _parse_pipe_row(r) and any(_parse_pipe_row(r))]
        if rows:
            sections_map[sec] = table_header + "\n".join(rows)

    return sections_map


# === MCP tool: find_socwatch_files ===
@mcp.tool()
@async_tool
@embed_if_large(threshold=3000)
def find_socwatch_files(parent_folder, force_reparse: bool = False, debug: bool = False) -> Dict[str, Any]:
    """
    Find all SocWatch CSV files in subdirectories of parent_folder using hybrid detection
    (filename regex + content marker scan). Files are NOT copied — their original paths
    are returned directly. The Analysis/socwatch_output/ folder is only used for
    output artifacts (JSON, Excel, Markdown) when parse_socwatch_data is called.

    Detection is two-stage:
      1. Filename regex  soc.*watch  (fast path)
      2. Content markers scan — first 50 lines, requires ≥4 of 5 SocWatch markers
         AND 'Package C-State Summary' must be present (catches renamed files)

    Files already inside any 'socwatch_output*' folder are excluded from discovery
    to avoid re-processing previously written artifacts.

    Args:
        parent_folder: Path to the parent directory to scan for SocWatch CSVs
        force_reparse: If True, ignore cached summary and re-discover files.
        debug: If True, include per-file detection decisions in the return value
               (use this when files aren't being found to diagnose why).

    Returns:
        Dict with:
          found        (bool)  – True if any SocWatch files were detected
          file_paths   (list)  – Full paths to discovered SocWatch CSV files
          output_folder (str)  – Path where artifacts will be written (Analysis/socwatch_output/)
          file_count   (int)  – Number of SocWatch files found
          file_names   (list) – Basenames of found files
          content_detected (list) – Files found via content scan (not filename match)
          already_parsed (bool) – True if cached summary already exists
          message      (str)  – Human-readable discovery summary
          debug_log    (list) – Per-file detection decisions (only when debug=True)
    """
    
    parent_path = Path(parent_folder)
    
    # Check if parent folder exists
    if not parent_path.exists():
        return {
            "found": False, "output_folder": "", "file_count": 0,
            "file_names": [], "content_detected": [],
            "message": f"Folder not found: {parent_folder}"
        }
    
    # Output artifacts folder (JSON, Excel, Markdown) — created only when parse_socwatch_data runs
    analysis_folder = parent_path / "Analysis"
    output_folder = analysis_folder / "socwatch_output"

    # Shared exclusion pattern — derived/analysis files that share content markers
    # but are NOT raw SocWatch output. Applied in ALL code paths (live scan + cache).
    _EXCL = re.compile(
        r'(WakeupAnalysis|wakeup_analysis|_analysis|_summary_report|_processed|_output)',
        re.IGNORECASE
    )
    _SW_PAT = re.compile(r'soc.*watch', re.IGNORECASE)

    def _is_kpi_csv(csv_path: Path) -> bool:
        """Return True if this looks like a KPI/workload companion CSV (not socwatch, not output)."""
        name = csv_path.name
        if _SW_PAT.search(name):
            return False  # it IS a socwatch file
        if _EXCL.search(name):
            return False  # derived file
        # Skip if inside any artifact output folder
        if any(p.name.lower().startswith(("socwatch_output", "power_output", "analysis"))
               for p in csv_path.parents):
            return False
        # Accept files that look like workload result files (non-timestamped, or contain a
        # meaningful label like a KPI name rather than a bare timestamp)
        stem = csv_path.stem
        has_timestamp = bool(re.match(r'\d{8}T\d{6}', stem))
        return not has_timestamp  # timestamp-named files are raw measurements, not KPI labels
    # ── Cache check ─────────────────────────────────────────────────────────────
    # Fast path: the output markdown existing with content is sufficient proof that
    # parsing already completed — no manifest required.  The expensive glob + content
    # scan only runs on first call or when force_reparse=True.
    if not force_reparse:
        cached_md = output_folder / "socwatch_output_summary.md"
        manifest_file = output_folder / "socwatch_parse_manifest.json"
        if cached_md.exists() and cached_md.stat().st_size > 0:
            # Try to enrich the response with the file list from manifest, but
            # never fall through to a live scan just because the manifest is absent.
            cached_paths = []
            try:
                if manifest_file.exists():
                    import json as _json_sw
                    with open(manifest_file, "r", encoding="utf-8") as _mf:
                        _manifest = _json_sw.load(_mf)
                    cached_paths = _manifest.get("parsed_files", [])
            except Exception:
                pass
            _cn = [Path(p).name for p in cached_paths]
            _sample = _cn[:5]
            _extra = len(_cn) - 5
            _names_str = ", ".join(_sample) + (f" ... and {_extra} more" if _extra > 0 else "")
            return {
                "found": True,
                "output_folder": str(output_folder),
                "file_count": len(cached_paths),
                "file_paths": cached_paths,
                "file_names": _sample + ([f"... and {_extra} more"] if _extra > 0 else []),
                "content_detected": [],
                "already_parsed": True,
                "kpi_context_files": [],
                "message": (
                    f"Cached results found"
                    + (f" ({len(cached_paths)} SocWatch file(s): {_names_str})" if _names_str else "")
                    + f". socwatch_output_summary.md already exists. "
                    f"Pass force_reparse=true to re-parse."
                ),
            }
    # ── End cache check ─────────────────────────────────────────────────────────
    
    # Find all CSV files recursively, but exclude files from folders starting with 'socwatch_output'
    all_csv_files = []
    kpi_context_files: List[str] = []
    for csv_file in parent_path.glob("**/*.csv"):
        if any(p.name.lower().startswith('socwatch_output') for p in csv_file.parents):
            continue
        if _EXCL.search(csv_file.name):
            continue
        if _SW_PAT.search(csv_file.name):
            all_csv_files.append(csv_file)
        elif _is_kpi_csv(csv_file):
            kpi_context_files.append(str(csv_file))

    _SOCWATCH_CONTENT_MARKERS = [
        "Intel(R) SoC Watch",
        "SoC Watch for Windows",
        "Package C-State Summary",
        "Core C-State Summary",
        "CPU P-State/Frequency Summary",
        "C-State Residency",
        "SoCWatch",
    ]

    debug_log = []  # collects per-file decisions when debug=True

    def _is_socwatch_by_content(csv_path) -> bool:
        """Return True if the CSV file contains ≥3 SocWatch content markers in its first 300 lines."""
        try:
            with open(csv_path, 'r', encoding='utf-8', errors='ignore') as fh:
                head = ''.join(fh.readline() for _ in range(300))
            hit_count = sum(1 for m in _SOCWATCH_CONTENT_MARKERS if m in head)
            if debug:
                hits = [m for m in _SOCWATCH_CONTENT_MARKERS if m in head]
                debug_log.append({
                    "file": str(csv_path),
                    "markers_found": hits,
                    "marker_count": hit_count,
                    "detected": hit_count >= 3,
                })
            return hit_count >= 3
        except Exception as exc:
            if debug:
                debug_log.append({"file": str(csv_path), "error": str(exc), "detected": False})
            return False

    socwatch_files = list(all_csv_files)  # filename-matched, exclusions already applied above
    content_detected_files: List[str] = []

    # Content-scan pass: non-socwatch-named CSVs that might be renamed SocWatch files
    for csv_file in parent_path.glob("**/*.csv"):
        if any(p.name.lower().startswith('socwatch_output') for p in csv_file.parents):
            continue
        if _EXCL.search(csv_file.name):
            continue
        if _SW_PAT.search(csv_file.name):
            continue  # already in socwatch_files
        if csv_file in all_csv_files:
            continue
        if _is_socwatch_by_content(csv_file):
            socwatch_files.append(csv_file)
            content_detected_files.append(csv_file.name)
        elif debug and not any(d.get("file") == str(csv_file) for d in debug_log):
            debug_log.append({"file": str(csv_file), "detected": False, "method": "content_scan", "note": "<3 markers"})

    _kpi_ctx_note = (
        f" KPI companion files in same folder: {kpi_context_files[:8]}."
        if kpi_context_files else ""
    )

    if socwatch_files:
        _names_all = [f.name for f in socwatch_files]
        _sample_names = _names_all[:5]
        _extra_n = len(_names_all) - 5
        _names_display = ", ".join(_sample_names) + (f" ... and {_extra_n} more" if _extra_n > 0 else "")
        _content_note = f" ({len(content_detected_files)} via content scan)" if content_detected_files else ""
        return {
            "found": True,
            "output_folder": str(output_folder),
            "file_count": len(socwatch_files),
            "file_paths": [str(f) for f in socwatch_files],
            "file_names": _sample_names + ([f"... and {_extra_n} more"] if _extra_n > 0 else []),
            "content_detected": content_detected_files[:5],
            "kpi_context_files": kpi_context_files[:8],
            "already_parsed": False,
            "message": f"Found {len(socwatch_files)} SocWatch file(s){_content_note}: {_names_display}.",
            **({
                "debug_log": [e for e in debug_log if e.get("detected") or e.get("error") or e.get("marker_count", 0) > 0]
            } if debug else {}),
        }
    else:
        return {
            "found": False, "output_folder": "", "file_count": 0,
            "file_names": [], "content_detected": [],
            "kpi_context_files": kpi_context_files[:8],
            "already_parsed": False,
            "message": "No SocWatch CSV files found in the provided folder or its subdirectories.",
            **({
                "debug_log": [e for e in debug_log if e.get("detected") or e.get("error") or e.get("marker_count", 0) > 0]
            } if debug else {}),
        }


# === MCP tool: parse_socwatch_data ===

@mcp.tool()
@async_tool
@embed_if_large(threshold=3000)
def parse_socwatch_data(
    parent_folder: str,
    force_reparse: bool = False,
) -> Dict[str, Any]:
    """
    Parse SocWatch CSV files and write Excel + Markdown summary to disk.

    Call find_socwatch_files(parent_folder) first to confirm files exist,
    then call this. Internally re-discovers SocWatch CSVs from parent_folder.
    Returns compact metadata only (section names, counts, file paths).
    Use query_socwatch_data(parent_folder, sections=[...]) to view section content.

    Args:
        parent_folder (str): Root folder — same as used for find_socwatch_files.
        force_reparse (bool): Re-parse even if socwatch_output_summary.md already exists.

    Returns:
        dict with:
            success, cached, section_names, section_count, files_parsed,
            full_table_path, excel_path, message.
    """
    parent_path = Path(parent_folder)
    output_dir = parent_path / "Analysis" / "socwatch_output"
    output_folder = str(output_dir)
    cached_md = str(output_dir / "socwatch_output_summary.md")

    def _build_compact_metadata(files_parsed: int, cached: bool) -> Dict[str, Any]:
        """Return compact metadata — reads section names from disk, no content in response."""
        section_names: List[str] = []
        if os.path.exists(cached_md):
            try:
                with open(cached_md, "r", encoding="utf-8") as _f:
                    _md = _f.read()
                section_names = list(_extract_sections_from_md(_md).keys())
            except Exception:
                pass
        excel_path = str(output_dir / "socwatch_output_summary.xlsx")
        return {
            "success": True,
            "cached": cached,
            "section_names": section_names,
            "section_count": len(section_names),
            "files_parsed": files_parsed,
            "full_table_path": cached_md,
            "excel_path": excel_path,
            "message": (
                f"Pipeline {'loaded from cache' if cached else 'compiled'}. "
                f"{len(section_names)} sections available. "
                "Use query_socwatch_data(parent_folder, sections=[...]) to view section content."
            ),
        }

    # ── Cache check ──────────────────────────────────────────────────────────────
    # Fast path: the output markdown existing with content is sufficient proof that
    # parsing already completed — no manifest required.  A full re-scan only
    # happens when force_reparse=True.
    if not force_reparse and os.path.exists(cached_md) and os.path.getsize(cached_md) > 0:
        return _build_compact_metadata(files_parsed=0, cached=True)
    # ── End cache check ──────────────────────────────────────────────────────────

    # Ensure output folder exists
    os.makedirs(output_folder, exist_ok=True)
    output_excel_path = str(output_dir / "socwatch_output_summary.xlsx")

    # ── Internal file discovery using filename regex + content scan ──────────────
    _sw_re = re.compile(r'soc.*watch', re.IGNORECASE)
    _excl_re = re.compile(
        r'(WakeupAnalysis|wakeup_analysis|_analysis|_summary_report|_processed|_output)',
        re.IGNORECASE
    )
    _SOCWATCH_CONTENT_MARKERS = [
        "Intel(R) SoC Watch",
        "SoC Watch for Windows",
        "Package C-State Summary",
        "Core C-State Summary",
        "CPU P-State/Frequency Summary",
        "C-State Residency",
        "SoCWatch",
    ]

    def _is_socwatch_by_content_parse(csv_path: Path) -> bool:
        """Return True if the CSV file contains ≥3 SocWatch content markers in its first 300 lines."""
        try:
            with open(csv_path, 'r', encoding='utf-8', errors='ignore') as fh:
                head = ''.join(fh.readline() for _ in range(300))
            return sum(1 for m in _SOCWATCH_CONTENT_MARKERS if m in head) >= 3
        except Exception:
            return False

    csv_files: List[str] = []
    _seen: set = set()
    for _f in parent_path.glob("**/*.csv"):
        if any(p.name.lower().startswith("socwatch_output") for p in _f.parents):
            continue
        if _excl_re.search(_f.name):
            continue
        _fp = str(_f)
        if _fp in _seen:
            continue
        # Stage 1: filename regex match
        if _sw_re.search(_f.name):
            csv_files.append(_fp)
            _seen.add(_fp)
        # Stage 2: content scan for files without "socwatch" in the name
        elif _is_socwatch_by_content_parse(_f):
            csv_files.append(_fp)
            _seen.add(_fp)

    if not csv_files:
        return {
            "success": False,
            "error": f"No SocWatch CSV files found under {parent_folder}.",
            "hint": "Run find_socwatch_files(parent_folder) first to confirm SocWatch files exist.",
        }

    safe_print(f"Found {len(csv_files)} CSV files:")
    for file in csv_files:
        safe_print(f"  - {os.path.basename(file)}")

    # ── Parallel CSV parsing ─────────────────────────────────────────────────────
    # csv_files only contains confirmed SocWatch files (by filename OR content scan),
    # so always route to parse_socwatch_csv regardless of filename.
    def _parse_one_csv(csv_file: str):
        """Parse a single CSV and return (file_key, parsed_data) or (file_key, Exception)."""
        file_key = os.path.splitext(os.path.basename(csv_file))[0]
        try:
            parsed_data = parse_socwatch_csv(csv_file, json_output_dir=output_folder)
            return file_key, parsed_data, None
        except Exception as e:
            return file_key, None, f"{os.path.basename(csv_file)}: {type(e).__name__}: {str(e)}"

    max_workers = min(len(csv_files), 8)
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        parse_results = list(pool.map(_parse_one_csv, csv_files))

    all_parsed_data: Dict[str, Any] = {}
    failed_files: List[str] = []

    for file_key, parsed_data, error in parse_results:
        if error:
            safe_print(f"  FAIL {error}")
            failed_files.append(error)
        else:
            safe_print(f"  OK {file_key}")
            all_parsed_data[file_key] = parsed_data
    # ── End parallel parsing ─────────────────────────────────────────────────────

    if not all_parsed_data:
        error_summary = "; ".join(failed_files)
        return {
            "success": False,
            "error": f"No files were successfully parsed. Errors: {error_summary}",
            "failed_files": failed_files,
        }

    # Create Excel comparison file (markdown is created automatically inside)
    safe_print(f"\nCreating Excel comparison file: {os.path.basename(output_excel_path)}")
    create_excel_comparison(all_parsed_data, output_excel_path)

    safe_print(f"\n{'='*60}")
    safe_print(f"PROCESSING SUMMARY")
    safe_print(f"{'='*60}")
    safe_print(f"Total files found: {len(csv_files)}")
    safe_print(f"Successfully parsed: {len(all_parsed_data)}")
    safe_print(f"Failed to parse: {len(failed_files)}")

    if failed_files:
        safe_print(f"\nFailed files:")
        for detail in failed_files:
            safe_print(f"  - {detail}")

    safe_print(f"\nOutput files created:")
    safe_print(f"  - Excel comparison: {output_excel_path}")
    for fk in all_parsed_data.keys():
        safe_print(f"  - JSON: {fk}.json")

    # Write manifest so future cache checks know which files this parse covered.
    try:
        import json as _json_w
        _manifest_out = output_dir / "socwatch_parse_manifest.json"
        with open(_manifest_out, "w", encoding="utf-8") as _mfw:
            _json_w.dump({"parsed_files": sorted(csv_files)}, _mfw, indent=2)
    except Exception:
        pass

    return _build_compact_metadata(files_parsed=len(all_parsed_data), cached=False)


# =====================================================================
# MCP TOOL — Query SocWatch sections (filtered, low token cost)
# =====================================================================
@mcp.tool(
    description=(
        "Query the compiled SocWatch markdown summary and return specific sections. "
        "Use this for ALL user-facing content after parse_socwatch_data has run. "
        "IMPORTANT: maximum 4 sections per call to keep LLM context small and fast. "
        "If you need more sections, call this tool multiple times with up to 4 sections each. "
        "Returns only the requested section(s) as compact text. "
        "Never return summary_table from parse_socwatch_data directly — use this tool."
    ),
    tags={"socwatch", "query", "filter", "sections"},
)
@async_tool
def query_socwatch_data(
    parent_folder: str,
    sections: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Query the compiled SocWatch markdown for specific sections. Low token cost."""
    _MAX_SECTIONS_PER_CALL = 4
    _MAX_SECTION_CHARS = 1500  # per-section content hard cap
    _PRIORITY_SECTIONS = [
        "Package C-State", "Core C-State", "CPU P-State", "MEMSS P-State", "Thread Wakeups"
    ]
    md_path = Path(parent_folder) / "Analysis" / "socwatch_output" / "socwatch_output_summary.md"
    if not md_path.exists():
        return {
            "success": False,
            "error": "socwatch_output_summary.md not found. Call parse_socwatch_data(parent_folder) first.",
        }
    try:
        md_text = md_path.read_text(encoding="utf-8")
    except Exception as e:
        return {"success": False, "error": f"Failed to read markdown: {e}"}

    sections_map = _extract_sections_from_md(md_text)
    all_section_names = list(sections_map.keys())

    if not all_section_names:
        return {
            "success": False,
            "error": (
                "The markdown file was found but contains 0 parseable sections. "
                "This usually means the file was just created but is empty or malformed. "
                "Try calling parse_socwatch_data(parent_folder, force_reparse=True) "
                "to regenerate it."
            ),
            "markdown_path": str(md_path),
            "file_size_bytes": md_path.stat().st_size,
        }

    remaining_sections: List[str] = []
    if sections:
        secs_lower = [s.lower() for s in sections]
        matched = {
            k: v for k, v in sections_map.items()
            if any(s in k.lower() for s in secs_lower)
        }
        if not matched:
            return {
                "success": False,
                "error": f"No sections matched {sections}.",
                "all_sections": all_section_names,
            }
        # Enforce per-call section cap — return first N, report the rest
        if len(matched) > _MAX_SECTIONS_PER_CALL:
            matched_items = list(matched.items())
            remaining_sections = [k for k, _ in matched_items[_MAX_SECTIONS_PER_CALL:]]
            matched = dict(matched_items[:_MAX_SECTIONS_PER_CALL])
    else:
        # Default: 5 priority sections (capped to 4)
        p_lower = [p.lower() for p in _PRIORITY_SECTIONS]
        matched = {
            k: v for k, v in sections_map.items()
            if any(p in k.lower() for p in p_lower)
        }
        if not matched:
            # Fallback: return first 3 sections
            matched = dict(list(sections_map.items())[:3])
        if len(matched) > _MAX_SECTIONS_PER_CALL:
            matched_items = list(matched.items())
            remaining_sections = [k for k, _ in matched_items[_MAX_SECTIONS_PER_CALL:]]
            matched = dict(matched_items[:_MAX_SECTIONS_PER_CALL])

    # Build content with explicit section name headers so agent context is clear
    # Apply per-section content cap to prevent context window bloat
    content_parts = []
    for sec_name, sec_content in matched.items():
        if len(sec_content) > _MAX_SECTION_CHARS:
            sec_content = sec_content[:_MAX_SECTION_CHARS] + "\n... [truncated — section too long]"
        content_parts.append(f"## {sec_name}\n{sec_content}")
    content = "\n\n".join(content_parts)

    result: Dict[str, Any] = {
        "success": True,
        "content": content,
        "sections_shown": list(matched.keys()),
    }
    if remaining_sections:
        result["remaining_sections"] = remaining_sections
        result["hint"] = (
            f"Only {_MAX_SECTIONS_PER_CALL} sections returned to limit context size. "
            f"Call query_socwatch_data again with sections={remaining_sections} for the rest."
        )
    return result
